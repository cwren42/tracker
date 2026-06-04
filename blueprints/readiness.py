import csv
import io
import os
import zipfile
from datetime import datetime

from flask import Blueprint, flash, redirect, render_template, request, send_file, url_for
from flask_login import login_required, current_user
from openpyxl import load_workbook
from sqlalchemy import or_

from extensions import db
from models import (
    Asset,
    Employee,
    SOC2InternalAudit,
    SOC2InternalAuditFinding,
    SOC2ManagementReview,
    SOC2ManagementReviewAction,
    SOC2PhishingCampaign,
    SOC2PhishingResult,
    SOC2PolicyAcknowledgement,
    SOC2ReadinessItem,
    SOC2SecurityTrainingRecord,
    SOC2ReadinessUpdate,
    SOC2Vendor,
    SOC2VendorReview,
    SystemDescription,
    _log_audit,
)
from soc2_models import SOC2Control
from soc2_artifact_service import build_system_description_markdown, get_training_completion_snapshot
from utils import admin_required


bp = Blueprint('readiness', __name__)


# Auto-collectable evidence, reconciled to the current 93-item StrikeGraphEvidence
# catalog (rebuild commits 5a4c9e7/34aac23/1eda2ba). Each entry's evidence_name
# is a real catalog row; the generator is wired in
# EvidenceFileService.generate_evidence_file_by_name. The 'category' selects the
# on-disk cache directory used to detect previously generated artifacts.
AUTOMATED_EVIDENCE_EXPORTS = [
    # M365 / Intune (identity, devices)
    {'evidence_name': 'Administrator Access to Application', 'zip_path': 'automated_evidence/admin_access_application.xlsx', 'category': 'M365'},
    {'evidence_name': 'Administrator Access to Database', 'zip_path': 'automated_evidence/admin_access_database.xlsx', 'category': 'M365'},
    {'evidence_name': 'Administrator Access to Network/Cloud', 'zip_path': 'automated_evidence/admin_access_network_cloud.xlsx', 'category': 'M365'},
    {'evidence_name': 'Administrator Access to Operating System', 'zip_path': 'automated_evidence/admin_access_os.xlsx', 'category': 'M365'},
    {'evidence_name': 'Application User List', 'zip_path': 'automated_evidence/application_user_list.xlsx', 'category': 'M365'},
    {'evidence_name': 'Database User List', 'zip_path': 'automated_evidence/database_user_list.xlsx', 'category': 'M365'},
    {'evidence_name': 'Network/Cloud User List', 'zip_path': 'automated_evidence/network_cloud_user_list.xlsx', 'category': 'M365'},
    {'evidence_name': 'Operating System User List', 'zip_path': 'automated_evidence/os_user_list.xlsx', 'category': 'M365'},
    {'evidence_name': 'Asset Inventory', 'zip_path': 'automated_evidence/asset_inventory.xlsx', 'category': 'M365'},
    {'evidence_name': 'Device Disk Encryption', 'zip_path': 'automated_evidence/device_disk_encryption.xlsx', 'category': 'M365'},
    # RMM-backed (endpoint protection, vulnerability, patch)
    {'evidence_name': 'Antivirus Configuration - Workstation', 'zip_path': 'automated_evidence/antivirus_workstation.xlsx', 'category': 'RMM'},
    {'evidence_name': 'Antivirus Configuration - Server', 'zip_path': 'automated_evidence/antivirus_server.xlsx', 'category': 'RMM'},
    {'evidence_name': 'Vulnerability Scan Results', 'zip_path': 'automated_evidence/vulnerability_scan_results.xlsx', 'category': 'RMM'},
    {'evidence_name': 'Vulnerability Remediation', 'zip_path': 'automated_evidence/vulnerability_remediation.xlsx', 'category': 'RMM'},
    {'evidence_name': 'Patch Scan', 'zip_path': 'automated_evidence/patch_scan.xlsx', 'category': 'RMM'},
    {'evidence_name': 'Server Scan and Patch', 'zip_path': 'automated_evidence/server_scan_and_patch.xlsx', 'category': 'RMM'},
    # ISMS-manual policy evidence (Policy-type catalog rows auto-generated as
    # PDFs by extracting the relevant IS-section(s) from the published ISMS
    # Manual; see EvidenceFileService.generate_isms_section_pdf). "Code of
    # Conduct" is excluded (it lives in the Employee Handbook). Acceptable Use
    # Policy, Business Continuity Plan, and Vendor Management Policy and
    # Procedures were reverted to manual/HR-sourced evidence and are excluded
    # here (see POLICY_EVIDENCE_REVERTED_TO_MANUAL in evidence_file_service.py).
    {'evidence_name': 'Access Removal Procedures/Checklist', 'zip_path': 'automated_evidence/access_removal_procedures.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Backup Policy', 'zip_path': 'automated_evidence/backup_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Backup Restoration Procedures', 'zip_path': 'automated_evidence/backup_restoration_procedures.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Change Management Policy', 'zip_path': 'automated_evidence/change_management_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Data Classification Policy', 'zip_path': 'automated_evidence/data_classification_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Data Management Policy', 'zip_path': 'automated_evidence/data_management_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Incident Response Plan', 'zip_path': 'automated_evidence/incident_response_plan.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Information Security Policy', 'zip_path': 'automated_evidence/information_security_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Logical Access Policy and Procedures', 'zip_path': 'automated_evidence/logical_access_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Password Policy', 'zip_path': 'automated_evidence/password_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Patch Management Policy', 'zip_path': 'automated_evidence/patch_management_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Record Retention Schedule', 'zip_path': 'automated_evidence/record_retention_schedule.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Risk Management Policy and Procedures', 'zip_path': 'automated_evidence/risk_management_policy.pdf', 'category': 'ISMS'},
    {'evidence_name': 'System Description Document', 'zip_path': 'automated_evidence/system_description_document.pdf', 'category': 'ISMS'},
    {'evidence_name': 'Vulnerability Management Policy', 'zip_path': 'automated_evidence/vulnerability_management_policy.pdf', 'category': 'ISMS'},
]


EMPTY_AUTOMATED_EVIDENCE_DETAILS = {}

LIVE_REFRESH_EXCLUDED = {}

AUTOMATED_EVIDENCE_CACHE_DIRS = {
    'M365': '/var/www/tracker/static/evidence/m365',
    'Intune': '/var/www/tracker/static/evidence/m365',
    'Defender': '/var/www/tracker/static/evidence/M365/Defender',
    'RMM': '/var/www/tracker/static/evidence/rmm',
    'ISMS': '/var/www/tracker/static/evidence/isms',
}


STATUS_OPTIONS = [
    'Not In Place',
    'Partially In Place',
    'In Place',
    'Open',
    'Blocked',
    'Closed',
]


def _summary_counts(items):
    total = len(items)
    return {
        'total': total,
        'open': sum(1 for item in items if item.status in {'Not In Place', 'Partially In Place', 'Open', 'Blocked'}),
        'closed': sum(1 for item in items if item.status in {'In Place', 'Closed'}),
        'critical': sum(1 for item in items if item.priority == 'P1-Critical'),
        'manual': sum(1 for item in items if item.source_type == 'manual'),
    }


class _ControlReadinessRow:
    """Read-only readiness row projected directly from a SOC2Control.

    Control implementation status lives in the single-source catalog
    (soc2_control). The readiness view derives the ``type1_control`` rows
    from it so the status shown here always matches the SOC 2 dashboard and
    can never drift into a separate editable copy. ``control_id`` links the
    row to its evidence page on the SOC 2 dashboard.
    """

    source_type = 'type1_control'
    is_control = True
    owner = None

    def __init__(self, control):
        self.id = None
        self.control_id = control.id
        self.title = control.control_name
        self.domain = control.control_frequency
        self.audit_alignment = control.audit_alignment
        self.status = control.control_progress or 'Not In Place'
        self.owner = control.control_owner
        self.frequency = control.control_frequency
        # Controls do not carry a separate priority; surface open work as P2.
        self.priority = 'P1-Critical' if self.status == 'Not In Place' else 'P2-High'
        self.next_step = None


def _control_readiness_rows(search_term=None, status=None):
    query = SOC2Control.query.filter_by(is_active=True)
    if status:
        query = query.filter(SOC2Control.control_progress == status)
    if search_term:
        ilike_value = f'%{search_term}%'
        query = query.filter(or_(
            SOC2Control.control_name.ilike(ilike_value),
            SOC2Control.audit_alignment.ilike(ilike_value),
            SOC2Control.control_owner.ilike(ilike_value),
            SOC2Control.control_frequency.ilike(ilike_value),
        ))
    controls = query.order_by(SOC2Control.control_name.asc()).all()
    return [_ControlReadinessRow(control) for control in controls]


def _write_csv_to_zip(archive, filename, header, rows):
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    archive.writestr(filename, output.getvalue())


def _detail_for_empty_automated_evidence(evidence_name):
    return EMPTY_AUTOMATED_EVIDENCE_DETAILS.get(evidence_name, ('attention', 'Generated file contains no data rows'))


def _sanitize_evidence_name(evidence_name):
    safe_name = ''.join(character for character in evidence_name if character.isalnum() or character in (' ', '-', '_')).rstrip()
    return safe_name.replace(' ', '_')


def _inspect_evidence_file(file_path, evidence_name):
    status = 'exported'
    detail = ''
    if file_path.endswith('.xlsx'):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        if worksheet.max_row <= 1:
            status, detail = _detail_for_empty_automated_evidence(evidence_name)
        workbook.close()
    return status, detail


def _find_cached_evidence_file(item):
    cache_dir = AUTOMATED_EVIDENCE_CACHE_DIRS.get(item['category'])
    if not cache_dir or not os.path.isdir(cache_dir):
        return None

    prefix = f"{_sanitize_evidence_name(item['evidence_name'])}_"
    # ISMS policy evidence is generated as PDF; the live-data exports are .xlsx.
    allowed_ext = ('.pdf',) if item.get('category') == 'ISMS' else ('.xlsx',)
    candidates = [
        os.path.join(cache_dir, name)
        for name in os.listdir(cache_dir)
        if name.startswith(prefix) and name.endswith(allowed_ext)
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def _evidence_static_url(file_path):
    static_prefix = '/var/www/tracker/static/'
    if file_path and file_path.startswith(static_prefix):
        return '/static/' + file_path[len(static_prefix):]
    return None


def _attach_artifact_links(results):
    """Annotate automation-status result rows with a download URL/name for the
    most recent cached artifact, so the UI can offer a download control."""
    export_by_name = {item['evidence_name']: item for item in AUTOMATED_EVIDENCE_EXPORTS}
    for row in results:
        export_item = export_by_name.get(row['evidence_name'])
        artifact_url = None
        artifact_name = None
        if export_item:
            cached = _find_cached_evidence_file(export_item)
            if cached and os.path.exists(cached):
                artifact_url = _evidence_static_url(cached)
                artifact_name = os.path.basename(cached)
        row['artifact_url'] = artifact_url
        row['artifact_name'] = artifact_name
    return results


def _summarize_automated_evidence_results(results):
    return {
        'total': len(results),
        'exported': sum(1 for item in results if item['status'] == 'exported'),
        'follow_up': sum(1 for item in results if item['status'] != 'exported'),
        'errors': sum(1 for item in results if item['status'] == 'error'),
        'no_resources': sum(1 for item in results if item['status'] == 'no-resources'),
        'attention': sum(1 for item in results if item['status'] == 'attention'),
    }


def _collect_automated_evidence_exports(archive=None):
    from evidence_file_service import EvidenceFileService

    evidence_service = EvidenceFileService()
    results = []

    for item in AUTOMATED_EVIDENCE_EXPORTS:
        try:
            result = evidence_service.generate_and_record(item['evidence_name'])
            file_path = result.get('file_path')
            if file_path and os.path.exists(file_path):
                if archive is not None:
                    archive.write(file_path, item['zip_path'])
                status, detail = _inspect_evidence_file(file_path, item['evidence_name'])
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': status,
                    'detail': detail,
                })
            else:
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': 'missing',
                    'detail': 'Generator returned no file path',
                })
        except Exception as exc:
            results.append({
                'category': item['category'],
                'evidence_name': item['evidence_name'],
                'zip_path': item['zip_path'],
                'status': 'error',
                'detail': str(exc),
            })

    return results


def _collect_cached_automated_evidence_results():
    results = []
    latest_generated_at = None

    for item in AUTOMATED_EVIDENCE_EXPORTS:
        try:
            file_path = _find_cached_evidence_file(item)
            if file_path and os.path.exists(file_path):
                status, detail = _inspect_evidence_file(file_path, item['evidence_name'])
                generated_at = datetime.utcfromtimestamp(os.path.getmtime(file_path))
                if latest_generated_at is None or generated_at > latest_generated_at:
                    latest_generated_at = generated_at
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': status,
                    'detail': detail,
                })
            else:
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': 'missing',
                    'detail': 'No cached evidence file yet. Export the Type 1 pack or run a live refresh.',
                })
        except Exception as exc:
            results.append({
                'category': item['category'],
                'evidence_name': item['evidence_name'],
                'zip_path': item['zip_path'],
                'status': 'error',
                'detail': str(exc),
            })

    return results, latest_generated_at


def _get_cached_result_for_item(item):
    file_path = _find_cached_evidence_file(item)
    if file_path and os.path.exists(file_path):
        status, detail = _inspect_evidence_file(file_path, item['evidence_name'])
        generated_at = datetime.utcfromtimestamp(os.path.getmtime(file_path))
        return {
            'category': item['category'],
            'evidence_name': item['evidence_name'],
            'zip_path': item['zip_path'],
            'status': status,
            'detail': detail,
        }, generated_at

    return {
        'category': item['category'],
        'evidence_name': item['evidence_name'],
        'zip_path': item['zip_path'],
        'status': 'missing',
        'detail': 'No cached evidence file yet. Export the Type 1 pack to generate one.',
    }, None


def _collect_interactive_live_automation_results():
    from evidence_file_service import EvidenceFileService

    evidence_service = EvidenceFileService()
    results = []
    latest_generated_at = None

    for item in AUTOMATED_EVIDENCE_EXPORTS:
        if item['evidence_name'] in LIVE_REFRESH_EXCLUDED:
            cached_result, cached_generated_at = _get_cached_result_for_item(item)
            cached_result['detail'] = LIVE_REFRESH_EXCLUDED[item['evidence_name']]
            results.append(cached_result)
            if cached_generated_at and (latest_generated_at is None or cached_generated_at > latest_generated_at):
                latest_generated_at = cached_generated_at
            continue

        try:
            result = evidence_service.generate_and_record(item['evidence_name'])
            file_path = result.get('file_path')
            if file_path and os.path.exists(file_path):
                status, detail = _inspect_evidence_file(file_path, item['evidence_name'])
                generated_at = datetime.utcfromtimestamp(os.path.getmtime(file_path))
                if latest_generated_at is None or generated_at > latest_generated_at:
                    latest_generated_at = generated_at
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': status,
                    'detail': detail,
                })
            else:
                results.append({
                    'category': item['category'],
                    'evidence_name': item['evidence_name'],
                    'zip_path': item['zip_path'],
                    'status': 'missing',
                    'detail': 'Generator returned no file path',
                })
        except Exception as exc:
            results.append({
                'category': item['category'],
                'evidence_name': item['evidence_name'],
                'zip_path': item['zip_path'],
                'status': 'error',
                'detail': str(exc),
            })

    return results, latest_generated_at


def _build_type1_manifest(readiness_items, audits, findings, assets, employees, vendors, vendor_reviews, management_reviews, management_review_actions, training_records, system_description_sections, policy_acknowledgements, phishing_campaigns, phishing_results, automated_evidence_results):
    generated_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
    open_items = sum(1 for item in readiness_items if item.status in {'Not In Place', 'Partially In Place', 'Open', 'Blocked'})
    closed_items = sum(1 for item in readiness_items if item.status in {'In Place', 'Closed'})
    automated_exported = sum(1 for item in automated_evidence_results if item['status'] == 'exported')
    automated_missing = sum(1 for item in automated_evidence_results if item['status'] != 'exported')
    return "\n".join([
        '# SOC 2 Type 1 Evidence Pack',
        '',
        f'- Generated at: {generated_at}',
        '- Audit scope: Security and Confidentiality',
        '- Audit window: 2026-04-01 through 2026-06-30',
        '',
        '## Included exports',
        '',
        f'- Active readiness items: {len(readiness_items)}',
        f'- Open readiness items: {open_items}',
        f'- Closed or in-place readiness items: {closed_items}',
        f'- Internal audits: {len(audits)}',
        f'- Internal audit findings: {len(findings)}',
        f'- Assets exported: {len(assets)}',
        f'- Employees exported: {len(employees)}',
        f'- Vendors exported: {len(vendors)}',
        f'- Vendor reviews exported: {len(vendor_reviews)}',
        f'- Management reviews exported: {len(management_reviews)}',
        f'- Management review actions exported: {len(management_review_actions)}',
        f'- Security training records exported: {len(training_records)}',
        f'- System description sections exported: {len(system_description_sections)}',
        f'- Policy acknowledgements exported: {len(policy_acknowledgements)}',
        f'- Phishing campaigns exported: {len(phishing_campaigns)}',
        f'- Phishing results exported: {len(phishing_results)}',
        f'- Automated evidence files exported: {automated_exported}',
        f'- Automated evidence files needing follow-up: {automated_missing}',
        '',
        '## Pack contents',
        '',
        '- manifest.md',
        '- readiness_items.csv',
        '- internal_audits.csv',
        '- internal_audit_findings.csv',
        '- asset_inventory.csv',
        '- employee_directory.csv',
        '- vendor_register.csv',
        '- vendor_reviews.csv',
        '- management_reviews.csv',
        '- management_review_actions.csv',
        '- security_training_records.csv',
        '- policy_acknowledgements.csv',
        '- phishing_campaigns.csv',
        '- phishing_results.csv',
        '- automated_evidence_status.csv',
        '- automated_evidence_gaps.csv',
        '- system_description.md',
        '- automated_evidence/*',
        '',
        '## Notes',
        '',
        '- This pack is a point-in-time export from Tracker.',
        '- Automated evidence status and gaps are listed in the automated_evidence CSV files.',
        '- Source of truth for policy and control narrative remains the ISMS Manual.',
        '- Open readiness items identify evidence still to be completed for the audit package.',
    ])


@bp.route('/soc2/readiness/automation-status')
@login_required
@admin_required
def automation_status_dashboard():
    refresh_live = (request.args.get('refresh') or '').strip() == '1'
    if refresh_live:
        automated_evidence_results, latest_generated_at = _collect_interactive_live_automation_results()
        generated_at = latest_generated_at.strftime('%Y-%m-%d %H:%M:%S UTC') if latest_generated_at else datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        source_label = 'live refresh (interactive subset)'
    else:
        automated_evidence_results, cached_generated_at = _collect_cached_automated_evidence_results()
        generated_at = cached_generated_at.strftime('%Y-%m-%d %H:%M:%S UTC') if cached_generated_at else 'No cached run yet'
        source_label = 'cached files'
    _attach_artifact_links(automated_evidence_results)
    automated_evidence_results.sort(key=lambda item: (item['status'] == 'exported', item['category'], item['evidence_name']))
    return render_template(
        'soc2_readiness_automation_status.html',
        results=automated_evidence_results,
        summary=_summarize_automated_evidence_results(automated_evidence_results),
        generated_at=generated_at,
        source_label=source_label,
        refresh_live=refresh_live,
    )


@bp.route('/soc2/readiness/automation-status/generate', methods=['POST'])
@login_required
@admin_required
def generate_automated_evidence_item():
    """Generate a single auto-collectable evidence artifact on demand and stamp
    it back to the catalog (StrikeGraphEvidence + EvidenceSnapshot audit row)."""
    from evidence_file_service import EvidenceFileService

    evidence_name = (request.form.get('evidence_name') or '').strip()
    return_control_id = (request.form.get('control_id') or '').strip()

    def _back():
        if return_control_id.isdigit():
            return redirect(url_for('soc2.soc2_evidence', control_id=int(return_control_id)))
        return redirect(url_for('readiness.automation_status_dashboard'))

    valid_names = {item['evidence_name'] for item in AUTOMATED_EVIDENCE_EXPORTS}
    if evidence_name not in valid_names:
        flash('Unknown or non-automatable evidence item.', 'danger')
        return _back()

    collected_by = getattr(current_user, 'email', None) or getattr(current_user, 'username', 'admin')
    try:
        result = EvidenceFileService().generate_and_record(evidence_name, collected_by=collected_by)
    except Exception as exc:  # pragma: no cover - surfaced to the operator
        db.session.rollback()
        flash(f'Failed to generate "{evidence_name}": {exc}', 'danger')
        return _back()

    if result.get('file_path'):
        count = result.get('record_count')
        count_text = f' ({count} rows)' if count is not None else ''
        stamped = 'and recorded to the catalog' if result.get('stamped') else 'but the catalog row was not found'
        flash(f'Generated "{evidence_name}"{count_text} {stamped}.', 'success')
    else:
        flash(f'Generator for "{evidence_name}" returned no file (no source data or missing integration).', 'warning')
    return _back()


@bp.route('/soc2/readiness')
@login_required
@admin_required
def readiness_dashboard():
    status = (request.args.get('status') or '').strip()
    priority = (request.args.get('priority') or '').strip()
    source_type = (request.args.get('source_type') or 'type1_control').strip()
    search_term = (request.args.get('q') or '').strip()

    # Control implementation status is owned by the single-source catalog
    # (soc2_control). The "Current controls" (type1_control) view is derived
    # straight from it so it always matches the SOC 2 dashboard. The genuinely
    # separate workstreams (Type 1 pack tasks, manual follow-up) remain
    # editable rows in soc2_readiness_item.
    show_controls = source_type in ('type1_control', 'all')
    show_workstream = source_type != 'type1_control'

    items = []

    if show_controls:
        control_rows = _control_readiness_rows(search_term=search_term or None, status=status or None)
        if priority:
            control_rows = [row for row in control_rows if row.priority == priority]
        items.extend(control_rows)

    if show_workstream:
        query = SOC2ReadinessItem.query.filter_by(is_active=True)
        # type1_control rows are now derived from the catalog above; never
        # surface the (deactivated) legacy duplicates from this table.
        query = query.filter(SOC2ReadinessItem.source_type != 'type1_control')
        if source_type and source_type != 'all':
            query = query.filter(SOC2ReadinessItem.source_type == source_type)
        if status:
            query = query.filter(SOC2ReadinessItem.status == status)
        if priority:
            query = query.filter(SOC2ReadinessItem.priority == priority)
        if search_term:
            ilike_value = f'%{search_term}%'
            query = query.filter(or_(
                SOC2ReadinessItem.title.ilike(ilike_value),
                SOC2ReadinessItem.domain.ilike(ilike_value),
                SOC2ReadinessItem.audit_alignment.ilike(ilike_value),
                SOC2ReadinessItem.owner.ilike(ilike_value),
            ))
        items.extend(query.order_by(SOC2ReadinessItem.priority.asc(), SOC2ReadinessItem.title.asc()).all())

    return render_template(
        'soc2_readiness_dashboard.html',
        items=items,
        summary=_summary_counts(items),
        status_options=STATUS_OPTIONS,
        selected_status=status,
        selected_priority=priority,
        selected_source_type=source_type,
        search_term=search_term,
        priority_options=['P1-Critical', 'P2-High', 'P3-Validate'],
        source_options=['type1_control', 'soc2_type1', 'manual', 'all'],
    )


@bp.route('/soc2/readiness/export/type1-pack')
@login_required
@admin_required
def export_type1_pack():
    # Control rows are derived from the single-source catalog; workstream rows
    # (Type 1 pack tasks, manual follow-up) come from soc2_readiness_item.
    control_rows = _control_readiness_rows()
    workstream_rows = (
        SOC2ReadinessItem.query
        .filter(SOC2ReadinessItem.is_active.is_(True), SOC2ReadinessItem.source_type != 'type1_control')
        .order_by(SOC2ReadinessItem.priority.asc(), SOC2ReadinessItem.title.asc())
        .all()
    )
    readiness_items = control_rows + workstream_rows
    audits = SOC2InternalAudit.query.order_by(SOC2InternalAudit.planned_date.asc().nullslast(), SOC2InternalAudit.id.asc()).all()
    findings = SOC2InternalAuditFinding.query.order_by(SOC2InternalAuditFinding.audit_id.asc(), SOC2InternalAuditFinding.id.asc()).all()
    assets = Asset.query.order_by(Asset.asset_tag.asc()).all()
    employees = Employee.query.order_by(Employee.name.asc()).all()
    vendors = SOC2Vendor.query.filter_by(is_active=True).order_by(SOC2Vendor.vendor_name.asc()).all()
    vendor_reviews = SOC2VendorReview.query.order_by(SOC2VendorReview.review_date.desc(), SOC2VendorReview.id.desc()).all()
    management_reviews = SOC2ManagementReview.query.order_by(SOC2ManagementReview.review_date.desc(), SOC2ManagementReview.id.desc()).all()
    management_review_actions = SOC2ManagementReviewAction.query.order_by(SOC2ManagementReviewAction.due_date.asc().nullslast(), SOC2ManagementReviewAction.id.asc()).all()
    training_records = SOC2SecurityTrainingRecord.query.order_by(SOC2SecurityTrainingRecord.training_date.desc(), SOC2SecurityTrainingRecord.id.desc()).all()
    system_description_sections = SystemDescription.query.order_by(SystemDescription.section_order.asc(), SystemDescription.id.asc()).all()
    policy_acknowledgements = SOC2PolicyAcknowledgement.query.order_by(SOC2PolicyAcknowledgement.acknowledged_on.desc(), SOC2PolicyAcknowledgement.id.desc()).all()
    phishing_campaigns = SOC2PhishingCampaign.query.order_by(SOC2PhishingCampaign.campaign_date.desc(), SOC2PhishingCampaign.id.desc()).all()
    phishing_results = SOC2PhishingResult.query.order_by(SOC2PhishingResult.training_completed_on.desc().nullslast(), SOC2PhishingResult.id.desc()).all()
    training_snapshot = get_training_completion_snapshot()

    pack_buffer = io.BytesIO()
    with zipfile.ZipFile(pack_buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        automated_evidence_results = _collect_automated_evidence_exports(archive)
        archive.writestr('manifest.md', _build_type1_manifest(readiness_items, audits, findings, assets, employees, vendors, vendor_reviews, management_reviews, management_review_actions, training_records, system_description_sections, policy_acknowledgements, phishing_campaigns, phishing_results, automated_evidence_results))

        _write_csv_to_zip(
            archive,
            'readiness_items.csv',
            ['Item Key', 'Title', 'Domain', 'Audit Alignment', 'Priority', 'Status', 'Owner', 'Frequency', 'Source', 'Evidence Reference', 'Next Step', 'Manual Reference'],
            [[
                getattr(item, 'item_key', None) or (f"control-{item.control_id}" if getattr(item, 'is_control', False) else ''),
                item.title,
                item.domain or '',
                item.audit_alignment or '',
                item.priority or '',
                item.status or '',
                item.owner or '',
                item.frequency or '',
                item.source_type or '',
                getattr(item, 'evidence_reference', None) or '',
                getattr(item, 'next_step', None) or '',
                getattr(item, 'manual_reference', None) or '',
            ] for item in readiness_items],
        )

        _write_csv_to_zip(
            archive,
            'internal_audits.csv',
            ['Audit Key', 'Title', 'Status', 'Owner', 'Audit Period Start', 'Audit Period End', 'Planned Date', 'Performed Date', 'Evidence Reference', 'Summary'],
            [[
                audit.audit_key,
                audit.title,
                audit.status or '',
                audit.owner or '',
                audit.audit_period_start.isoformat() if audit.audit_period_start else '',
                audit.audit_period_end.isoformat() if audit.audit_period_end else '',
                audit.planned_date.isoformat() if audit.planned_date else '',
                audit.performed_date.isoformat() if audit.performed_date else '',
                audit.evidence_reference or '',
                audit.summary or '',
            ] for audit in audits],
        )

        _write_csv_to_zip(
            archive,
            'internal_audit_findings.csv',
            ['Finding Key', 'Audit Key', 'Title', 'Severity', 'Status', 'Criteria Reference', 'Owner', 'Due Date', 'Linked Readiness Item', 'Evidence Reference', 'Recommendation'],
            [[
                finding.finding_key,
                finding.audit.audit_key if finding.audit else '',
                finding.title,
                finding.severity or '',
                finding.status or '',
                finding.criteria_reference or '',
                finding.owner or '',
                finding.due_date.isoformat() if finding.due_date else '',
                finding.linked_readiness_item.title if getattr(finding, 'linked_readiness_item', None) else '',
                finding.evidence_reference or '',
                finding.recommendation or '',
            ] for finding in findings],
        )

        _write_csv_to_zip(
            archive,
            'asset_inventory.csv',
            ['Asset Tag', 'Name', 'Category', 'Status', 'Manufacturer', 'Model', 'Serial Number', 'Assigned To', 'Location'],
            [[
                asset.asset_tag,
                asset.name,
                asset.category or '',
                asset.status or '',
                asset.manufacturer or '',
                asset.model or '',
                asset.serial_number or '',
                asset.assigned_employee.name if asset.assigned_employee else '',
                asset.location or '',
            ] for asset in assets],
        )

        _write_csv_to_zip(
            archive,
            'employee_directory.csv',
            ['Name', 'Email', 'Department', 'Position', 'Employment Status'],
            [[
                employee.name,
                employee.email or '',
                employee.department or '',
                employee.position or '',
                'Active' if employee.ad_enabled is True else 'Disabled' if employee.ad_enabled is False else 'Unknown',
            ] for employee in employees],
        )

        _write_csv_to_zip(
            archive,
            'vendor_register.csv',
            ['Vendor Key', 'Vendor Name', 'Service Description', 'Vendor Type', 'Criticality', 'Risk Level', 'Owner', 'Data Access Scope', 'Contract Status', 'Assurance Status', 'Last Review', 'Next Review', 'Evidence Reference'],
            [[
                vendor.vendor_key,
                vendor.vendor_name,
                vendor.service_description or '',
                vendor.vendor_type or '',
                vendor.criticality or '',
                vendor.risk_level or '',
                vendor.owner or '',
                vendor.data_access_scope or '',
                vendor.contract_status or '',
                vendor.assurance_status or '',
                vendor.last_review_date.isoformat() if vendor.last_review_date else '',
                vendor.next_review_date.isoformat() if vendor.next_review_date else '',
                vendor.evidence_reference or '',
            ] for vendor in vendors],
        )

        _write_csv_to_zip(
            archive,
            'vendor_reviews.csv',
            ['Vendor Name', 'Review Date', 'Review Type', 'Status', 'Reviewer', 'Summary', 'Findings', 'Evidence Reference'],
            [[
                review.vendor.vendor_name if review.vendor else '',
                review.review_date.isoformat() if review.review_date else '',
                review.review_type or '',
                review.status or '',
                review.reviewer or '',
                review.summary or '',
                review.findings or '',
                review.evidence_reference or '',
            ] for review in vendor_reviews],
        )

        _write_csv_to_zip(
            archive,
            'management_reviews.csv',
            ['Review Key', 'Title', 'Review Date', 'Period Start', 'Period End', 'Chairperson', 'Minute Taker', 'Location', 'Status', 'Attendees', 'Agenda Summary', 'Decisions Summary', 'Effectiveness Summary', 'Resource Summary', 'Evidence Reference'],
            [[
                review.review_key,
                review.title,
                review.review_date.isoformat() if review.review_date else '',
                review.review_period_start.isoformat() if review.review_period_start else '',
                review.review_period_end.isoformat() if review.review_period_end else '',
                review.chairperson or '',
                review.minute_taker or '',
                review.location or '',
                review.status or '',
                review.attendees or '',
                review.agenda_summary or '',
                review.decisions_summary or '',
                review.effectiveness_summary or '',
                review.resource_summary or '',
                review.evidence_reference or '',
            ] for review in management_reviews],
        )

        _write_csv_to_zip(
            archive,
            'management_review_actions.csv',
            ['Action Key', 'Review Key', 'Title', 'Owner', 'Due Date', 'Status', 'Notes'],
            [[
                action.action_key,
                action.review.review_key if action.review else '',
                action.title,
                action.owner or '',
                action.due_date.isoformat() if action.due_date else '',
                action.status or '',
                action.notes or '',
            ] for action in management_review_actions],
        )

        _write_csv_to_zip(
            archive,
            'security_training_records.csv',
            ['Record Key', 'Employee', 'Email', 'Department', 'Role', 'Training Date', 'Training Topic', 'Provider Method', 'Duration', 'Completion Status', 'Score', 'Notes'],
            [[
                record.record_key,
                record.trainee_name,
                record.trainee_email or '',
                record.department or '',
                record.role_title or '',
                record.training_date.isoformat() if record.training_date else '',
                record.training_topic,
                record.provider_method or '',
                record.duration or '',
                record.completion_status or '',
                record.score if record.score is not None else '',
                record.notes or '',
            ] for record in training_records],
        )

        _write_csv_to_zip(
            archive,
            'policy_acknowledgements.csv',
            ['Acknowledgement Key', 'Person', 'Email', 'Department', 'Type', 'Policy Name', 'Policy Version', 'Acknowledged On', 'Status', 'Evidence Reference', 'Notes'],
            [[
                record.acknowledgement_key,
                record.person_name,
                record.person_email or '',
                record.department or '',
                record.acknowledgement_type or '',
                record.policy_name,
                record.policy_version or '',
                record.acknowledged_on.isoformat() if record.acknowledged_on else '',
                record.status or '',
                record.evidence_reference or '',
                record.notes or '',
            ] for record in policy_acknowledgements],
        )

        _write_csv_to_zip(
            archive,
            'phishing_campaigns.csv',
            ['Campaign Key', 'Title', 'Campaign Date', 'Provider', 'Scope', 'Status', 'Follow-up Training Topic', 'Summary', 'Evidence Reference'],
            [[
                campaign.campaign_key,
                campaign.title,
                campaign.campaign_date.isoformat() if campaign.campaign_date else '',
                campaign.provider or '',
                campaign.scope or '',
                campaign.status or '',
                campaign.follow_up_training_topic or '',
                campaign.summary or '',
                campaign.evidence_reference or '',
            ] for campaign in phishing_campaigns],
        )

        _write_csv_to_zip(
            archive,
            'phishing_results.csv',
            ['Result Key', 'Campaign Key', 'Employee', 'Email', 'Department', 'Delivered', 'Opened', 'Clicked', 'Reported', 'Training Completed', 'Training Completed On', 'Outcome', 'Notes'],
            [[
                result.result_key,
                result.campaign.campaign_key if result.campaign else '',
                result.employee_name,
                result.employee_email or '',
                result.department or '',
                'Yes' if result.delivered else 'No',
                'Yes' if result.opened else 'No',
                'Yes' if result.clicked else 'No',
                'Yes' if result.reported else 'No',
                'Yes' if result.training_completed else 'No',
                result.training_completed_on.isoformat() if result.training_completed_on else '',
                result.outcome or '',
                result.notes or '',
            ] for result in phishing_results],
        )

        _write_csv_to_zip(
            archive,
            'automated_evidence_status.csv',
            ['Category', 'Evidence Name', 'Zip Path', 'Status', 'Detail'],
            [[
                result['category'],
                result['evidence_name'],
                result['zip_path'],
                result['status'],
                result['detail'],
            ] for result in automated_evidence_results],
        )

        _write_csv_to_zip(
            archive,
            'automated_evidence_gaps.csv',
            ['Category', 'Evidence Name', 'Expected Zip Path', 'Status', 'Detail'],
            [[
                result['category'],
                result['evidence_name'],
                result['zip_path'],
                result['status'],
                result['detail'],
            ] for result in automated_evidence_results if result['status'] != 'exported'],
        )

        archive.writestr('system_description.md', build_system_description_markdown())

    _log_audit('soc2_type1_evidence_pack', 0, 'export', {
        'readiness_items': len(readiness_items),
        'internal_audits': len(audits),
        'internal_audit_findings': len(findings),
        'assets': len(assets),
        'employees': len(employees),
        'vendors': len(vendors),
        'vendor_reviews': len(vendor_reviews),
        'management_reviews': len(management_reviews),
        'management_review_actions': len(management_review_actions),
        'security_training_records': len(training_records),
        'system_description_sections': len(system_description_sections),
        'policy_acknowledgements': len(policy_acknowledgements),
        'phishing_campaigns': len(phishing_campaigns),
        'phishing_results': len(phishing_results),
        'automated_evidence_exported': sum(1 for item in automated_evidence_results if item['status'] == 'exported'),
        'automated_evidence_gaps': sum(1 for item in automated_evidence_results if item['status'] != 'exported'),
        'training_completed_current_cycle': training_snapshot['completed_current_cycle'],
    })
    db.session.commit()

    pack_buffer.seek(0)
    filename = f"soc2-type1-evidence-pack-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}.zip"
    return send_file(pack_buffer, as_attachment=True, download_name=filename, mimetype='application/zip')


@bp.route('/soc2/readiness/<int:item_id>')
@login_required
@admin_required
def readiness_item_detail(item_id):
    item = SOC2ReadinessItem.query.get_or_404(item_id)
    return render_template(
        'soc2_readiness_item_detail.html',
        item=item,
        status_options=STATUS_OPTIONS,
    )


@bp.route('/soc2/readiness/<int:item_id>/update', methods=['POST'])
@login_required
@admin_required
def update_readiness_item(item_id):
    item = SOC2ReadinessItem.query.get_or_404(item_id)
    previous_status = item.status

    item.status = (request.form.get('status') or item.status).strip()
    item.owner = (request.form.get('owner') or '').strip() or None
    item.frequency = (request.form.get('frequency') or '').strip() or None
    item.evidence_reference = (request.form.get('evidence_reference') or '').strip() or None
    item.next_step = (request.form.get('next_step') or '').strip() or None
    item.notes = (request.form.get('notes') or '').strip() or None

    due_date = (request.form.get('due_date') or '').strip()
    item.due_date = datetime.strptime(due_date, '%Y-%m-%d').date() if due_date else None

    update_note = (request.form.get('update_note') or '').strip()
    if previous_status != item.status or update_note:
        db.session.add(SOC2ReadinessUpdate(
            readiness_item_id=item.id,
            update_type='status_change' if previous_status != item.status else 'note',
            previous_status=previous_status,
            new_status=item.status,
            note=update_note or 'Readiness item updated',
            created_by=getattr(current_user, 'email', None) or getattr(current_user, 'username', 'unknown'),
        ))

    _log_audit('soc2_readiness_item', item.id, 'update', {
        'previous_status': previous_status,
        'new_status': item.status,
        'owner': item.owner,
        'due_date': due_date,
    })
    db.session.commit()
    flash('Readiness item updated.', 'success')
    return redirect(url_for('readiness.readiness_item_detail', item_id=item.id))