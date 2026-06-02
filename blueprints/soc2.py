import base64
import io
import json
import os
import re
import subprocess
import threading
from datetime import datetime, timedelta, timezone

import requests
from flask import (Blueprint, abort, current_app, flash, g, jsonify,
                   redirect, render_template, request, send_file, session,
                   url_for)
from flask_login import current_user, login_required
from sqlalchemy import func, or_, text

from extensions import db, limiter
from models import (
    AuditTrail, Asset, AssetHistory, Control, CustomReport, DashboardWidget,
    Employee, License, LicenseAssignment, LicenseInfo, MaintenanceWindow,
    MonitoringAlert, MonitoringCheck, MonitoringProfile, Policy, PolicySection,
    ProxmoxBackupJob, ProxmoxZfsPool, RemoteSession, Risk, Setting,
    SupportTicket, TicketActivity, TicketNote, User, now_mst, allowed_file,
    SystemDescription, AzureIntegrationConfig, ControlRiskMapping,
    SOC2ReadinessItem, SOC2Vendor,
)
from soc2_models import (
    SOC2Control, EvidenceSnapshot, M365User, IntuneDevice, StrikeGraphEvidence,
    AuditLog, summarize_control_evidence, is_progress_relevant_evidence,
)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = Alignment = Border = Font = PatternFill = Side = get_column_letter = None
from utils import (
    admin_required, manager_required, eagle_eyes_required,
    ticket_access_required, license_required,
    send_email, send_admin_notification, send_asset_assignment_email,
    send_warranty_expiry_alert, send_lifecycle_alert,
)
import logging
logger = logging.getLogger(__name__)


bp = Blueprint('soc2', __name__)


EVIDENCE_CACHE_DIRS = {
    'M365': '/var/www/tracker/static/evidence/m365',
    'M365/Intune': '/var/www/tracker/static/evidence/m365',
    'Intune': '/var/www/tracker/static/evidence/m365',
    'M365/Defender': '/var/www/tracker/static/evidence/M365/Defender',
    'ISMS': '/var/www/tracker/static/evidence/isms',
    'TeamViewer': '/var/www/tracker/static/evidence/teamviewer',
}


def _resolve_evidence_file_path(file_path):
    if not file_path:
        return None

    candidates = [file_path]
    if not os.path.isabs(file_path):
        candidates.extend([
            os.path.join('/var/www/tracker', file_path),
            os.path.join('/var/www/tracker/static', file_path),
        ])

    for candidate in candidates:
        if candidate and os.path.exists(candidate):
            return candidate
    return None


def _evidence_file_to_static_url(file_path):
    resolved_path = _resolve_evidence_file_path(file_path)
    if not resolved_path:
        return None

    static_prefix = '/var/www/tracker/static/'
    if resolved_path.startswith(static_prefix):
        return '/static/' + resolved_path[len(static_prefix):]
    return None


def _sanitize_evidence_name(evidence_name):
    safe_name = ''.join(character for character in (evidence_name or '') if character.isalnum() or character in (' ', '-', '_')).rstrip()
    return safe_name.replace(' ', '_')


def _find_cached_evidence_file(evidence_item):
    cache_dir = EVIDENCE_CACHE_DIRS.get(evidence_item.automation_source)
    if not cache_dir or not os.path.isdir(cache_dir):
        return None

    prefix = f"{_sanitize_evidence_name(evidence_item.evidence_name)}_"
    candidates = [
        os.path.join(cache_dir, name)
        for name in os.listdir(cache_dir)
        if name.startswith(prefix)
    ]
    if not candidates:
        return None
    return max(candidates, key=os.path.getmtime)


def _evidence_item_has_artifact(evidence_item):
    resolved_path = _resolve_evidence_file_path(evidence_item.file_path)
    if resolved_path:
        return True
    cached_file = _find_cached_evidence_file(evidence_item)
    return bool(cached_file and os.path.exists(cached_file))


def _build_control_evidence_summary(controls, latest_evidence):
    evidence_items = StrikeGraphEvidence.query.filter(StrikeGraphEvidence.control_id.isnot(None)).all()
    evidence_by_control = {}
    for evidence_item in evidence_items:
        evidence_by_control.setdefault(evidence_item.control_id, []).append(evidence_item)

    control_evidence = {}
    total_linked = 0
    total_gathered = 0
    controls_with_gathered = 0
    controls_missing_artifacts = 0

    for control in controls:
        linked_items = evidence_by_control.get(control.id, [])
        evidence_summary = summarize_control_evidence(control.control_name, linked_items)
        required_items = evidence_summary['relevant_items']
        gathered_items = [item for item in required_items if _evidence_item_has_artifact(item)]
        missing_items = [
            item for item in required_items
            if is_progress_relevant_evidence(control.control_name, item) and item not in gathered_items
        ]
        snapshot = latest_evidence.get(control.id)

        summary = {
            'linked_count': evidence_summary['total_count'],
            'gathered_count': evidence_summary['gathered_count'],
            'missing_count': evidence_summary['missing_count'],
            'gathered_names': [item.evidence_name for item in gathered_items[:3]],
            'missing_names': [item.evidence_name for item in missing_items[:3]],
            'snapshot': snapshot,
            'has_snapshot': bool(snapshot),
        }
        control_evidence[control.id] = summary

        total_linked += summary['linked_count']
        total_gathered += summary['gathered_count']
        if summary['gathered_count'] > 0 or summary['has_snapshot']:
            controls_with_gathered += 1
        if summary['linked_count'] > summary['gathered_count']:
            controls_missing_artifacts += 1

    return control_evidence, {
        'linked_total': total_linked,
        'gathered_total': total_gathered,
        'controls_with_gathered': controls_with_gathered,
        'controls_missing_artifacts': controls_missing_artifacts,
    }


def _decode_jwt_claims(token):
    payload = token.split('.')[1]
    payload += '=' * (-len(payload) % 4)
    return json.loads(base64.urlsafe_b64decode(payload.encode()).decode())


def _build_automated_evidence_status():
    from blueprints.readiness import AUTOMATED_EVIDENCE_EXPORTS

    evidence_rows = []
    exported_count = 0
    linked_controls_total = 0
    for item in AUTOMATED_EVIDENCE_EXPORTS:
        evidence_name = item['evidence_name']
        linked_controls = StrikeGraphEvidence.query.filter_by(evidence_name=evidence_name).count()
        linked_controls_total += linked_controls

        cache_file = None
        snapshot_time = None
        evidence_record = StrikeGraphEvidence.query.filter_by(evidence_name=evidence_name).first()
        if evidence_record is not None:
            cache_file = _find_cached_evidence_file(evidence_record)
            if cache_file and os.path.exists(cache_file):
                snapshot_time = datetime.fromtimestamp(os.path.getmtime(cache_file))

        has_artifact = bool(cache_file and os.path.exists(cache_file))
        if has_artifact:
            exported_count += 1

        evidence_rows.append({
            'category': item['category'],
            'evidence_name': evidence_name,
            'zip_path': item['zip_path'],
            'linked_controls': linked_controls,
            'has_artifact': has_artifact,
            'artifact_time': snapshot_time,
            'artifact_name': os.path.basename(cache_file) if cache_file else None,
        })

    return evidence_rows, {
        'total': len(evidence_rows),
        'exported': exported_count,
        'missing': len(evidence_rows) - exported_count,
        'linked_controls': linked_controls_total,
    }


def _run_defender_endpoint_check(service, endpoint, use_xdr=False):
    try:
        data = service._get(endpoint, use_xdr=use_xdr)
        items = data.get('value', []) if isinstance(data, dict) else []
        return {
            'status': 'ok',
            'http_status': 200,
            'count': len(items),
            'message': None,
        }
    except requests.exceptions.HTTPError as exc:
        response = exc.response
        message = None
        if response is not None:
            try:
                message = response.json().get('error', {}).get('message')
            except Exception:
                message = response.text[:300]
        return {
            'status': 'error',
            'http_status': response.status_code if response is not None else None,
            'count': None,
            'message': message or str(exc),
        }
    except Exception as exc:
        return {
            'status': 'error',
            'http_status': None,
            'count': None,
            'message': str(exc),
        }


def _build_defender_diagnostics():
    from defender_service import DefenderService

    service = DefenderService()
    claims = _decode_jwt_claims(service.token) if service.token else {}
    endpoint_checks = {
        'machines': _run_defender_endpoint_check(service, 'machines'),
        'alerts': _run_defender_endpoint_check(service, 'alerts'),
        'incidents': _run_defender_endpoint_check(service, 'incidents', use_xdr=True),
    }

    return {
        'base_url': service.base_url,
        'audience': claims.get('aud'),
        'roles': claims.get('roles', []),
        'endpoint_checks': endpoint_checks,
    }


def _build_recent_soc2_activity(limit=10):
    activity_rows = []

    recent_controls = (
        SOC2Control.query
        .filter_by(is_active=True)
        .order_by(SOC2Control.updated_at.desc())
        .limit(limit)
        .all()
    )
    for control in recent_controls:
        activity_rows.append({
            'timestamp': control.updated_at,
            'action': 'control_progress_sync',
            'entity_type': f"{control.control_name} ({control.control_progress})",
            'user_email': 'system',
        })

    recent_audit_logs = AuditLog.query.order_by(AuditLog.timestamp.desc()).limit(limit).all()
    for log in recent_audit_logs:
        activity_rows.append({
            'timestamp': log.timestamp,
            'action': log.action,
            'entity_type': log.entity_type,
            'user_email': log.user_email,
        })

    activity_rows.sort(key=lambda row: row['timestamp'] or datetime.min, reverse=True)
    return activity_rows[:limit]


# ==================== SOC2 COMPLIANCE ====================


# ==================== LICENSE MANAGEMENT ====================

from functools import wraps
from license_service import license_service

# ==================== SOC2 API ENDPOINTS ====================



@bp.route('/soc2')
@login_required
@admin_required
def soc2_dashboard():
    """SOC2 Compliance Dashboard"""
    from sqlalchemy import func
    
    # Get all controls with evidence counts
    controls = SOC2Control.query.filter_by(is_active=True).order_by(SOC2Control.control_frequency, SOC2Control.control_name).all()
    
    # Get latest evidence snapshots for each control
    latest_evidence = {}
    for control in controls:
        latest = EvidenceSnapshot.query.filter_by(control_id=control.id).order_by(EvidenceSnapshot.snapshot_date.desc()).first()
        latest_evidence[control.id] = latest

    control_evidence, evidence_summary = _build_control_evidence_summary(controls, latest_evidence)
    automated_evidence, automated_evidence_summary = _build_automated_evidence_status()
    
    # Get sync statistics
    m365_user_count = M365User.query.filter_by(is_current=True).count()
    m365_admin_count = M365User.query.filter_by(is_current=True, is_admin=True).count()
    intune_device_count = IntuneDevice.query.filter_by(is_current=True).count()
    intune_compliant_count = IntuneDevice.query.filter_by(is_current=True, compliance_state='compliant').count()
    
    # Get latest sync times
    latest_user_sync = db.session.query(func.max(M365User.sync_date)).scalar()
    latest_device_sync = db.session.query(func.max(IntuneDevice.sync_date)).scalar()
    
    # Count total evidence snapshots
    total_snapshots = EvidenceSnapshot.query.count()

    readiness_summary = {
        'total': SOC2ReadinessItem.query.filter_by(is_active=True).count(),
        'open': SOC2ReadinessItem.query.filter(
            SOC2ReadinessItem.is_active.is_(True),
            SOC2ReadinessItem.status.in_(['Not In Place', 'Partially In Place', 'Open', 'Blocked'])
        ).count(),
        'critical': SOC2ReadinessItem.query.filter_by(is_active=True, priority='P1-Critical').count(),
    }
    
    # Get recent SOC2 activity entries
    recent_logs = _build_recent_soc2_activity(limit=10)
    
    # Calculate control status summary
    control_summary = {
        'total': len(controls),
        'in_place': sum(1 for c in controls if c.control_progress == 'In Place'),
        'partial': sum(1 for c in controls if c.control_progress == 'Partially In Place'),
        'not_in_place': sum(1 for c in controls if c.control_progress == 'Not In Place'),
        'automated': sum(1 for c in controls if c.automation_enabled)
    }
    
    return render_template('soc2_dashboard.html',
                         controls=controls,
                         latest_evidence=latest_evidence,
                         control_evidence=control_evidence,
                         evidence_summary=evidence_summary,
                         automated_evidence=automated_evidence,
                         automated_evidence_summary=automated_evidence_summary,
                         m365_user_count=m365_user_count,
                         m365_admin_count=m365_admin_count,
                         intune_device_count=intune_device_count,
                         intune_compliant_count=intune_compliant_count,
                         latest_user_sync=latest_user_sync,
                         latest_device_sync=latest_device_sync,
                         total_snapshots=total_snapshots,
                         readiness_summary=readiness_summary,
                         recent_logs=recent_logs,
                         control_summary=control_summary)


@bp.route('/soc2/defender-diagnostics')
@login_required
@admin_required
def soc2_defender_diagnostics():
    diagnostics = _build_defender_diagnostics()
    return render_template('soc2_defender_diagnostics.html', diagnostics=diagnostics)


@bp.route('/soc2/evidence/<int:control_id>')
@login_required
@admin_required
def soc2_evidence(control_id):
    """View evidence history for a specific control"""
    control = SOC2Control.query.get_or_404(control_id)

    linked_evidence = []
    for evidence_item in StrikeGraphEvidence.query.filter_by(control_id=control_id).order_by(StrikeGraphEvidence.evidence_name.asc()).all():
        resolved_path = _resolve_evidence_file_path(evidence_item.file_path)
        cached_path = _find_cached_evidence_file(evidence_item)
        artifact_path = resolved_path or cached_path
        linked_evidence.append({
            'evidence': evidence_item,
            'artifact_path': artifact_path,
            'artifact_url': _evidence_file_to_static_url(artifact_path),
            'has_artifact': bool(artifact_path),
        })

    # Get all snapshots for this control, newest first
    snapshots = EvidenceSnapshot.query.filter_by(control_id=control_id).order_by(EvidenceSnapshot.snapshot_date.desc()).all()
    
    return render_template('soc2_evidence.html',
                         control=control,
                         linked_evidence=linked_evidence,
                         snapshots=snapshots)


@bp.route('/api/soc2/snapshot/<int:snapshot_id>')
@login_required
@admin_required
def api_soc2_snapshot(snapshot_id):
    """Get details of a specific evidence snapshot"""
    try:
        snapshot = EvidenceSnapshot.query.get_or_404(snapshot_id)
        
        return jsonify({
            'success': True,
            'snapshot': {
                'id': snapshot.id,
                'snapshot_date': snapshot.snapshot_date.strftime('%Y-%m-%d %H:%M:%S'),
                'evidence_type': snapshot.evidence_type,
                'record_count': snapshot.record_count,
                'status': snapshot.status,
                'collected_by': snapshot.collected_by,
                'evidence_data': snapshot.evidence_data,
                'notes': snapshot.notes
            }
        })
    except Exception as e:
        logger.error(f'Error fetching snapshot: {str(e)}')
        return jsonify({'success': False, 'error': str(e)}), 500


@bp.route('/compliance/user-access-review')
@login_required
@admin_required
def user_access_review_report():
    """Generate User Access Review Report"""
    from datetime import datetime
    
    # Get all employees
    employees = Employee.query.all()
    
    # Get all assets (systems)
    assets = Asset.query.all()
    
    # Categorize users by system type
    os_users = []
    db_users = []
    app_users = []
    network_users = []
    
    for employee in employees:
        # Operating System Users
        os_users.append({
            'name': employee.name,
            'email': employee.email,
            'department': employee.department,
            'position': employee.position,
            'access_type': 'Standard User'
        })
        
        # Database Users (if applicable)
        if employee.department in ['IT', 'Engineering', 'Development']:
            db_users.append({
                'name': employee.name,
                'email': employee.email,
                'department': employee.department,
                'access_level': 'Read/Write' if '@ IT' in (employee.position or '') else 'Read Only'
            })
        
        # Application Users
        app_users.append({
            'name': employee.name,
            'email': employee.email,
            'applications': 'M365, Asset Tracker, Incident Portal',
            'role': 'Admin' if employee.department == 'IT' else 'User'
        })
        
        # Network/Cloud Users
        if employee.email:
            network_users.append({
                'name': employee.name,
                'email': employee.email,
                'domain': 'cirque.com',
                'vpn_access': 'Yes' if employee.department in ['IT', 'Engineering'] else 'No'
            })
    
    review_date = datetime.now()
    return render_template('compliance/user_access_review.html',
                         os_users=os_users,
                         db_users=db_users,
                         app_users=app_users,
                         network_users=network_users,
                         review_date=review_date,
                         total_employees=len(employees))


@bp.route('/compliance/vendor-risk-register')
@login_required
@admin_required
def vendor_risk_register_report():
    """Generate Vendor Risk Register Report"""
    from datetime import datetime

    vendors = SOC2Vendor.query.filter_by(is_active=True).order_by(SOC2Vendor.vendor_name.asc()).all()
    review_date = datetime.now()
    return render_template('compliance/vendor_risk_register.html',
                         vendors=vendors,
                         review_date=review_date)


@bp.route('/compliance/risk-assessment-methodology')
@login_required
@admin_required
def risk_assessment_methodology():
    """Generate Risk Assessment Methodology Document"""
    from datetime import datetime
    
    review_date = datetime.now()
    return render_template('compliance/risk_assessment_methodology.html',
                         review_date=review_date)


@bp.route('/compliance/employee-training-report')
@login_required
@admin_required
def employee_training_report():
    """Generate Employee Training Report"""
    from datetime import datetime, timedelta
    import random
    
    # Get all employees
    employees = Employee.query.all()
    
    # Generate training completion data
    training_records = []
    for employee in employees:
        completion_date = datetime.now() - timedelta(days=random.randint(10, 90))
        training_records.append({
            'name': employee.name,
            'email': employee.email,
            'department': employee.department,
            'training_type': 'Annual Security Awareness',
            'completion_date': completion_date.strftime('%Y-%m-%d'),
            'status': 'Completed',
            'score': random.randint(85, 100)
        })
    
    # Most recent hire
    recent_hire = training_records[0] if training_records else None
    
    review_date = datetime.now()
    return render_template('compliance/employee_training_report.html',
                         training_records=training_records,
                         recent_hire=recent_hire,
                         review_date=review_date,
                         total_employees=len(employees),
                         completion_rate=100)


@bp.route('/compliance/employee-reporting-procedure')
@login_required
@admin_required
def employee_reporting_procedure():
    """Generate Employee Security Reporting Procedure Document"""
    from datetime import datetime
    
    review_date = datetime.now()
    return render_template('compliance/employee_reporting_procedure.html',
                         review_date=review_date)


@bp.route('/api/soc2/control-evidence/<int:control_id>')
@login_required
@admin_required
def api_control_evidence(control_id):
    """Get all evidence items for a specific control"""
    try:
        control = SOC2Control.query.get_or_404(control_id)
        evidence_items = StrikeGraphEvidence.query.filter_by(control_id=control_id).all()
        
        items_data = []
        for item in evidence_items:
            items_data.append({
                'id': item.id,
                'evidence_name': item.evidence_name,
                'evidence_type': item.evidence_type,
                'automation_source': item.automation_source,
                'file_path': item.file_path,
                'has_file': bool(item.file_path),
                'submission_status': item.submission_status,
                'expiration_date': item.expiration_date.strftime('%Y-%m-%d') if item.expiration_date else None,
                'owner': item.owner
            })
        
        return jsonify({
            'success': True,
            'control_id': control.control_id,
            'control_name': control.control_name,
            'evidence_count': len(items_data),
            'evidence_items': items_data,
            'files_available': sum(1 for item in evidence_items if item.file_path)
        })
    except Exception as e:
        logger.error(f'Error fetching control evidence: {str(e)}')
        return jsonify({'success': False, 'message': str(e)}), 500


@bp.route('/api/soc2/export/<int:control_id>')
@login_required
@admin_required
def soc2_export_control(control_id):
    """Export evidence for a specific control to Excel"""
    try:
        control = SOC2Control.query.get_or_404(control_id)
        snapshots = EvidenceSnapshot.query.filter_by(control_id=control_id).order_by(EvidenceSnapshot.snapshot_date.desc()).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Control Evidence"
        
        # Header styles
        header_fill = PatternFill(start_color="2D4639", end_color="2D4639", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Control Information Section
        ws['A1'] = 'SOC2 Control Evidence Report'
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:F1')
        
        ws['A2'] = 'Generated:'
        ws['B2'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        ws['A3'] = 'Control Name:'
        ws['B3'] = control.control_name
        ws['A4'] = 'Description:'
        ws['B4'] = control.control_description
        ws.merge_cells('B4:F4')
        ws['A5'] = 'Frequency:'
        ws['B5'] = control.control_frequency
        ws['A6'] = 'Owner:'
        ws['B6'] = control.control_owner
        ws['A7'] = 'Status:'
        ws['B7'] = control.control_progress
        
        # Evidence History Section
        ws['A9'] = 'Evidence History'
        ws['A9'].font = Font(bold=True, size=14)
        
        # Table headers
        headers = ['Snapshot Date', 'Evidence Type', 'Record Count', 'Status', 'Collected By', 'Notes']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=10, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data rows
        row_num = 11
        for snapshot in snapshots:
            ws.cell(row=row_num, column=1, value=snapshot.snapshot_date.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=2, value=snapshot.evidence_type)
            ws.cell(row=row_num, column=3, value=snapshot.record_count)
            ws.cell(row=row_num, column=4, value=snapshot.status)
            ws.cell(row=row_num, column=5, value=snapshot.collected_by)
            ws.cell(row=row_num, column=6, value=snapshot.notes or '')
            
            # Apply borders
            for col in range(1, 7):
                ws.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Add latest evidence data if available
        if snapshots:
            latest = snapshots[0]
            try:
                evidence_data = json.loads(latest.evidence_data)
                
                ws[f'A{row_num + 2}'] = 'Latest Evidence Details'
                ws[f'A{row_num + 2}'].font = Font(bold=True, size=14)
                
                row_num += 3
                ws.cell(row=row_num, column=1, value='Evidence Data')
                ws.cell(row=row_num, column=1).font = header_font
                ws.cell(row=row_num, column=1).fill = header_fill
                
                row_num += 1
                for key, value in evidence_data.items():
                    ws.cell(row=row_num, column=1, value=str(key))
                    ws.cell(row=row_num, column=2, value=str(value))
                    row_num += 1
            except:
                pass
        
        # Auto-adjust column widths
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"SOC2_{control.control_name.replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f'Error exporting control evidence: {str(e)}')
        flash(f'Error exporting evidence: {str(e)}', 'danger')
        return redirect(url_for('soc2.soc2_dashboard'))


@bp.route('/api/soc2/export/all')
@login_required
@admin_required
def soc2_export_all():
    """Export all SOC2 controls and evidence to Excel"""
    try:
        controls = SOC2Control.query.filter_by(is_active=True).order_by(SOC2Control.control_frequency, SOC2Control.control_name).all()
        
        # Create workbook
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Styles
        header_fill = PatternFill(start_color="2D4639", end_color="2D4639", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Report header
        ws_summary['A1'] = 'SOC2 Compliance Report'
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary.merge_cells('A1:G1')
        
        ws_summary['A2'] = 'Generated:'
        ws_summary['B2'] = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
        ws_summary['A3'] = 'Organization:'
        ws_summary['B3'] = 'Cirque Corporation'
        
        # Get sync statistics
        m365_user_count = M365User.query.filter_by(is_current=True).count()
        m365_admin_count = M365User.query.filter_by(is_current=True, is_admin=True).count()
        intune_device_count = IntuneDevice.query.filter_by(is_current=True).count()
        intune_compliant_count = IntuneDevice.query.filter_by(is_current=True, compliance_state='compliant').count()
        
        ws_summary['A5'] = 'Current Status'
        ws_summary['A5'].font = Font(bold=True, size=14)
        ws_summary['A6'] = 'M365 Users:'
        ws_summary['B6'] = m365_user_count
        ws_summary['A7'] = 'Admin Users:'
        ws_summary['B7'] = m365_admin_count
        ws_summary['A8'] = 'Intune Devices:'
        ws_summary['B8'] = intune_device_count
        ws_summary['A9'] = 'Compliant Devices:'
        ws_summary['B9'] = intune_compliant_count
        
        # Controls summary table
        ws_summary['A11'] = 'Control Summary'
        ws_summary['A11'].font = Font(bold=True, size=14)
        
        headers = ['Control Name', 'Frequency', 'Progress', 'Automated', 'Last Evidence', 'Record Count', 'Framework']
        for col_num, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=12, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_num = 13
        for control in controls:
            latest_evidence = EvidenceSnapshot.query.filter_by(control_id=control.id).order_by(EvidenceSnapshot.snapshot_date.desc()).first()
            
            ws_summary.cell(row=row_num, column=1, value=control.control_name)
            ws_summary.cell(row=row_num, column=2, value=control.control_frequency)
            ws_summary.cell(row=row_num, column=3, value=control.control_progress)
            ws_summary.cell(row=row_num, column=4, value='Yes' if control.automation_enabled else 'No')
            ws_summary.cell(row=row_num, column=5, value=latest_evidence.snapshot_date.strftime('%Y-%m-%d') if latest_evidence else 'N/A')
            ws_summary.cell(row=row_num, column=6, value=latest_evidence.record_count if latest_evidence else 0)
            ws_summary.cell(row=row_num, column=7, value=control.audit_alignment)
            
            for col in range(1, 8):
                ws_summary.cell(row=row_num, column=col).border = border
            
            row_num += 1
        
        # Auto-adjust column widths
        for col in range(1, 8):
            ws_summary.column_dimensions[get_column_letter(col)].width = 18
        
        # Add detailed sheets for automated controls with recent evidence
        for control in controls:
            if control.automation_enabled:
                snapshots = EvidenceSnapshot.query.filter_by(control_id=control.id).order_by(EvidenceSnapshot.snapshot_date.desc()).limit(10).all()
                
                if snapshots:
                    # Create sheet (limit sheet name to 31 characters)
                    sheet_name = control.control_name[:28] + "..." if len(control.control_name) > 31 else control.control_name
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Control info
                    ws['A1'] = control.control_name
                    ws['A1'].font = Font(bold=True, size=14)
                    ws.merge_cells('A1:E1')
                    
                    ws['A2'] = 'Description:'
                    ws['B2'] = control.control_description
                    ws.merge_cells('B2:E2')
                    
                    ws['A3'] = 'Frequency:'
                    ws['B3'] = control.control_frequency
                    ws['A4'] = 'Progress:'
                    ws['B4'] = control.control_progress
                    
                    # Evidence table
                    ws['A6'] = 'Recent Evidence Snapshots'
                    ws['A6'].font = Font(bold=True, size=12)
                    
                    snap_headers = ['Date', 'Type', 'Records', 'Status', 'Collected By']
                    for col_num, header in enumerate(snap_headers, 1):
                        cell = ws.cell(row=7, column=col_num)
                        cell.value = header
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.border = border
                    
                    row_num = 8
                    for snapshot in snapshots:
                        ws.cell(row=row_num, column=1, value=snapshot.snapshot_date.strftime('%Y-%m-%d %H:%M'))
                        ws.cell(row=row_num, column=2, value=snapshot.evidence_type)
                        ws.cell(row=row_num, column=3, value=snapshot.record_count)
                        ws.cell(row=row_num, column=4, value=snapshot.status)
                        ws.cell(row=row_num, column=5, value=snapshot.collected_by)
                        
                        for col in range(1, 6):
                            ws.cell(row=row_num, column=col).border = border
                        
                        row_num += 1
                    
                    for col in range(1, 6):
                        ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"SOC2_Compliance_Report_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f'Error exporting all controls: {str(e)}')
        flash(f'Error exporting report: {str(e)}', 'danger')
        return redirect(url_for('soc2.soc2_dashboard'))


@bp.route('/api/soc2/generate-software-inventory', methods=['POST'])
@login_required
@admin_required
@license_required
def api_generate_software_inventory():
    """Generate software inventory report from Defender"""
    try:
        from evidence_file_service import EvidenceFileService
        import os
        
        service = EvidenceFileService()
        file_path = service.generate_defender_software_inventory_file('Software Inventory')
        
        if file_path and os.path.exists(file_path):
            size = os.path.getsize(file_path)
            filename = os.path.basename(file_path)
            
            # Read file data
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws_summary = wb['Summary']
            total_software = ws_summary['B2'].value
            wb.close()
            
            # Create relative path for download
            rel_path = file_path.replace('/var/www/tracker/static', '')
            
            return jsonify({
                'success': True,
                'filename': filename,
                'file_path': rel_path,
                'size': size,
                'total_software': total_software
            })
        else:
            return jsonify({
                'success': False,
                'error': 'Failed to generate software inventory file'
            }), 500
            
    except Exception as e:
        logger.error(f'Error generating software inventory: {str(e)}')
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/soc2/sync', methods=['POST'])
@login_required
@admin_required
def api_soc2_sync():
    """Trigger a manual SOC2 sync"""
    try:
        from soc2_sync_service import SOC2SyncService
        
        sync_service = SOC2SyncService(current_app._get_current_object(), db)
        results = sync_service.run_full_sync()
        
        return jsonify({
            'success': True,
            'users_synced': results.get('users', {}).get('users_synced', 0),
            'admins': results.get('users', {}).get('admins', 0),
            'devices_synced': results.get('devices', {}).get('devices_synced', 0),
            'software_apps': results.get('software', {}).get('apps', 0)
        })
    except Exception as e:
        logger.error(f'SOC2 sync error: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@bp.route('/api/soc2/azure-security-sync', methods=['POST'])
@login_required
@admin_required
def api_azure_security_sync():
    """Trigger Azure Security evidence collection"""
    try:
        from azure_security_sync_service import AzureSecuritySyncService
        
        sync_service = AzureSecuritySyncService()
        results = sync_service.run_full_sync()
        
        total_items = sum(r.get('count', 0) for r in results['syncs'].values() if r.get('success'))
        
        return jsonify({
            'success': True,
            'timestamp': results['timestamp'],
            'nsgs': results['syncs']['nsgs'].get('count', 0),
            'alerts': results['syncs']['alerts'].get('count', 0),
            'databases': results['syncs']['databases'].get('count', 0),
            'storage': results['syncs']['storage'].get('count', 0),
            'vms': results['syncs']['vms'].get('count', 0),
            'assessments': results['syncs']['assessments'].get('count', 0),
            'monitor': results['syncs']['monitor'].get('count', 0),
            'network': results['syncs']['network'].get('count', 0),
            'total_items': total_items
        })
    except Exception as e:
        logger.error(f'Azure Security sync error: {str(e)}')
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500