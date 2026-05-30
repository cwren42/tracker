"""Azure evidence-file generators — mixin for EvidenceFileService.
Split from evidence_file_service.py."""

import os
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from app import db
from soc2_models import (
    EvidenceSnapshot, StrikeGraphEvidence, SOC2Control,
    M365User, IntuneDevice, DeviceSoftware, AdminRoleSnapshot,
    AzureNetworkSecurityGroup, AzureSecurityAlert, AzureDatabase,
    AzureStorageAccount, AzureVirtualMachine, AzureSecurityAssessment,
    AzureMonitorAlert, AzureNetworkTopology
)
from teamviewer_evidence_service import TeamViewerEvidenceService
from defender_service import DefenderService



class EvidenceAzureMixin:
    """Azure evidence generators (mixed into EvidenceFileService)."""
    def generate_azure_nsg_file(self, evidence_name):
        """Generate Azure Network Security Groups file"""
        nsgs = AzureNetworkSecurityGroup.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Network Security Groups')
        headers = ['NSG Name', 'Resource Group', 'Location', 'Security Rules', 'Associated Subnets', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, nsg in enumerate(nsgs, 2):
            rules_data = json.loads(nsg.security_rules) if nsg.security_rules else []
            rules_count = len(rules_data)
            
            ws.cell(row_idx, 1, nsg.name or '')
            ws.cell(row_idx, 2, nsg.resource_group or '')
            ws.cell(row_idx, 3, nsg.location or '')
            ws.cell(row_idx, 4, f'{rules_count} rules')
            ws.cell(row_idx, 5, '')
            ws.cell(row_idx, 6, nsg.sync_date.strftime('%Y-%m-%d %H:%M') if nsg.sync_date else '')
        
        # Add detailed rules sheet
        ws_rules = wb.create_sheet('Security Rules')
        rule_headers = ['NSG Name', 'Rule Name', 'Priority', 'Direction', 'Access', 'Protocol', 'Source', 'Destination', 'Ports']
        self.style_header_row(ws_rules, rule_headers)
        
        rule_row = 2
        for nsg in nsgs:
            rules_data = json.loads(nsg.security_rules) if nsg.security_rules else []
            for rule in rules_data:
                ws_rules.cell(rule_row, 1, nsg.name)
                ws_rules.cell(rule_row, 2, rule.get('name', ''))
                ws_rules.cell(rule_row, 3, rule.get('priority', ''))
                ws_rules.cell(rule_row, 4, rule.get('direction', ''))
                ws_rules.cell(rule_row, 5, rule.get('access', ''))
                ws_rules.cell(rule_row, 6, rule.get('protocol', ''))
                ws_rules.cell(rule_row, 7, rule.get('sourceAddressPrefix', ''))
                ws_rules.cell(rule_row, 8, rule.get('destinationAddressPrefix', ''))
                ws_rules.cell(rule_row, 9, rule.get('destinationPortRange', ''))
                rule_row += 1
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_security_alerts_file(self, evidence_name):
        """Generate Azure Security Alerts file"""
        alerts = AzureSecurityAlert.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Security Alerts')
        headers = ['Alert Name', 'Severity', 'Status', 'Description', 'Affected Resource', 'Detection Time', 'Remediation Steps']
        self.style_header_row(ws, headers)
        
        for row_idx, alert in enumerate(alerts, 2):
            ws.cell(row_idx, 1, alert.alert_name or '')
            ws.cell(row_idx, 2, alert.severity or '')
            ws.cell(row_idx, 3, alert.status or '')
            ws.cell(row_idx, 4, alert.description or '')
            ws.cell(row_idx, 5, alert.affected_resource or '')
            ws.cell(row_idx, 6, alert.detection_time.strftime('%Y-%m-%d %H:%M') if alert.detection_time else '')
            ws.cell(row_idx, 7, alert.remediation_steps or '')
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_databases_file(self, evidence_name):
        """Generate Azure SQL Databases file"""
        databases = AzureDatabase.query.filter_by(is_current=True).all()
        if not databases:
            from azure_security_service import AzureSecurityService

            azure_service = AzureSecurityService()
            databases = azure_service.get_sql_databases()
        
        wb, ws = self.create_styled_workbook('SQL Databases')
        headers = ['Database Name', 'Server Name', 'Resource Group', 'Location', 'TDE Enabled', 'Firewall Rules', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, db in enumerate(databases, 2):
            ws.cell(row_idx, 1, getattr(db, 'database_name', None) or db.get('database_name', ''))
            ws.cell(row_idx, 2, getattr(db, 'server_name', None) or db.get('server_name', ''))
            ws.cell(row_idx, 3, getattr(db, 'resource_group', None) or db.get('resource_group', ''))
            ws.cell(row_idx, 4, getattr(db, 'location', None) or db.get('location', ''))
            ws.cell(row_idx, 5, 'Yes' if (getattr(db, 'tde_enabled', None) if not isinstance(db, dict) else db.get('tde_enabled')) else 'No')
            ws.cell(row_idx, 6, getattr(db, 'firewall_rules', None) or db.get('firewall_rules', ''))
            sync_date = getattr(db, 'sync_date', None) if not isinstance(db, dict) else None
            ws.cell(row_idx, 7, sync_date.strftime('%Y-%m-%d %H:%M') if sync_date else datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_storage_file(self, evidence_name):
        """Generate Azure Storage Accounts file"""
        storage = AzureStorageAccount.query.filter_by(is_current=True).all()
        if not storage:
            from azure_security_service import AzureSecurityService

            azure_service = AzureSecurityService()
            storage = azure_service.get_storage_accounts()
        
        wb, ws = self.create_styled_workbook('Storage Accounts')
        headers = ['Storage Account', 'Resource Group', 'Location', 'Encryption Enabled', 'HTTPS Only', 'Min TLS Version', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, account in enumerate(storage, 2):
            ws.cell(row_idx, 1, getattr(account, 'storage_account_name', None) or account.get('name', ''))
            ws.cell(row_idx, 2, getattr(account, 'resource_group', None) or account.get('resource_group', ''))
            ws.cell(row_idx, 3, getattr(account, 'location', None) or account.get('location', ''))
            ws.cell(row_idx, 4, 'Yes' if (getattr(account, 'encryption_enabled', None) if not isinstance(account, dict) else account.get('encryption_enabled')) else 'No')
            ws.cell(row_idx, 5, 'Yes' if (getattr(account, 'https_only', None) if not isinstance(account, dict) else account.get('https_only')) else 'No')
            ws.cell(row_idx, 6, getattr(account, 'min_tls_version', None) or account.get('tls_version', ''))
            sync_date = getattr(account, 'sync_date', None) if not isinstance(account, dict) else None
            ws.cell(row_idx, 7, sync_date.strftime('%Y-%m-%d %H:%M') if sync_date else datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_vms_file(self, evidence_name):
        """Generate Azure Virtual Machines file"""
        vms = AzureVirtualMachine.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Virtual Machines')
        headers = ['VM Name', 'Resource Group', 'Location', 'VM Size', 'OS Type', 'Disk Encryption', 'Power State', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, vm in enumerate(vms, 2):
            ws.cell(row_idx, 1, vm.name or '')
            ws.cell(row_idx, 2, vm.resource_group or '')
            ws.cell(row_idx, 3, vm.location or '')
            ws.cell(row_idx, 4, vm.vm_size or '')
            ws.cell(row_idx, 5, vm.os_type or '')
            ws.cell(row_idx, 6, 'Yes' if vm.disk_encryption else 'No')
            ws.cell(row_idx, 7, '')
            ws.cell(row_idx, 8, vm.sync_date.strftime('%Y-%m-%d %H:%M') if vm.sync_date else '')
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_assessments_file(self, evidence_name):
        """Generate Azure Security Assessments file"""
        assessments = AzureSecurityAssessment.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Security Assessments')
        headers = ['Assessment Name', 'Resource', 'Status', 'Severity', 'Description', 'Remediation', 'Last Sync']
        self.style_header_row(ws, headers)
        
        if assessments:
            for row_idx, assessment in enumerate(assessments, 2):
                ws.cell(row_idx, 1, assessment.name or '')
                ws.cell(row_idx, 2, assessment.resource_id or '')
                ws.cell(row_idx, 3, assessment.status or '')
                ws.cell(row_idx, 4, assessment.severity or '')
                ws.cell(row_idx, 5, assessment.description or '')
                ws.cell(row_idx, 6, assessment.remediation or '')
                ws.cell(row_idx, 7, assessment.sync_date.strftime('%Y-%m-%d %H:%M') if assessment.sync_date else '')
        else:
            # No data yet - add placeholder row
            ws.cell(2, 1, 'No assessments available')
            ws.cell(2, 2, 'Azure Security Sync has not been run yet')
            ws.cell(2, 3, 'Pending')
            ws.cell(2, 4, 'Info')
            ws.cell(2, 5, 'Run Azure Security Sync from SOC2 Dashboard to collect vulnerability scan results')
            ws.cell(2, 6, 'Click "Azure Security Sync" button on dashboard')
            ws.cell(2, 7, datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_monitor_alerts_file(self, evidence_name):
        """Generate Azure Monitor Alerts file"""
        alerts = AzureMonitorAlert.query.filter_by(is_current=True).all()
        if not alerts:
            from azure_security_service import AzureSecurityService

            azure_service = AzureSecurityService()
            alerts = azure_service.get_monitor_alerts()
        
        wb, ws = self.create_styled_workbook('Monitor Alerts')
        headers = ['Alert Name', 'Resource Group', 'Target Resource', 'Condition', 'Severity', 'Enabled', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, alert in enumerate(alerts, 2):
            if isinstance(alert, dict):
                criteria_data = alert.get('criteria') or {}
                all_of = criteria_data.get('allOf') or []
                condition = all_of[0].get('metricName', '') if all_of else criteria_data.get('odata.type', '')
                resource_group = alert.get('resource_group', '')
                target_resource = alert.get('target_resource', '')
                sync_date = None
            else:
                criteria_data = json.loads(alert.criteria) if alert.criteria else {}
                condition = criteria_data.get('allOf', [{}])[0].get('metricName', '') if criteria_data else ''
                resource_group = alert.resource_group or ''
                target_resource = alert.target_resource or ''
                sync_date = alert.sync_date
            
            ws.cell(row_idx, 1, getattr(alert, 'alert_name', None) or alert.get('name', ''))
            ws.cell(row_idx, 2, resource_group)
            ws.cell(row_idx, 3, target_resource)
            ws.cell(row_idx, 4, condition)
            ws.cell(row_idx, 5, getattr(alert, 'severity', None) or alert.get('severity', ''))
            ws.cell(row_idx, 6, 'Yes' if (getattr(alert, 'enabled', None) if not isinstance(alert, dict) else alert.get('enabled')) else 'No')
            ws.cell(row_idx, 7, sync_date.strftime('%Y-%m-%d %H:%M') if sync_date else datetime.utcnow().strftime('%Y-%m-%d %H:%M'))
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
    def generate_azure_network_topology_file(self, evidence_name):
        """Generate Azure Network Topology file"""
        networks = AzureNetworkTopology.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Network Topology')
        headers = ['VNet Name', 'Resource Group', 'Location', 'Address Space', 'Subnets', 'Peerings', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, vnet in enumerate(networks, 2):
            subnets_data = json.loads(vnet.subnets) if vnet.subnets else []
            
            subnet_count = len(subnets_data)
            
            ws.cell(row_idx, 1, vnet.name or '')
            ws.cell(row_idx, 2, vnet.resource_group or '')
            ws.cell(row_idx, 3, vnet.location or '')
            ws.cell(row_idx, 4, vnet.address_space or '')
            ws.cell(row_idx, 5, f'{subnet_count} subnets')
            ws.cell(row_idx, 6, '')
            ws.cell(row_idx, 7, vnet.sync_date.strftime('%Y-%m-%d %H:%M') if vnet.sync_date else '')
        
        # Add detailed subnets sheet
        ws_subnets = wb.create_sheet('Subnets')
        subnet_headers = ['VNet Name', 'Subnet Name', 'Address Prefix', 'NSG Attached']
        self.style_header_row(ws_subnets, subnet_headers)
        
        subnet_row = 2
        for vnet in networks:
            subnets_data = json.loads(vnet.subnets) if vnet.subnets else []
            for subnet in subnets_data:
                ws_subnets.cell(subnet_row, 1, vnet.name)
                ws_subnets.cell(subnet_row, 2, subnet.get('name', ''))
                ws_subnets.cell(subnet_row, 3, subnet.get('addressPrefix', ''))
                ws_subnets.cell(subnet_row, 4, 'Yes' if subnet.get('networkSecurityGroup') else 'No')
                subnet_row += 1
        
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path
    
