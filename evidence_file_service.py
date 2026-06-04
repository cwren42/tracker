"""
Evidence File Generation Service for StrikeGraph
Generates downloadable evidence files for SOC2 compliance
"""

import os
import re
import json
from datetime import datetime, timedelta
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from app import db
from soc2_models import (
    EvidenceSnapshot, StrikeGraphEvidence, SOC2Control,
    M365User, IntuneDevice, DeviceSoftware, AdminRoleSnapshot,
    AzureNetworkSecurityGroup, AzureSecurityAlert, AzureDatabase,
    AzureStorageAccount, AzureVirtualMachine, AzureSecurityAssessment,
    AzureMonitorAlert, AzureNetworkTopology
)
from teamviewer_evidence_service import TeamViewerEvidenceService
from defender_service import DefenderService
from evidence_azure import EvidenceAzureMixin


# ---------------------------------------------------------------------------
# ISMS-manual policy evidence resolution
# ---------------------------------------------------------------------------
# The published ISMS Manual (isms_document slug=isms-manual, current version in
# isms_document_version.markdown_body) is section-addressable: each policy /
# procedure is a markdown heading of the form ``## IS-...-CIRQ##-...: Title``.
# For Policy-type StrikeGraphEvidence items we resolve the policy to one or more
# of these IS-section IDs and extract the section text to a dated PDF artifact.
#
# Resolution order at runtime (see EvidenceFileService._resolve_policy_sections):
#   (a) the linked SOC2Control.authoritative_docs IS-* IDs (single source of truth)
#   (b) an explicit, manually verified name->section map for items whose control
#       has no authoritative_docs / no control link (see below)
#   (c) a last-resort title search of the manual headings
#
# Every entry below was verified against ISMS Manual v4 (document_id=1) before
# being committed: each IS-ID exists as a real heading section in markdown_body.
POLICY_EVIDENCE_SECTION_MAP = {
    # "Access Removal" lives in the Access Control Procedure (§4.2.1 Employee
    # Termination: keycards/accounts revoked on termination).
    'Access Removal Procedures/Checklist': ['IS-AIR01-CIRQ04-A00'],
    # no control link at all -> resolved by name against manual headings
    'Backup Policy': ['IS-AIR01-CIRQ08-A00'],
    'Backup Restoration Procedures': ['IS-AIR01-CIRQ08-A00'],
    'Information Security Policy': ['IS-APM01-CIRQ01-A00'],
    # Technical Vulnerability Management is §4.8 of the Operations Security
    # Policy (cf. control "Vulnerability Scan" authoritative_docs).
    'Vulnerability Management Policy': ['IS-AIR01-CIRQ03-A00'],
}

# Policy items deliberately reverted from ISMS auto-generation back to manual /
# HR-sourced evidence (the real StrikeGraph evidence is a signed HR document or a
# dedicated source document, NOT the ISMS manual's policy text). The resolver
# short-circuits these to ([], None) BEFORE any resolution path -- including the
# linked control's authoritative_docs -- so a stray "Generate now" can never
# re-pull the wrong-source section. The control's authoritative_docs pointer is
# intentionally LEFT intact as a documentation reference; only the auto-gen
# evidence path is neutralized.
POLICY_EVIDENCE_REVERTED_TO_MANUAL = {
    'Acceptable Use Policy',                     # signed AUP from HR
    'Business Continuity Plan',                  # dedicated BCP source
    'Vendor Management Policy and Procedures',   # different source than Supplier Relationships
}

# Policy items the ISMS manual genuinely does NOT contain (so they must stay
# Manual and be sourced elsewhere). Recorded here so the resolver logs rather
# than fabricates. Code of Conduct is in the Employee Handbook, not the ISMS.
POLICY_EVIDENCE_NOT_IN_MANUAL = {
    'Code of Conduct',
}

# IS-section reference tokens that appear in authoritative_docs but are NOT
# extractable manual sections (external files / formats) -> ignored on resolve.
_NON_SECTION_AUTH_TOKENS = ('.docx', '.pdf', '.xlsx')

_IS_SECTION_RE = re.compile(r'IS-[A-Z0-9]+-CIRQ[0-9]+(?:-[A-Z0-9]+)?')
_IS_HEADING_RE = re.compile(
    r'^(#{1,6})\s+(IS-[A-Z0-9]+-CIRQ[0-9]+(?:-[A-Z0-9]+)?)\s*:\s*(.+)$'
)


class EvidenceFileService(EvidenceAzureMixin):
    """Service to generate evidence files for StrikeGraph upload"""
    
    def __init__(self):
        self.evidence_dir = '/var/www/tracker/static/evidence'
        self.ensure_directories()
    
    def _sanitize_for_excel(self, text):
        """Remove illegal characters that Excel doesn't allow in cells"""
        if not text:
            return ''
        # Remove control characters and illegal XML characters
        import re
        # Keep only valid characters (printable ASCII + extended ASCII - control chars)
        return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', str(text))
    
    def ensure_directories(self):
        """Create evidence directories if they don't exist"""
        dirs = [
            self.evidence_dir,
            f'{self.evidence_dir}/m365',
            f'{self.evidence_dir}/M365',
            f'{self.evidence_dir}/M365/Defender',
            f'{self.evidence_dir}/azure',
            f'{self.evidence_dir}/isms',
            f'{self.evidence_dir}/manual',
            f'{self.evidence_dir}/teamviewer',
            f'{self.evidence_dir}/rmm',
        ]
        for dir_path in dirs:
            os.makedirs(dir_path, exist_ok=True)
    
    def generate_filename(self, evidence_name, extension='xlsx'):
        """Generate a standardized filename"""
        # Sanitize filename
        safe_name = "".join(c for c in evidence_name if c.isalnum() or c in (' ', '-', '_')).rstrip()
        safe_name = safe_name.replace(' ', '_')
        timestamp = datetime.now().strftime('%Y%m%d')
        return f"{safe_name}_{timestamp}.{extension}"
    
    def get_file_path(self, evidence_name, automation_source, extension='xlsx'):
        """Get full file path for evidence"""
        filename = self.generate_filename(evidence_name, extension)
        
        if automation_source in ['M365/Intune', 'M365']:
            return f'{self.evidence_dir}/m365/{filename}'
        elif automation_source == 'M365/Defender':
            return f'{self.evidence_dir}/M365/Defender/{filename}'
        elif automation_source == 'Azure':
            return f'{self.evidence_dir}/azure/{filename}'
        elif automation_source == 'ISMS':
            return f'{self.evidence_dir}/isms/{filename}'
        elif automation_source == 'TeamViewer':
            return f'{self.evidence_dir}/teamviewer/{filename}'
        elif automation_source == 'RMM':
            return f'{self.evidence_dir}/rmm/{filename}'
        else:
            return f'{self.evidence_dir}/manual/{filename}'
    
    def create_styled_workbook(self, title):
        """Create a styled Excel workbook"""
        wb = Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel sheet name limit
        return wb, ws
    
    def style_header_row(self, ws, headers):
        """Apply styling to header row"""
        header_fill = PatternFill(start_color="2D4639", end_color="2D4639", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(1, col_idx, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[cell.column_letter].width = 20
    
    def generate_m365_users_file(self, evidence_name):
        """Generate M365 Users list file"""
        users = M365User.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('M365 Users')
        headers = ['Display Name', 'Email', 'Job Title', 'Department', 'Office', 'Account Enabled', 'Is Admin']
        self.style_header_row(ws, headers)
        
        for row_idx, user in enumerate(users, 2):
            ws.cell(row_idx, 1, user.display_name or '')
            ws.cell(row_idx, 2, user.user_principal_name or '')
            ws.cell(row_idx, 3, user.job_title or '')
            ws.cell(row_idx, 4, user.department or '')
            ws.cell(row_idx, 5, user.office_location or '')
            ws.cell(row_idx, 6, 'Yes' if user.account_enabled else 'No')
            ws.cell(row_idx, 7, 'Yes' if user.is_admin else 'No')
        
        file_path = self.get_file_path(evidence_name, 'M365/Intune')
        wb.save(file_path)
        return file_path
    
    def generate_admin_users_file(self, evidence_name):
        """Generate Administrator Access list file"""
        admin_users = M365User.query.filter_by(is_admin=True, is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Admin Users')
        headers = ['Display Name', 'Email', 'Job Title', 'Department', 'Admin Roles', 'Last Sync']
        self.style_header_row(ws, headers)
        
        for row_idx, user in enumerate(admin_users, 2):
            # Get admin roles from AdminRoleSnapshot
            roles = AdminRoleSnapshot.query.filter_by(
                user_principal_name=user.user_principal_name,
                status='active'
            ).all()
            role_names = ', '.join([r.role_name for r in roles]) if roles else 'Administrator'
            
            ws.cell(row_idx, 1, user.display_name or '')
            ws.cell(row_idx, 2, user.user_principal_name or '')
            ws.cell(row_idx, 3, user.job_title or '')
            ws.cell(row_idx, 4, user.department or '')
            ws.cell(row_idx, 5, role_names)
            ws.cell(row_idx, 6, user.sync_date.strftime('%Y-%m-%d %H:%M') if user.sync_date else '')
        
        file_path = self.get_file_path(evidence_name, 'M365/Intune')
        wb.save(file_path)
        return file_path
    
    def generate_intune_devices_file(self, evidence_name):
        """Generate Intune Devices list file"""
        devices = IntuneDevice.query.filter_by(is_current=True).all()
        
        wb, ws = self.create_styled_workbook('Intune Devices')
        headers = ['Device Name', 'User', 'Model', 'OS Version', 'Compliant', 'Encrypted', 'Last Sync', 'Management Agent']
        self.style_header_row(ws, headers)
        
        for row_idx, device in enumerate(devices, 2):
            ws.cell(row_idx, 1, device.device_name or '')
            ws.cell(row_idx, 2, device.user_display_name or device.user_principal_name or '')
            ws.cell(row_idx, 3, f"{device.manufacturer or ''} {device.model or ''}".strip())
            ws.cell(row_idx, 4, device.os_version or '')
            ws.cell(row_idx, 5, 'Yes' if device.compliance_state == 'compliant' else 'No')
            ws.cell(row_idx, 6, 'Yes' if device.is_encrypted else 'No')
            ws.cell(row_idx, 7, device.last_sync_datetime.strftime('%Y-%m-%d %H:%M') if device.last_sync_datetime else '')
            ws.cell(row_idx, 8, device.management_agent or '')
        
        file_path = self.get_file_path(evidence_name, 'M365/Intune')
        wb.save(file_path)
        return file_path
    
    def generate_device_software_file(self, evidence_name):
        """Generate Device Software/Antivirus inventory file using Defender data"""
        # Use Defender for antivirus/security software information
        defender_service = DefenderService()
        machines = defender_service.get_machines()
        
        wb, ws = self.create_styled_workbook('Antivirus Configuration')
        headers = ['Device Name', 'Health State', 'Risk Score', 'OS Platform', 'OS Version', 
                   'Antivirus Status', 'Last Seen', 'Onboarded']
        self.style_header_row(ws, headers)
        
        for row_idx, machine in enumerate(machines, 2):
            ws.cell(row_idx, 1, machine.get('computerDnsName', 'Unknown'))
            ws.cell(row_idx, 2, machine.get('healthStatus', 'Unknown'))
            ws.cell(row_idx, 3, machine.get('riskScore', 'None'))
            ws.cell(row_idx, 4, machine.get('osPlatform', 'Unknown'))
            ws.cell(row_idx, 5, machine.get('osVersion', 'Unknown'))
            
            # Determine antivirus status from health state
            health = machine.get('healthStatus', '').lower()
            if 'active' in health:
                av_status = 'Protected'
            elif 'inactive' in health:
                av_status = 'Not Protected'
            else:
                av_status = 'Active' if health else 'Unknown'
            
            ws.cell(row_idx, 6, av_status)
            ws.cell(row_idx, 7, machine.get('lastSeen', '')[:19] if machine.get('lastSeen') else '')
            ws.cell(row_idx, 8, machine.get('onboardingStatus', 'Unknown'))
        
        file_path = self.get_file_path(evidence_name, 'M365/Defender')
        wb.save(file_path)
        return file_path
    
    def generate_isms_policy_pdf(self, evidence_name):
        """Generate PDF for ISMS policy document"""
        try:
            # Read ISMS manual to extract policy
            from docx import Document
            isms_path = '/home/webuser/ISMS-Manual2025v1.docx'
            
            if not os.path.exists(isms_path):
                print(f"ISMS manual not found at {isms_path}")
                return None
            
            doc = Document(isms_path)
            
            # Policy name mapping
            policy_map = {
                'Acceptable Use Policy': 'Acceptable Use Policy',
                'Access Removal Procedures/Checklist': 'Access Removal',
                'Backup Policy': 'Backup Policy',
                'Backup Restoration Procedures': 'Backup',
                'Change Management Policy': 'Change Management',
                'Code of Conduct': 'Code of Conduct',
                'Data Classification Policy': 'Data Classification',
                'Data Deletion': 'Data Disposal',
                'Data Management Policy': 'Data Management',
                'Incident Response Plan': 'Incident Response',
                'Information Security Policy': 'Information Security',
                'Logical Access Policy and Procedures': 'Logical Access',
                'Password Policy': 'Password',
                'Patch Management Policy': 'Patch Management',
                'Vulnerability Management Policy': 'Vulnerability'
            }
            
            search_term = policy_map.get(evidence_name, evidence_name)
            
            # Extract policy content
            policy_content = []
            in_policy = False
            
            for para in doc.paragraphs:
                text = para.text.strip()
                
                # Start capturing when we find the policy title
                if search_term.lower() in text.lower() and (
                    para.style.name.startswith('Heading') or 
                    any(run.bold for run in para.runs)
                ):
                    in_policy = True
                    policy_content.append(('heading', text))
                    continue
                
                # Stop at next policy/section
                if in_policy and para.style.name.startswith('Heading') and len(policy_content) > 5:
                    break
                
                # Capture content
                if in_policy and text:
                    if para.style.name.startswith('Heading'):
                        policy_content.append(('subheading', text))
                    else:
                        policy_content.append(('body', text))
            
            if not policy_content:
                print(f"Policy content not found for: {evidence_name}")
                return None
            
            # Generate PDF
            file_path = self.get_file_path(evidence_name, 'ISMS', 'pdf')
            pdf = SimpleDocTemplate(file_path, pagesize=letter)
            story = []
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                textColor='#2D4639',
                spaceAfter=30,
                alignment=TA_CENTER
            )
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=14,
                textColor='#2D4639',
                spaceAfter=12,
                spaceBefore=12
            )
            body_style = ParagraphStyle(
                'CustomBody',
                parent=styles['BodyText'],
                fontSize=11,
                alignment=TA_JUSTIFY,
                spaceAfter=12
            )
            
            # Add header
            story.append(Paragraph("Cirque Corporation", title_style))
            story.append(Paragraph(evidence_name, heading_style))
            story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
            story.append(Spacer(1, 0.3*inch))
            
            # Add policy content
            for content_type, text in policy_content:
                if content_type == 'heading':
                    story.append(Paragraph(text, title_style))
                elif content_type == 'subheading':
                    story.append(Paragraph(text, heading_style))
                else:
                    # Clean up text for PDF
                    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    story.append(Paragraph(text, body_style))
            
            # Build PDF
            pdf.build(story)
            return file_path
            
        except Exception as e:
            print(f"Error generating ISMS policy PDF: {e}")
            return None
    
    def generate_employee_handbook_pdf(self, section_name):
        """Generate PDF for Employee Handbook section"""
        try:
            import PyPDF2
            handbook_path = '/home/webuser/NEW Cirque_Corporation_Employee_Handbook_1-2022 2.pdf'
            
            if not os.path.exists(handbook_path):
                print(f"Employee Handbook not found at {handbook_path}")
                return None
            
            # Extract the section number (e.g., "1-6" from "1-6. Non-Disclosure...")
            import re
            section_match = re.match(r'^(\d+-\d+)\.?\s*(.+)', section_name)
            if not section_match:
                print(f"Could not parse section number from: {section_name}")
                return None
            
            section_num = section_match.group(1)
            section_title = section_match.group(2).strip()
            
            # Read the PDF
            with open(handbook_path, 'rb') as file:
                pdf = PyPDF2.PdfReader(file)
                
                # Extract all text to find the section
                full_text = ""
                for page in pdf.pages:
                    full_text += page.extract_text() + "\n"
                
                # Find the section content
                lines = full_text.split('\n')
                section_content = []
                in_section = False
                
                for line in lines:
                    line_stripped = line.strip()
                    
                    # Start capturing when we find the section
                    if section_num in line_stripped and (section_title.lower() in line_stripped.lower() or
                                                         line_stripped.startswith(section_num)):
                        in_section = True
                        section_content.append(('heading', line_stripped))
                        continue
                    
                    # Stop at next section
                    if in_section and re.match(r'^\d+-\d+\.', line_stripped) and len(section_content) > 5:
                        break
                    
                    # Capture content
                    if in_section and line_stripped:
                        # Skip page headers/footers
                        if 'Employee Handbook' in line_stripped or 'Copyright, Cirque' in line_stripped:
                            continue
                        if re.match(r'^\d+$', line_stripped):  # Just a page number
                            continue
                        section_content.append(('body', line_stripped))
                
                if not section_content:
                    print(f"Section content not found for: {section_name}")
                    return None
                
                # Generate PDF
                file_path = self.get_file_path(section_title, 'Employee_Handbook', 'pdf')
                pdf_doc = SimpleDocTemplate(file_path, pagesize=letter)
                story = []
                
                # Styles
                styles = getSampleStyleSheet()
                title_style = ParagraphStyle(
                    'CustomTitle',
                    parent=styles['Heading1'],
                    fontSize=18,
                    textColor='#2D4639',
                    spaceAfter=30,
                    alignment=TA_CENTER
                )
                heading_style = ParagraphStyle(
                    'CustomHeading',
                    parent=styles['Heading2'],
                    fontSize=14,
                    textColor='#2D4639',
                    spaceAfter=12,
                    spaceBefore=12
                )
                body_style = ParagraphStyle(
                    'CustomBody',
                    parent=styles['BodyText'],
                    fontSize=11,
                    alignment=TA_JUSTIFY,
                    spaceAfter=12
                )
                
                # Add header
                story.append(Paragraph("Cirque Corporation", title_style))
                story.append(Paragraph("Employee Handbook - Extract", heading_style))
                story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y')}", styles['Normal']))
                story.append(Spacer(1, 0.3*inch))
                
                # Add section content
                for content_type, text in section_content:
                    # Clean up text for PDF
                    text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    if content_type == 'heading':
                        story.append(Paragraph(text, heading_style))
                    else:
                        story.append(Paragraph(text, body_style))
                
                # Build PDF
                pdf_doc.build(story)
                return file_path
                
        except Exception as e:
            print(f"Error generating Employee Handbook PDF: {e}")
            return None
    
    def generate_teamviewer_patch_scan_file(self, evidence_name):
        """Generate TeamViewer patch scan/device status file"""
        try:
            tv_service = TeamViewerEvidenceService()
            report = tv_service.generate_patch_report_data()
            
            wb, ws = self.create_styled_workbook('Patch Scan')
            headers = ['Device Name', 'OS Version', 'Online', 'Managed', 'Policy', 'Last Seen', 'Days Offline', 'Asset Tracked']
            self.style_header_row(ws, headers)
            
            for row_idx, device in enumerate(report['devices'], 2):
                ws.cell(row_idx, 1, device['device_name'])
                ws.cell(row_idx, 2, device['os_version'])
                ws.cell(row_idx, 3, device['online_state'])
                ws.cell(row_idx, 4, 'Yes' if device['managed'] else 'No')
                ws.cell(row_idx, 5, 'Yes' if device['policy_applied'] else 'No')
                ws.cell(row_idx, 6, device['last_seen'] or 'Never')
                ws.cell(row_idx, 7, device['days_offline'] if device['days_offline'] is not None else 'N/A')
                ws.cell(row_idx, 8, 'Yes' if device['asset_tracked'] else 'No')
            
            # Add summary sheet
            ws_summary = wb.create_sheet('Summary')
            summary_data = [
                ('Report Date', report['report_date']),
                ('Total Devices', report['total_devices']),
                ('Online Devices', report['online_devices']),
                ('Managed Devices', report['managed_devices']),
                ('Devices with Policy', report['devices_with_policy']),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Intune')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating TeamViewer patch scan: {e}")
            return None
    
    def generate_teamviewer_vulnerability_scan_file(self, evidence_name):
        """Generate TeamViewer/Intune vulnerability scan file"""
        try:
            tv_service = TeamViewerEvidenceService()
            vulnerabilities = tv_service.get_vulnerability_summary()
            
            wb, ws = self.create_styled_workbook('Vulnerability Scan')
            headers = ['Device Name', 'OS Version', 'Severity', 'Finding', 'Remediation', 'Last Seen']
            self.style_header_row(ws, headers)
            
            for row_idx, vuln in enumerate(vulnerabilities, 2):
                ws.cell(row_idx, 1, vuln['device_name'])
                ws.cell(row_idx, 2, vuln['os_version'])
                ws.cell(row_idx, 3, vuln['severity'])
                ws.cell(row_idx, 4, vuln['finding'])
                ws.cell(row_idx, 5, vuln['remediation'])
                ws.cell(row_idx, 6, vuln['last_seen'] or 'Never')
            
            # Add summary information
            from soc2_models import IntuneDevice
            total_devices = IntuneDevice.query.filter_by(is_current=True).count()
            critical = sum(1 for v in vulnerabilities if v['severity'] == 'Critical')
            high = sum(1 for v in vulnerabilities if v['severity'] == 'High')
            medium = sum(1 for v in vulnerabilities if v['severity'] == 'Medium')
            
            ws_summary = wb.create_sheet('Summary')
            summary_data = [
                ('Scan Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Devices Scanned', total_devices),
                ('Devices with Findings', len([v for v in vulnerabilities if v['severity'] != 'Info'])),
                ('Critical Findings', critical),
                ('High Findings', high),
                ('Medium Findings', medium),
                ('Scan Type', 'Intune Compliance + OS Version Analysis'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Intune')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating TeamViewer vulnerability scan: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_defender_vulnerability_remediation_file(self, evidence_name):
        """Generate Microsoft Defender vulnerability remediation action plan"""
        try:
            defender_service = DefenderService()
            recommendations = defender_service.get_recommendations()
            machines = defender_service.get_machines()
            
            # Filter for vulnerability-related recommendations
            vuln_recommendations = [
                r for r in recommendations 
                if any(keyword in r.get('recommendationName', '').lower() 
                       for keyword in ['vulnerability', 'cve', 'security update', 'exploit'])
            ]
            
            wb, ws = self.create_styled_workbook('Remediation Actions')
            headers = ['Priority', 'Recommendation', 'Product', 'Severity', 'Affected Machines', 
                      'Status', 'Action Required', 'Remediation Type', 'Due Date']
            self.style_header_row(ws, headers)
            
            # Sort by severity and exposed machines
            severity_order = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3}
            vuln_recommendations.sort(
                key=lambda x: (severity_order.get(x.get('severity', 'Low'), 4), 
                              -x.get('exposedMachinesCount', 0))
            )
            
            for row_idx, rec in enumerate(vuln_recommendations, 2):
                severity = rec.get('severity', 'Low')
                exposed = rec.get('exposedMachinesCount', 0)
                
                # Determine priority based on severity and exposure
                if severity == 'Critical' and exposed > 10:
                    priority = 'P1 - Immediate'
                elif severity in ['Critical', 'High'] and exposed > 5:
                    priority = 'P2 - High'
                elif severity in ['Critical', 'High']:
                    priority = 'P3 - Medium'
                else:
                    priority = 'P4 - Low'
                
                # Calculate due date based on priority
                from datetime import timedelta
                due_days = {'P1 - Immediate': 7, 'P2 - High': 30, 'P3 - Medium': 60, 'P4 - Low': 90}
                due_date = (datetime.utcnow() + timedelta(days=due_days[priority])).strftime('%Y-%m-%d')
                
                ws.cell(row_idx, 1, priority)
                ws.cell(row_idx, 2, rec.get('recommendationName', '')[:100])
                ws.cell(row_idx, 3, rec.get('productName', 'Unknown'))
                ws.cell(row_idx, 4, severity)
                ws.cell(row_idx, 5, exposed)
                ws.cell(row_idx, 6, rec.get('status', 'Active'))
                ws.cell(row_idx, 7, rec.get('recommendedAction', 'Apply security update')[:100])
                ws.cell(row_idx, 8, rec.get('remediationType', 'Update'))
                ws.cell(row_idx, 9, due_date)
            
            # Add summary sheet
            ws_summary = wb.create_sheet('Remediation Summary', 0)
            
            p1_count = sum(1 for r in vuln_recommendations 
                          if r.get('severity') == 'Critical' and r.get('exposedMachinesCount', 0) > 10)
            p2_count = sum(1 for r in vuln_recommendations 
                          if r.get('severity') in ['Critical', 'High'] and r.get('exposedMachinesCount', 0) > 5)
            total_machines_affected = len(set(m['id'] for m in machines 
                                             if m.get('riskScore') in ['High', 'Medium']))
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Remediation Items', len(vuln_recommendations)),
                ('P1 - Immediate (7 days)', p1_count),
                ('P2 - High (30 days)', p2_count),
                ('P3 - Medium (60 days)', len(vuln_recommendations) - p1_count - p2_count),
                ('Total Machines Requiring Action', total_machines_affected),
                ('Data Source', 'Microsoft Defender for Endpoint'),
                ('Remediation Focus', 'Vulnerability & Security Updates'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Defender vulnerability remediation: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_defender_vulnerability_scan_file(self, evidence_name):
        """Generate Microsoft Defender vulnerability scan file with real CVE data"""
        try:
            defender_service = DefenderService()
            vulnerabilities = defender_service.generate_vulnerability_report_data()
            
            wb, ws = self.create_styled_workbook('Vulnerabilities')
            headers = ['CVE/Vulnerability ID', 'Name', 'Severity', 'CVSS Score', 'Exposed Machines', 
                      'Description', 'Published Date', 'Updated Date']
            self.style_header_row(ws, headers)
            
            for row_idx, vuln in enumerate(vulnerabilities, 2):
                ws.cell(row_idx, 1, vuln['vulnerability_id'])
                ws.cell(row_idx, 2, vuln['name'][:100])  # Truncate long names
                ws.cell(row_idx, 3, vuln['severity'])
                ws.cell(row_idx, 4, vuln['cvss_score'])
                ws.cell(row_idx, 5, vuln['exposed_machines'])
                ws.cell(row_idx, 6, vuln['description'][:200])  # Truncate description
                ws.cell(row_idx, 7, vuln['published_date'][:10] if vuln['published_date'] else '')
                ws.cell(row_idx, 8, vuln['updated_date'][:10] if vuln['updated_date'] else '')
            
            # Add summary sheet
            summary = defender_service.get_vulnerability_summary()
            if summary:
                ws_summary = wb.create_sheet('Summary', 0)  # Insert at beginning
                summary_data = [
                    ('Scan Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                    ('Total Vulnerabilities', summary['total_vulnerabilities']),
                    ('Total Machines', summary['total_machines']),
                    ('Critical Vulnerabilities', summary['by_severity']['Critical']),
                    ('High Vulnerabilities', summary['by_severity']['High']),
                    ('Medium Vulnerabilities', summary['by_severity']['Medium']),
                    ('Low Vulnerabilities', summary['by_severity']['Low']),
                    ('Data Source', 'Microsoft Defender for Endpoint'),
                ]
                
                for row_idx, (label, value) in enumerate(summary_data, 1):
                    ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                    ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Defender vulnerability scan: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_defender_software_inventory_file(self, evidence_name):
        """Generate Microsoft Defender software inventory report"""
        try:
            defender_service = DefenderService()
            software_data = defender_service.generate_software_inventory_report_data()
            
            wb, ws = self.create_styled_workbook('Software Inventory')
            headers = ['Vendor', 'Product', 'Version', 'Installed On', 'Has Vulnerabilities',
                      'Critical Vulns', 'High Vulns', 'Total Vulns', 'Category', 'Support Status']
            self.style_header_row(ws, headers)
            
            for row_idx, sw in enumerate(software_data, 2):
                # Sanitize strings for Excel (remove illegal characters)
                vendor = self._sanitize_for_excel(sw['vendor'])
                product = self._sanitize_for_excel(sw['product'][:100])
                version = self._sanitize_for_excel(sw['version'])
                category = self._sanitize_for_excel(sw['category'])
                eol = self._sanitize_for_excel(sw['end_of_life'])
                
                ws.cell(row_idx, 1, vendor)
                ws.cell(row_idx, 2, product)
                ws.cell(row_idx, 3, version)
                ws.cell(row_idx, 4, sw['installed_on'])
                ws.cell(row_idx, 5, sw['has_vulnerabilities'])
                ws.cell(row_idx, 6, sw['critical_vulns'])
                ws.cell(row_idx, 7, sw['high_vulns'])
                ws.cell(row_idx, 8, sw['total_vulns'])
                ws.cell(row_idx, 9, category)
                ws.cell(row_idx, 10, eol)
            
            # Add summary sheet
            ws_summary = wb.create_sheet('Summary', 0)
            
            total_software = len(software_data)
            software_with_vulns = sum(1 for sw in software_data if sw['has_vulnerabilities'] == 'Yes')
            total_critical = sum(sw['critical_vulns'] for sw in software_data)
            total_high = sum(sw['high_vulns'] for sw in software_data)
            end_of_life = sum(1 for sw in software_data if sw['end_of_life'] != 'Supported')
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Software Products', total_software),
                ('Software with Known Vulnerabilities', software_with_vulns),
                ('Total Critical Vulnerabilities', total_critical),
                ('Total High Vulnerabilities', total_high),
                ('End-of-Life Software', end_of_life),
                ('Data Source', 'Microsoft Defender for Endpoint'),
                ('Inventory Type', 'Organization-wide Software Catalog'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Defender software inventory: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_defender_patch_scan_file(self, evidence_name):
        """Generate Microsoft Defender missing patch report"""
        try:
            defender_service = DefenderService()
            patches = defender_service.generate_patch_report_data()
            
            wb, ws = self.create_styled_workbook('Missing Patches')
            headers = ['Recommendation ID', 'Name', 'Product', 'Severity', 'Exposed Machines', 
                      'Status', 'Remediation Type', 'Recommended Update']
            self.style_header_row(ws, headers)
            
            for row_idx, patch in enumerate(patches, 2):
                ws.cell(row_idx, 1, patch['recommendation_id'])
                ws.cell(row_idx, 2, patch['name'][:100])
                ws.cell(row_idx, 3, patch['product'])
                ws.cell(row_idx, 4, patch['severity'])
                ws.cell(row_idx, 5, patch['exposed_machines'])
                ws.cell(row_idx, 6, patch['status'])
                ws.cell(row_idx, 7, patch['remediation_type'])
                ws.cell(row_idx, 8, patch['recommendation'][:100] if patch['recommendation'] else '')
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            critical = sum(1 for p in patches if p['severity'] == 'Critical')
            high = sum(1 for p in patches if p['severity'] == 'High')
            total_exposed = sum(p['exposed_machines'] for p in patches)
            
            summary_data = [
                ('Scan Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Recommendations', len(patches)),
                ('Critical', critical),
                ('High', high),
                ('Total Exposed Machines', total_exposed),
                ('Data Source', 'Microsoft Defender for Endpoint'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Defender patch scan: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_mfa_status_file(self, evidence_name):
        """Generate MFA status report for all users"""
        try:
            from m365_service import M365Service
            from app import Setting
            
            from m365_config import get_m365_credentials
            tenant_id, client_id, client_secret = get_m365_credentials()

            m365_service = M365Service(tenant_id, client_id, client_secret)
            mfa_data = m365_service.get_users_mfa_status()
            
            wb, ws = self.create_styled_workbook('MFA Status')
            headers = ['Display Name', 'User Principal Name', 'Email', 'MFA Enabled', 'MFA Methods', 'Method Count']
            self.style_header_row(ws, headers)
            
            for row_idx, user in enumerate(mfa_data, 2):
                ws.cell(row_idx, 1, user['displayName'])
                ws.cell(row_idx, 2, user['userPrincipalName'])
                ws.cell(row_idx, 3, user.get('mail', ''))
                ws.cell(row_idx, 4, 'Yes' if user['mfaEnabled'] else 'No')
                ws.cell(row_idx, 5, user['mfaMethods'])
                ws.cell(row_idx, 6, user['methodCount'])
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            total_users = len(mfa_data)
            mfa_enabled = sum(1 for u in mfa_data if u['mfaEnabled'])
            mfa_disabled = total_users - mfa_enabled
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Users', total_users),
                ('MFA Enabled', mfa_enabled),
                ('MFA Disabled', mfa_disabled),
                ('Compliance Rate', f"{(mfa_enabled/total_users*100):.1f}%" if total_users > 0 else '0%'),
                ('Data Source', 'Microsoft Graph API'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating MFA status: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_security_incidents_file(self, evidence_name):
        """Generate security incidents report from Defender"""
        try:
            defender_service = DefenderService()
            incidents = defender_service.get_incidents()
            
            wb, ws = self.create_styled_workbook('Security Incidents')
            headers = ['Incident ID', 'Title', 'Severity', 'Status', 'Classification', 
                      'Assigned To', 'Created', 'Last Updated', 'Alert Count', 'Affected Devices']
            self.style_header_row(ws, headers)
            
            for row_idx, incident in enumerate(incidents, 2):
                ws.cell(row_idx, 1, str(incident.get('incidentId', '')))
                ws.cell(row_idx, 2, self._sanitize_for_excel(incident.get('incidentName', '')[:100]))
                ws.cell(row_idx, 3, incident.get('severity', 'Unknown'))
                ws.cell(row_idx, 4, incident.get('status', 'Unknown'))
                ws.cell(row_idx, 5, incident.get('classification', 'Unknown'))
                ws.cell(row_idx, 6, incident.get('assignedTo', 'Unassigned'))
                ws.cell(row_idx, 7, incident.get('createdTime', '')[:19] if incident.get('createdTime') else '')
                ws.cell(row_idx, 8, incident.get('lastUpdateTime', '')[:19] if incident.get('lastUpdateTime') else '')
                ws.cell(row_idx, 9, len(incident.get('alerts', [])))
                ws.cell(row_idx, 10, len(incident.get('devices', [])))
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            total = len(incidents)
            by_severity = {}
            by_status = {}
            
            for incident in incidents:
                severity = incident.get('severity', 'Unknown')
                status = incident.get('status', 'Unknown')
                by_severity[severity] = by_severity.get(severity, 0) + 1
                by_status[status] = by_status.get(status, 0) + 1
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Incidents', total),
                ('Critical', by_severity.get('High', 0)),
                ('High', by_severity.get('Medium', 0)),
                ('Medium', by_severity.get('Low', 0)),
                ('Active', by_status.get('Active', 0)),
                ('Resolved', by_status.get('Resolved', 0)),
                ('Data Source', 'Microsoft Defender for Endpoint'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating security incidents: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_security_alerts_file(self, evidence_name):
        """Generate security alerts report from Defender"""
        try:
            defender_service = DefenderService()
            alerts = defender_service.get_alerts()
            
            wb, ws = self.create_styled_workbook('Security Alerts')
            headers = ['Alert ID', 'Title', 'Category', 'Severity', 'Status', 'Machine', 
                      'Detection Time', 'First Activity', 'Last Activity', 'Assigned To']
            self.style_header_row(ws, headers)
            
            for row_idx, alert in enumerate(alerts, 2):
                ws.cell(row_idx, 1, alert.get('id', '')[:50])
                ws.cell(row_idx, 2, self._sanitize_for_excel(alert.get('title', '')[:100]))
                ws.cell(row_idx, 3, alert.get('category', 'Unknown'))
                ws.cell(row_idx, 4, alert.get('severity', 'Unknown'))
                ws.cell(row_idx, 5, alert.get('status', 'Unknown'))
                ws.cell(row_idx, 6, alert.get('machineId', 'Unknown')[:30])
                ws.cell(row_idx, 7, alert.get('alertCreationTime', '')[:19] if alert.get('alertCreationTime') else '')
                ws.cell(row_idx, 8, alert.get('firstEventTime', '')[:19] if alert.get('firstEventTime') else '')
                ws.cell(row_idx, 9, alert.get('lastEventTime', '')[:19] if alert.get('lastEventTime') else '')
                ws.cell(row_idx, 10, alert.get('assignedTo', 'Unassigned'))
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            total = len(alerts)
            by_severity = {}
            by_category = {}
            
            for alert in alerts:
                severity = alert.get('severity', 'Unknown')
                category = alert.get('category', 'Unknown')
                by_severity[severity] = by_severity.get(severity, 0) + 1
                by_category[category] = by_category.get(category, 0) + 1
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Alerts', total),
                ('High Severity', by_severity.get('High', 0)),
                ('Medium Severity', by_severity.get('Medium', 0)),
                ('Low Severity', by_severity.get('Low', 0)),
                ('Informational', by_severity.get('Informational', 0)),
                ('Top Category', max(by_category.items(), key=lambda x: x[1])[0] if by_category else 'None'),
                ('Data Source', 'Microsoft Defender for Endpoint'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating security alerts: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_azure_rbac_file(self, evidence_name):
        """Generate Azure RBAC role assignments report"""
        try:
            from azure_security_service import AzureSecurityService
            
            azure_service = AzureSecurityService()
            assignments = azure_service.get_role_assignments()
            
            wb, ws = self.create_styled_workbook('Azure RBAC Assignments')
            headers = ['Principal ID', 'Principal Type', 'Role Name', 'Scope', 'Created On', 'Created By']
            self.style_header_row(ws, headers)
            
            for row_idx, assignment in enumerate(assignments, 2):
                ws.cell(row_idx, 1, assignment.get('principalId', '')[:50])
                ws.cell(row_idx, 2, assignment.get('principalType', 'Unknown'))
                ws.cell(row_idx, 3, assignment.get('roleName', 'Unknown'))
                scope = assignment.get('scope', '')
                # Simplify scope display
                if '/resourceGroups/' in scope:
                    scope_display = scope.split('/resourceGroups/')[-1]
                elif '/subscriptions/' in scope:
                    scope_display = 'Subscription'
                else:
                    scope_display = scope[-50:]
                ws.cell(row_idx, 4, scope_display)
                ws.cell(row_idx, 5, assignment.get('createdOn', '')[:19] if assignment.get('createdOn') else '')
                ws.cell(row_idx, 6, assignment.get('createdBy', 'Unknown')[:50])
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            total = len(assignments)
            by_type = {}
            by_role = {}
            
            for assignment in assignments:
                ptype = assignment.get('principalType', 'Unknown')
                role = assignment.get('roleName', 'Unknown')
                by_type[ptype] = by_type.get(ptype, 0) + 1
                by_role[role] = by_role.get(role, 0) + 1
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Role Assignments', total),
                ('User Assignments', by_type.get('User', 0)),
                ('Service Principal Assignments', by_type.get('ServicePrincipal', 0)),
                ('Group Assignments', by_type.get('Group', 0)),
                ('Top Role', max(by_role.items(), key=lambda x: x[1])[0] if by_role else 'None'),
                ('Data Source', 'Azure Resource Manager API'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'Azure')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Azure RBAC report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_conditional_access_file(self, evidence_name):
        """Generate Conditional Access policies report"""
        try:
            from m365_service import M365Service
            from app import Setting
            
            from m365_config import get_m365_credentials
            tenant_id, client_id, client_secret = get_m365_credentials()

            m365_service = M365Service(tenant_id, client_id, client_secret)
            policies = m365_service.get_conditional_access_policies()
            
            wb, ws = self.create_styled_workbook('Conditional Access Policies')
            headers = ['Policy Name', 'State', 'Created', 'Modified', 'Users/Groups', 'Applications', 'Grant Controls']
            self.style_header_row(ws, headers)
            
            for row_idx, policy in enumerate(policies, 2):
                ws.cell(row_idx, 1, self._sanitize_for_excel(policy.get('displayName', 'Unknown')[:100]))
                ws.cell(row_idx, 2, policy.get('state', 'Unknown'))
                ws.cell(row_idx, 3, policy.get('createdDateTime', '')[:19] if policy.get('createdDateTime') else '')
                ws.cell(row_idx, 4, policy.get('modifiedDateTime', '')[:19] if policy.get('modifiedDateTime') else '')
                
                conditions = policy.get('conditions', {})
                users = conditions.get('users', {})
                user_count = len(users.get('includeUsers', [])) + len(users.get('includeGroups', []))
                ws.cell(row_idx, 5, f"{user_count} users/groups")
                
                apps = conditions.get('applications', {})
                app_count = len(apps.get('includeApplications', []))
                ws.cell(row_idx, 6, f"{app_count} applications")
                
                grant_controls = policy.get('grantControls', {})
                built_in_controls = grant_controls.get('builtInControls', [])
                ws.cell(row_idx, 7, ', '.join(built_in_controls[:3]) if built_in_controls else 'None')
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            total = len(policies)
            enabled = sum(1 for p in policies if p.get('state') == 'enabled')
            disabled = sum(1 for p in policies if p.get('state') == 'disabled')
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Policies', total),
                ('Enabled', enabled),
                ('Disabled/Report-Only', disabled),
                ('Data Source', 'Microsoft Graph API'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating conditional access report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_software_inventory_by_asset_file(self, evidence_name):
        """Generate software inventory organized by asset"""
        try:
            defender_service = DefenderService()
            software_data = defender_service.get_software_by_machine()
            
            wb, ws = self.create_styled_workbook('Software Inventory by Asset')
            headers = ['Machine Name', 'Machine ID', 'OS Platform', 'Software Name', 
                      'Vendor', 'Version', 'Installed Count']
            self.style_header_row(ws, headers)
            
            for row_idx, item in enumerate(software_data, 2):
                ws.cell(row_idx, 1, item.get('machineName', 'Unknown'))
                ws.cell(row_idx, 2, item.get('machineId', '')[:30])
                ws.cell(row_idx, 3, item.get('osPlatform', 'Unknown'))
                ws.cell(row_idx, 4, self._sanitize_for_excel(item.get('softwareName', 'Unknown')[:100]))
                ws.cell(row_idx, 5, item.get('softwareVendor', 'Unknown')[:50])
                ws.cell(row_idx, 6, item.get('softwareVersion', 'Unknown')[:30])
                ws.cell(row_idx, 7, item.get('installedMachines', 0))
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            machines = set(item.get('machineName') for item in software_data)
            software_titles = set(item.get('softwareName') for item in software_data)
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Machines', len(machines)),
                ('Total Software Titles', len(software_titles)),
                ('Total Installations', len(software_data)),
                ('Data Source', 'Microsoft Defender for Endpoint'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating software inventory by asset: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_system_updates_file(self, evidence_name):
        """Generate system updates and missing hotfixes report"""
        try:
            defender_service = DefenderService()
            missing_updates = defender_service.get_missing_kbs()
            
            wb, ws = self.create_styled_workbook('System Updates & Hotfixes')
            headers = ['Machine Name', 'OS Platform', 'OS Version', 'Recommendation', 
                      'Product', 'Severity', 'Exposed Machines', 'Component']
            self.style_header_row(ws, headers)
            
            for row_idx, update in enumerate(missing_updates, 2):
                ws.cell(row_idx, 1, update.get('machineName', 'Unknown'))
                ws.cell(row_idx, 2, update.get('osPlatform', 'Unknown'))
                ws.cell(row_idx, 3, update.get('osVersion', 'Unknown')[:50])
                ws.cell(row_idx, 4, self._sanitize_for_excel(update.get('recommendationName', '')[:100]))
                ws.cell(row_idx, 5, update.get('productName', 'Unknown')[:50])
                ws.cell(row_idx, 6, update.get('severity', 'Unknown'))
                ws.cell(row_idx, 7, update.get('exposedMachines', 0))
                ws.cell(row_idx, 8, update.get('relatedComponent', 'Unknown')[:50])
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            machines = set(u.get('machineName') for u in missing_updates)
            by_severity = {}
            for update in missing_updates:
                sev = update.get('severity', 'Unknown')
                by_severity[sev] = by_severity.get(sev, 0) + 1
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Machines with Missing Updates', len(machines)),
                ('Total Missing Updates', len(missing_updates)),
                ('Critical', by_severity.get('Critical', 0)),
                ('High', by_severity.get('High', 0)),
                ('Medium', by_severity.get('Medium', 0)),
                ('Low', by_severity.get('Low', 0)),
                ('Data Source', 'Microsoft Defender for Endpoint'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'M365/Defender')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating system updates report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_security_baseline_file(self, evidence_name):
        """Generate security baseline compliance report (Secure Score)"""
        try:
            from azure_security_service import AzureSecurityService
            
            azure_service = AzureSecurityService()
            secure_score = azure_service.get_secure_score()
            
            wb, ws = self.create_styled_workbook('Security Baseline Compliance')
            headers = ['Control Name', 'Current Score', 'Max Score', 'Healthy Resources', 
                      'Unhealthy Resources', 'Not Applicable']
            self.style_header_row(ws, headers)
            
            for row_idx, control in enumerate(secure_score.get('controls', []), 2):
                ws.cell(row_idx, 1, control.get('displayName', 'Unknown')[:100])
                ws.cell(row_idx, 2, control.get('score', 0))
                ws.cell(row_idx, 3, control.get('maxScore', 0))
                ws.cell(row_idx, 4, control.get('healthyResources', 0))
                ws.cell(row_idx, 5, control.get('unhealthyResources', 0))
                ws.cell(row_idx, 6, control.get('notApplicableResources', 0))
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            controls = secure_score.get('controls', [])
            total_healthy = sum(c.get('healthyResources', 0) for c in controls)
            total_unhealthy = sum(c.get('unhealthyResources', 0) for c in controls)
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Overall Secure Score', f"{secure_score.get('score', 0)}/{secure_score.get('maxScore', 0)}"),
                ('Compliance Percentage', f"{secure_score.get('percentage', 0):.1f}%"),
                ('Total Controls', len(controls)),
                ('Healthy Resources', total_healthy),
                ('Unhealthy Resources', total_unhealthy),
                ('Data Source', 'Microsoft Defender for Cloud'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'Azure')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating security baseline report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_key_vault_policies_file(self, evidence_name):
        """Generate Key Vault access policies report"""
        try:
            from azure_security_service import AzureSecurityService
            
            azure_service = AzureSecurityService()
            policies = azure_service.get_key_vault_access_policies()
            
            wb, ws = self.create_styled_workbook('Key Vault Access Policies')
            headers = ['Vault Name', 'Resource Group', 'Object ID', 'Application ID', 
                      'Key Permissions', 'Secret Permissions', 'Certificate Permissions']
            self.style_header_row(ws, headers)
            
            for row_idx, policy in enumerate(policies, 2):
                ws.cell(row_idx, 1, policy.get('vaultName', 'Unknown'))
                ws.cell(row_idx, 2, policy.get('vaultResourceGroup', 'Unknown'))
                ws.cell(row_idx, 3, policy.get('objectId', 'Unknown')[:50])
                ws.cell(row_idx, 4, policy.get('applicationId', '')[:50])
                perms = policy.get('permissions', {})
                ws.cell(row_idx, 5, perms.get('keys', 'None')[:50])
                ws.cell(row_idx, 6, perms.get('secrets', 'None')[:50])
                ws.cell(row_idx, 7, perms.get('certificates', 'None')[:50])
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            vaults = set(p.get('vaultName') for p in policies)
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Key Vaults', len(vaults)),
                ('Total Access Policies', len(policies)),
                ('Data Source', 'Azure Key Vault'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'Azure')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating Key Vault policies report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def generate_network_traffic_logs_file(self, evidence_name):
        """Generate NSG Flow Logs configuration report"""
        try:
            from azure_security_service import AzureSecurityService
            
            azure_service = AzureSecurityService()
            flow_logs = azure_service.get_nsg_flow_logs()
            
            wb, ws = self.create_styled_workbook('Network Traffic Logs')
            headers = ['Flow Log Name', 'Location', 'Target Resource', 'Storage Account', 
                      'Enabled', 'Retention Days', 'Format', 'Version']
            self.style_header_row(ws, headers)
            
            for row_idx, log in enumerate(flow_logs, 2):
                ws.cell(row_idx, 1, log.get('name', 'Unknown'))
                ws.cell(row_idx, 2, log.get('location', 'Unknown'))
                target = log.get('targetResourceId', '')
                # Simplify resource ID display
                target_display = target.split('/')[-1] if '/' in target else target[:50]
                ws.cell(row_idx, 3, target_display)
                storage = log.get('storageId', '')
                storage_display = storage.split('/')[-1] if '/' in storage else storage[:50]
                ws.cell(row_idx, 4, storage_display)
                ws.cell(row_idx, 5, 'Yes' if log.get('enabled') else 'No')
                ws.cell(row_idx, 6, log.get('retentionDays', 0))
                ws.cell(row_idx, 7, log.get('format', 'Unknown'))
                ws.cell(row_idx, 8, log.get('version', 0))
            
            # Add summary
            ws_summary = wb.create_sheet('Summary', 0)
            enabled = sum(1 for log in flow_logs if log.get('enabled'))
            disabled = len(flow_logs) - enabled
            
            summary_data = [
                ('Report Date', datetime.utcnow().strftime('%Y-%m-%d %H:%M')),
                ('Total Flow Logs', len(flow_logs)),
                ('Enabled', enabled),
                ('Disabled', disabled),
                ('Data Source', 'Azure Network Watcher'),
            ]
            
            for row_idx, (label, value) in enumerate(summary_data, 1):
                ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
                ws_summary.cell(row_idx, 2, value)
            
            file_path = self.get_file_path(evidence_name, 'Azure')
            wb.save(file_path)
            return file_path
        except Exception as e:
            print(f"Error generating network traffic logs report: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    # ------------------------------------------------------------------
    # RMM-backed generators (local data: vulnerability_cache, cve_patch_job,
    # rmm_agent, rmm_telemetry, asset). These do not depend on any external
    # API so they always produce in-period data while the snapshots are fresh.
    # ------------------------------------------------------------------

    def _summary_sheet(self, wb, rows):
        ws_summary = wb.create_sheet('Summary', 0)
        for row_idx, (label, value) in enumerate(rows, 1):
            ws_summary.cell(row_idx, 1, label).font = Font(bold=True)
            ws_summary.cell(row_idx, 2, value)
        ws_summary.column_dimensions['A'].width = 32
        ws_summary.column_dimensions['B'].width = 40
        return ws_summary

    def generate_rmm_vulnerability_scan_file(self, evidence_name):
        """Vulnerability scan results from the live RMM vulnerability_cache.

        Source: vulnerability_cache (CVE catalog synced from the RMM scanner),
        ordered by severity then CVSS. Includes a dated summary sheet.
        """
        sev_rank = {'Critical': 0, 'High': 1, 'Medium': 2, 'Low': 3, 'Unknown': 4}
        rows = db.session.execute(text(
            """SELECT cve_id, name, severity, cvss, exposed_machines, description,
                      published_on, synced_at
               FROM vulnerability_cache"""
        )).fetchall()
        rows = sorted(rows, key=lambda r: (sev_rank.get(r[2], 5), -(r[3] or 0)))

        wb, ws = self.create_styled_workbook('Vulnerability Scan')
        headers = ['CVE ID', 'Name', 'Severity', 'CVSS', 'Exposed Machines',
                   'Description', 'Published', 'Last Synced']
        self.style_header_row(ws, headers)
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row_idx, 1, r[0] or '')
            ws.cell(row_idx, 2, self._sanitize_for_excel((r[1] or '')[:120]))
            ws.cell(row_idx, 3, r[2] or 'Unknown')
            ws.cell(row_idx, 4, r[3] if r[3] is not None else '')
            ws.cell(row_idx, 5, r[4] if r[4] is not None else 0)
            ws.cell(row_idx, 6, self._sanitize_for_excel((r[5] or '')[:300]))
            ws.cell(row_idx, 7, str(r[6])[:10] if r[6] else '')
            ws.cell(row_idx, 8, r[7].strftime('%Y-%m-%d %H:%M') if r[7] else '')

        by_sev = {}
        latest_sync = None
        for r in rows:
            by_sev[r[2] or 'Unknown'] = by_sev.get(r[2] or 'Unknown', 0) + 1
            if r[7] and (latest_sync is None or r[7] > latest_sync):
                latest_sync = r[7]
        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Last Scanner Sync', latest_sync.strftime('%Y-%m-%d %H:%M') if latest_sync else 'Unknown'),
            ('Total Vulnerabilities', len(rows)),
            ('Critical', by_sev.get('Critical', 0)),
            ('High', by_sev.get('High', 0)),
            ('Medium', by_sev.get('Medium', 0)),
            ('Low', by_sev.get('Low', 0)),
            ('Data Source', 'RMM vulnerability scanner (vulnerability_cache)'),
        ])
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_rmm_vulnerability_remediation_file(self, evidence_name):
        """Vulnerability remediation evidence from cve_patch_job joined to the
        CVE catalog and asset register. Shows what was remediated, on which
        device, and the current state (installed/superseded/no_patch).
        """
        rows = db.session.execute(text(
            """SELECT j.agent_id, COALESCE(a.name, j.agent_id) AS device,
                      j.cve_id, v.severity, v.cvss, j.status,
                      j.updates_found, j.reboot_required,
                      COALESCE(j.completed_at, j.updated_at, j.created_at) AS action_at,
                      j.deployed_by
               FROM cve_patch_job j
               LEFT JOIN asset a ON a.id = j.asset_id
               LEFT JOIN vulnerability_cache v ON v.cve_id = j.cve_id
               WHERE j.status IN ('installed', 'superseded')
               ORDER BY action_at DESC NULLS LAST"""
        )).fetchall()

        wb, ws = self.create_styled_workbook('Vuln Remediation')
        headers = ['Device', 'Agent ID', 'CVE ID', 'Severity', 'CVSS', 'Remediation Status',
                   'Updates Found', 'Reboot Required', 'Remediated At', 'Deployed By']
        self.style_header_row(ws, headers)
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row_idx, 1, r[1] or '')
            ws.cell(row_idx, 2, r[0] or '')
            ws.cell(row_idx, 3, r[2] or '')
            ws.cell(row_idx, 4, r[3] or 'Unknown')
            ws.cell(row_idx, 5, r[4] if r[4] is not None else '')
            ws.cell(row_idx, 6, r[5] or '')
            ws.cell(row_idx, 7, r[6] if r[6] is not None else 0)
            ws.cell(row_idx, 8, 'Yes' if r[7] else 'No')
            ws.cell(row_idx, 9, r[8].strftime('%Y-%m-%d %H:%M') if r[8] else '')
            ws.cell(row_idx, 10, r[9] or 'automated')

        installed = sum(1 for r in rows if r[5] == 'installed')
        superseded = sum(1 for r in rows if r[5] == 'superseded')
        devices = len({r[0] for r in rows if r[0]})
        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Total Remediation Records', len(rows)),
            ('Patches Installed', installed),
            ('Superseded', superseded),
            ('Devices Covered', devices),
            ('Data Source', 'RMM patch engine (cve_patch_job + vulnerability_cache)'),
        ])
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_rmm_patch_scan_file(self, evidence_name):
        """Patch scan / server scan-and-patch evidence: per-device patch posture
        from cve_patch_job rolled up by device, joined to rmm_agent last_seen.
        """
        rows = db.session.execute(text(
            """SELECT j.agent_id,
                      COALESCE(a.name, j.agent_id) AS device,
                      a.category,
                      COUNT(*) AS total_jobs,
                      COUNT(*) FILTER (WHERE j.status = 'installed') AS installed,
                      COUNT(*) FILTER (WHERE j.status = 'superseded') AS superseded,
                      COUNT(*) FILTER (WHERE j.status = 'no_patch') AS no_patch,
                      MAX(COALESCE(j.completed_at, j.updated_at, j.created_at)) AS last_action,
                      MAX(ag.last_seen_at) AS last_seen
               FROM cve_patch_job j
               LEFT JOIN asset a ON a.id = j.asset_id
               LEFT JOIN rmm_agent ag ON ag.agent_id = j.agent_id
               GROUP BY j.agent_id, a.name, a.category
               ORDER BY device"""
        )).fetchall()

        wb, ws = self.create_styled_workbook('Patch Scan')
        headers = ['Device', 'Agent ID', 'Category', 'Patch Jobs', 'Installed',
                   'Superseded', 'No Patch Needed', 'Last Patch Action', 'Agent Last Seen']
        self.style_header_row(ws, headers)
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row_idx, 1, r[1] or '')
            ws.cell(row_idx, 2, r[0] or '')
            ws.cell(row_idx, 3, r[2] or 'Workstation')
            ws.cell(row_idx, 4, r[3])
            ws.cell(row_idx, 5, r[4])
            ws.cell(row_idx, 6, r[5])
            ws.cell(row_idx, 7, r[6])
            ws.cell(row_idx, 8, r[7].strftime('%Y-%m-%d %H:%M') if r[7] else '')
            ws.cell(row_idx, 9, r[8].strftime('%Y-%m-%d %H:%M') if r[8] else 'Never')

        total_installed = sum(r[4] for r in rows)
        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Devices Scanned', len(rows)),
            ('Total Patch Jobs', sum(r[3] for r in rows)),
            ('Patches Installed', total_installed),
            ('Data Source', 'RMM patch engine (cve_patch_job + rmm_agent)'),
        ])
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_rmm_antivirus_file(self, evidence_name):
        """Antivirus/endpoint protection configuration from live RMM telemetry
        (rmm_telemetry.security_json: AV, firewall, antispyware, last scan).
        """
        rows = db.session.execute(text(
            """SELECT t.agent_id, COALESCE(a.name, t.hostname, t.agent_id) AS device,
                      t.os_name, t.security_json, t.last_seen
               FROM rmm_telemetry t
               LEFT JOIN asset a ON a.id = t.asset_id
               WHERE t.security_json IS NOT NULL AND t.security_json <> ''
               ORDER BY device"""
        )).fetchall()

        wb, ws = self.create_styled_workbook('Antivirus Config')
        headers = ['Device', 'OS', 'AV Product', 'AV Active', 'AV Updated',
                   'Firewall', 'Firewall Active', 'Last Scan', 'Telemetry As Of']
        self.style_header_row(ws, headers)
        protected = 0
        for row_idx, r in enumerate(rows, 2):
            try:
                sec = json.loads(r[3]) or {}
            except Exception:
                sec = {}
            av = (sec.get('av') or [{}])[0]
            fw = (sec.get('fw') or [{}])[0]
            last_scan = (sec.get('last_scan') or {}).get('time', '')
            av_active = bool(av.get('active'))
            if av_active:
                protected += 1
            ws.cell(row_idx, 1, self._sanitize_for_excel(r[1] or ''))
            ws.cell(row_idx, 2, self._sanitize_for_excel((r[2] or '')[:60]))
            ws.cell(row_idx, 3, self._sanitize_for_excel((av.get('name') or 'Unknown')[:60]))
            ws.cell(row_idx, 4, 'Yes' if av_active else 'No')
            ws.cell(row_idx, 5, 'Yes' if av.get('updated') else 'No')
            ws.cell(row_idx, 6, self._sanitize_for_excel((fw.get('name') or 'Unknown')[:80]))
            ws.cell(row_idx, 7, 'Yes' if fw.get('active') else 'No')
            ws.cell(row_idx, 8, str(last_scan)[:19])
            ws.cell(row_idx, 9, r[4].strftime('%Y-%m-%d %H:%M') if r[4] else '')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Devices Reporting', len(rows)),
            ('AV Active', protected),
            ('AV Not Active', len(rows) - protected),
            ('Data Source', 'RMM endpoint telemetry (rmm_telemetry.security_json)'),
        ])
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_admin_access_file(self, evidence_name):
        """Administrator access list per layer.

        Combines the Microsoft Entra directory-role snapshot
        (admin_role_snapshot, active assignments at the latest snapshot date)
        with M365 users flagged is_admin, so each privileged identity and its
        roles are captured for the access-review evidence.
        """
        # admin_role_snapshot stamps each role row with a per-row timestamp that
        # differs by microseconds within a single sync, so matching on the exact
        # MAX(snapshot_date) would return only one row. Use the most recent sync
        # *day* to capture the whole latest batch of active assignments.
        latest = db.session.execute(text(
            "SELECT MAX(snapshot_date) FROM admin_role_snapshot"
        )).scalar()
        role_rows = []
        if latest:
            role_rows = db.session.execute(text(
                """SELECT s.user_principal_name, s.role_name, s.assigned_date,
                          u.display_name, u.job_title, u.department, u.account_enabled
                   FROM admin_role_snapshot s
                   LEFT JOIN m365_user u
                     ON u.user_principal_name = s.user_principal_name AND u.is_current
                   WHERE s.status = 'active'
                     AND s.snapshot_date::date = (
                         SELECT MAX(snapshot_date)::date FROM admin_role_snapshot
                     )
                   ORDER BY s.user_principal_name, s.role_name"""
            )).fetchall()

        # group roles by user
        by_user = {}
        for r in role_rows:
            upn = r[0]
            entry = by_user.setdefault(upn, {
                'display_name': r[3] or '', 'job_title': r[4] or '',
                'department': r[5] or '', 'enabled': r[6], 'roles': [],
                'assigned': r[2],
            })
            entry['roles'].append(r[1])

        # include is_admin users that may not appear in the directory-role snapshot
        admin_users = M365User.query.filter_by(is_admin=True, is_current=True).all()
        for u in admin_users:
            if u.user_principal_name not in by_user:
                by_user[u.user_principal_name] = {
                    'display_name': u.display_name or '', 'job_title': u.job_title or '',
                    'department': u.department or '', 'enabled': u.account_enabled,
                    'roles': [x.strip() for x in (u.admin_roles or '').split(',') if x.strip()] or ['Administrator'],
                    'assigned': None,
                }

        wb, ws = self.create_styled_workbook('Admin Access')
        headers = ['User Principal Name', 'Display Name', 'Job Title', 'Department',
                   'Admin Roles', 'Account Enabled', 'Earliest Assignment']
        self.style_header_row(ws, headers)
        for row_idx, (upn, e) in enumerate(sorted(by_user.items()), 2):
            ws.cell(row_idx, 1, upn or '')
            ws.cell(row_idx, 2, e['display_name'])
            ws.cell(row_idx, 3, e['job_title'])
            ws.cell(row_idx, 4, e['department'])
            ws.cell(row_idx, 5, ', '.join(sorted(set(e['roles']))))
            ws.cell(row_idx, 6, 'Yes' if e['enabled'] else 'No')
            ws.cell(row_idx, 7, e['assigned'].strftime('%Y-%m-%d') if e['assigned'] else '')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Directory Snapshot Date', latest.strftime('%Y-%m-%d %H:%M') if latest else 'None'),
            ('Privileged Identities', len(by_user)),
            ('Distinct Roles', len({r[1] for r in role_rows})),
            ('Data Source', 'Entra directory roles (admin_role_snapshot) + m365_user.is_admin'),
        ])
        file_path = self.get_file_path(evidence_name, 'M365')
        wb.save(file_path)
        return file_path

    def generate_access_request_newhire_file(self, evidence_name):
        """SOC2 'Access Request - New Hire' evidence (control 97, Provisioning).

        Samples the most-recent APPROVED 'onboard_employee' command_ledger entry
        (the segregation-of-duties new-hire flow: HR requested, IT approved at
        /approvals and supplied the OU + groups, which triggered AD provisioning).
        Renders one artifact showing: employee identity, OU + groups granted,
        requested_by (HR), approved_by (IT) + approval date, and the tracking ticket.

        Raises ValueError('no approved onboarding to sample yet') when none exists
        — the row simply stays uncollected rather than fabricating evidence."""
        row = db.session.execute(text(
            """SELECT id, object_id, requested_by, before_state, after_state,
                      verification_detail, completed_at, created_at, status
               FROM command_ledger
               WHERE action_type='onboard_employee'
                 AND approval_status='approved'
                 AND status='succeeded'
               ORDER BY COALESCE(completed_at, created_at) DESC
               LIMIT 1"""
        )).fetchone()
        if not row:
            raise ValueError('no approved onboarding to sample yet')

        # before_state / after_state are JSON columns -> dicts (psycopg2 parses).
        bs = row[3] if isinstance(row[3], dict) else (json.loads(row[3]) if row[3] else {})
        after = row[4] if isinstance(row[4], dict) else (json.loads(row[4]) if row[4] else {})
        onboard = (bs or {}).get('onboard') or {}
        ver_detail = row[5] or ''
        approved_at = row[6] or row[7]
        requested_by = row[2] or ''

        # Approver — parse from verification_detail ("approved by <x>; OU=...; groups=...").
        approver = ''
        if 'approved by ' in ver_detail:
            approver = ver_detail.split('approved by ', 1)[1].split(';', 1)[0].strip()

        # OU + groups granted — prefer the recorded after_state (what actually ran).
        ou_dn = after.get('ou_dn') or ''
        ou_cn = ou_dn.split(',')[0].split('=')[-1] if '=' in ou_dn.split(',')[0] else ou_dn
        groups = after.get('groups') or []
        sam = after.get('sam') or onboard.get('sam') or ''

        # Tracking ticket reference ([ONBOARD] ticket for this hire).
        emp_name = onboard.get('name') or after.get('employee') or ''
        ticket_ref = ''
        if emp_name:
            t = db.session.execute(text(
                "SELECT id FROM support_ticket WHERE category='HR / Onboarding' "
                "AND subject LIKE :s ORDER BY id DESC LIMIT 1"
            ), {'s': f'%{emp_name}%'}).fetchone()
            if t:
                ticket_ref = f'#{t[0]}'

        wb, ws = self.create_styled_workbook('Access Request - New Hire')
        headers = ['Field', 'Value']
        self.style_header_row(ws, headers)
        rows = [
            ('Employee', emp_name),
            ('Email', onboard.get('email') or ''),
            ('sAMAccountName', sam),
            ('Department', onboard.get('dept') or ''),
            ('Job Title', onboard.get('title') or ''),
            ('Manager', onboard.get('manager') or ''),
            ('Start Date', onboard.get('start') or ''),
            ('Work Type', onboard.get('work_type') or ''),
            ('AD OU Granted', ou_cn or ou_dn),
            ('AD OU (DN)', ou_dn),
            ('Security Groups Granted', ', '.join(groups) if groups else ''),
            ('Requested By (HR)', requested_by),
            ('Approved By (IT)', approver),
            ('Approval Date', approved_at.strftime('%Y-%m-%d %H:%M') if approved_at else ''),
            ('Tracking Ticket', ticket_ref),
            ('Ledger Entry', f'command_ledger #{row[0]}'),
            ('Verification Detail', ver_detail),
        ]
        for row_idx, (label, value) in enumerate(rows, 2):
            ws.cell(row_idx, 1, label).font = Font(bold=True)
            ws.cell(row_idx, 2, self._sanitize_for_excel(value))
        ws.column_dimensions['A'].width = 28
        ws.column_dimensions['B'].width = 60

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Control', 'Provisioning (control 97)'),
            ('Sampled New Hire', emp_name),
            ('Segregation of Duties', 'HR requested; IT approved + provisioned'),
            ('Data Source', 'command_ledger (approved onboard_employee) + employee onboarding request'),
        ])
        file_path = self.get_file_path(evidence_name, 'M365')
        wb.save(file_path)
        return file_path

    def generate_asset_inventory_file(self, evidence_name):
        """Asset inventory from the local asset register, enriched with Intune
        enrollment/compliance where the device is managed.
        """
        assets = db.session.execute(text(
            """SELECT a.asset_tag, a.name, a.category, a.manufacturer, a.model,
                      a.serial_number, a.status, a.location, a.os_version,
                      a.last_seen, d.compliance_state, d.is_encrypted
               FROM asset a
               LEFT JOIN intune_device d ON d.asset_id = a.id AND d.is_current
               ORDER BY a.name"""
        )).fetchall()

        wb, ws = self.create_styled_workbook('Asset Inventory')
        headers = ['Asset Tag', 'Name', 'Category', 'Manufacturer', 'Model',
                   'Serial', 'Status', 'Location', 'OS Version', 'Last Seen',
                   'Intune Compliance', 'Encrypted']
        self.style_header_row(ws, headers)
        for row_idx, a in enumerate(assets, 2):
            ws.cell(row_idx, 1, a[0] or '')
            ws.cell(row_idx, 2, self._sanitize_for_excel(a[1] or ''))
            ws.cell(row_idx, 3, a[2] or '')
            ws.cell(row_idx, 4, a[3] or '')
            ws.cell(row_idx, 5, self._sanitize_for_excel(a[4] or ''))
            ws.cell(row_idx, 6, a[5] or '')
            ws.cell(row_idx, 7, a[6] or '')
            ws.cell(row_idx, 8, a[7] or '')
            ws.cell(row_idx, 9, (a[8] or '')[:40])
            ws.cell(row_idx, 10, a[9].strftime('%Y-%m-%d %H:%M') if a[9] else '')
            ws.cell(row_idx, 11, a[10] or 'Not Managed')
            ws.cell(row_idx, 12, ('Yes' if a[11] else 'No') if a[11] is not None else 'Unknown')

        managed = sum(1 for a in assets if a[10])
        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Total Assets', len(assets)),
            ('Intune-Managed', managed),
            ('Data Source', 'Asset register (asset) + Intune (intune_device)'),
        ])
        file_path = self.get_file_path(evidence_name, 'M365/Intune')
        wb.save(file_path)
        return file_path

    def generate_disk_encryption_file(self, evidence_name):
        """Device disk encryption status from Intune (intune_device.is_encrypted)."""
        devices = IntuneDevice.query.filter_by(is_current=True).order_by(IntuneDevice.device_name).all()
        wb, ws = self.create_styled_workbook('Disk Encryption')
        headers = ['Device Name', 'User', 'Manufacturer', 'Model', 'OS Version',
                   'Encrypted', 'Compliance State', 'Last Sync']
        self.style_header_row(ws, headers)
        encrypted = 0
        for row_idx, d in enumerate(devices, 2):
            if d.is_encrypted:
                encrypted += 1
            ws.cell(row_idx, 1, d.device_name or '')
            ws.cell(row_idx, 2, d.user_display_name or d.user_principal_name or '')
            ws.cell(row_idx, 3, d.manufacturer or '')
            ws.cell(row_idx, 4, d.model or '')
            ws.cell(row_idx, 5, d.os_version or '')
            ws.cell(row_idx, 6, 'Yes' if d.is_encrypted else 'No')
            ws.cell(row_idx, 7, d.compliance_state or '')
            ws.cell(row_idx, 8, d.last_sync_datetime.strftime('%Y-%m-%d %H:%M') if d.last_sync_datetime else '')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Total Devices', len(devices)),
            ('Encrypted', encrypted),
            ('Not Encrypted', len(devices) - encrypted),
            ('Data Source', 'Intune (intune_device.is_encrypted)'),
        ])
        file_path = self.get_file_path(evidence_name, 'M365/Intune')
        wb.save(file_path)
        return file_path

    # ------------------------------------------------------------------
    # Settings-evidence live-config generators (the "screenshot" equivalent:
    # a dated export of the actual current configuration values from the
    # Tracker's own data / live integrations). No fabricated values.
    # ------------------------------------------------------------------

    def generate_m365_password_settings_file(self, evidence_name):
        """Network/Cloud password settings, LIVE from Microsoft Entra (M365).

        Pulls the real password-expiration policy per verified domain
        (Graph /domains: passwordValidityPeriodInDays /
        passwordNotificationWindowInDays), the authentication-methods policy
        state, and the enforced Conditional Access policies (MFA). This is the
        authoritative, auditable network/cloud password configuration; no
        values are hard-coded. Returns None if the tenant is unreachable so the
        item stays "not collected" rather than reporting fabricated settings.
        """
        try:
            import requests
            from m365_config import get_m365_credentials
            from m365_service import M365Service

            tenant_id, client_id, client_secret = get_m365_credentials()
            if not all([tenant_id, client_id, client_secret]):
                print('[M365 password] credentials not configured')
                return None
            svc = M365Service(tenant_id, client_id, client_secret)
            token = svc.get_access_token()
            if not token:
                print('[M365 password] could not obtain access token')
                return None
            headers = {'Authorization': f'Bearer {token}'}

            domains_resp = requests.get(
                'https://graph.microsoft.com/v1.0/domains', headers=headers, timeout=25)
            if not domains_resp.ok:
                print(f'[M365 password] /domains -> {domains_resp.status_code}')
                return None
            domains = [d for d in domains_resp.json().get('value', []) if d.get('isVerified')]

            try:
                ca_policies = svc.get_conditional_access_policies()
            except Exception:
                ca_policies = []

            wb, ws = self.create_styled_workbook('Domain Password Policy')
            headers_row = ['Domain', 'Default', 'Password Validity (days)',
                           'Notification Window (days)', 'Never Expires', 'Authentication Type']
            self.style_header_row(ws, headers_row)
            never_expire_sentinel = 2147483647
            for row_idx, d in enumerate(domains, 2):
                validity = d.get('passwordValidityPeriodInDays')
                never = validity == never_expire_sentinel
                ws.cell(row_idx, 1, d.get('id', ''))
                ws.cell(row_idx, 2, 'Yes' if d.get('isDefault') else 'No')
                ws.cell(row_idx, 3, 'Never' if never else (validity if validity is not None else 'Inherited'))
                ws.cell(row_idx, 4, d.get('passwordNotificationWindowInDays') if d.get('passwordNotificationWindowInDays') is not None else 'Inherited')
                ws.cell(row_idx, 5, 'Yes' if never else 'No')
                ws.cell(row_idx, 6, d.get('authenticationType', ''))

            # Conditional Access (MFA enforcement) detail sheet
            ws_ca = wb.create_sheet('Conditional Access')
            ca_headers = ['Policy Name', 'State', 'Grant Controls']
            self.style_header_row(ws_ca, ca_headers)
            for row_idx, p in enumerate(ca_policies, 2):
                controls = (p.get('grantControls') or {}).get('builtInControls') or []
                ws_ca.cell(row_idx, 1, self._sanitize_for_excel(p.get('displayName', '')[:120]))
                ws_ca.cell(row_idx, 2, p.get('state', ''))
                ws_ca.cell(row_idx, 3, ', '.join(controls) if controls else 'None')

            enabled_ca = sum(1 for p in ca_policies if p.get('state') == 'enabled')
            mfa_ca = sum(1 for p in ca_policies
                         if 'mfa' in ((p.get('grantControls') or {}).get('builtInControls') or []))
            self._summary_sheet(wb, [
                ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
                ('Data Source', 'Microsoft Entra / Graph API (live)'),
                ('Verified Domains', len(domains)),
                ('Conditional Access Policies', len(ca_policies)),
                ('Conditional Access Enabled', enabled_ca),
                ('Policies Requiring MFA', mfa_ca),
                ('Layer', 'Network/Cloud (Microsoft Entra ID)'),
            ])
            file_path = self.get_file_path(evidence_name, 'M365/Intune')
            wb.save(file_path)
            return file_path
        except Exception as exc:
            print(f'[M365 password] error: {exc}')
            import traceback
            traceback.print_exc()
            return None

    def generate_monitoring_tools_file(self, evidence_name):
        """Enabled monitoring tools/checks, LIVE from the Tracker's own
        monitoring subsystem (monitoring_profile + monitoring_check) plus any
        synced Azure Monitor alert rules. Reports the actual enabled monitors.
        """
        profiles = db.session.execute(text(
            """SELECT name, device_type, os_family, severity_level,
                      check_interval_minutes, enabled
               FROM monitoring_profile ORDER BY name"""
        )).fetchall()
        checks = db.session.execute(text(
            """SELECT name, check_type, script_type, timeout_seconds, enabled
               FROM monitoring_check ORDER BY name"""
        )).fetchall()
        azure_alerts = AzureMonitorAlert.query.filter_by(is_current=True).all()

        wb, ws = self.create_styled_workbook('Monitoring Profiles')
        self.style_header_row(ws, ['Profile', 'Device Type', 'OS Family',
                                   'Severity', 'Interval (min)', 'Enabled'])
        for row_idx, p in enumerate(profiles, 2):
            ws.cell(row_idx, 1, p[0] or '')
            ws.cell(row_idx, 2, p[1] or '')
            ws.cell(row_idx, 3, p[2] or '')
            ws.cell(row_idx, 4, p[3] or '')
            ws.cell(row_idx, 5, p[4] if p[4] is not None else '')
            ws.cell(row_idx, 6, 'Yes' if p[5] else 'No')

        ws_checks = wb.create_sheet('Monitoring Checks')
        self.style_header_row(ws_checks, ['Check', 'Type', 'Script Type',
                                          'Timeout (s)', 'Enabled'])
        for row_idx, c in enumerate(checks, 2):
            ws_checks.cell(row_idx, 1, c[0] or '')
            ws_checks.cell(row_idx, 2, c[1] or '')
            ws_checks.cell(row_idx, 3, c[2] or '')
            ws_checks.cell(row_idx, 4, c[3] if c[3] is not None else '')
            ws_checks.cell(row_idx, 5, 'Yes' if c[4] else 'No')

        if azure_alerts:
            ws_az = wb.create_sheet('Azure Monitor Alerts')
            self.style_header_row(ws_az, ['Alert Name', 'Resource Group', 'Severity', 'Enabled'])
            for row_idx, a in enumerate(azure_alerts, 2):
                ws_az.cell(row_idx, 1, a.alert_name or '')
                ws_az.cell(row_idx, 2, a.resource_group or '')
                ws_az.cell(row_idx, 3, a.severity or '')
                ws_az.cell(row_idx, 4, 'Yes' if a.enabled else 'No')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Data Source', 'Tracker monitoring subsystem + Azure Monitor (synced)'),
            ('Monitoring Profiles', len(profiles)),
            ('Enabled Profiles', sum(1 for p in profiles if p[5])),
            ('Monitoring Checks', len(checks)),
            ('Enabled Checks', sum(1 for c in checks if c[4])),
            ('Azure Monitor Alert Rules', len(azure_alerts)),
        ])
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_monitoring_alert_config_file(self, evidence_name):
        """Performance/monitoring alert configuration, LIVE from the Tracker's
        monitoring_alert records (severity/status/check) joined to the check
        definitions, plus synced Azure Monitor alert rules. Reports the actual
        configured alerting, not a documented standard.
        """
        alerts = db.session.execute(text(
            """SELECT ma.severity, ma.status, mc.name, mc.check_type,
                      ma.message, ma.triggered_at
               FROM monitoring_alert ma
               LEFT JOIN monitoring_check mc ON mc.id = ma.check_id
               ORDER BY ma.triggered_at DESC NULLS LAST"""
        )).fetchall()
        azure_alerts = AzureMonitorAlert.query.filter_by(is_current=True).all()

        wb, ws = self.create_styled_workbook('Alert Configuration')
        self.style_header_row(ws, ['Severity', 'Status', 'Check', 'Check Type',
                                   'Message', 'Triggered At'])
        sev_counts = {}
        for row_idx, a in enumerate(alerts, 2):
            sev_counts[a[0] or 'Unknown'] = sev_counts.get(a[0] or 'Unknown', 0) + 1
            ws.cell(row_idx, 1, a[0] or '')
            ws.cell(row_idx, 2, a[1] or '')
            ws.cell(row_idx, 3, a[2] or '')
            ws.cell(row_idx, 4, a[3] or '')
            ws.cell(row_idx, 5, self._sanitize_for_excel((a[4] or '')[:200]))
            ws.cell(row_idx, 6, a[5].strftime('%Y-%m-%d %H:%M') if a[5] else '')

        if azure_alerts:
            ws_az = wb.create_sheet('Azure Monitor Alert Rules')
            self.style_header_row(ws_az, ['Alert Name', 'Resource Group',
                                          'Target Resource', 'Severity', 'Enabled'])
            for row_idx, a in enumerate(azure_alerts, 2):
                ws_az.cell(row_idx, 1, a.alert_name or '')
                ws_az.cell(row_idx, 2, a.resource_group or '')
                ws_az.cell(row_idx, 3, a.target_resource or '')
                ws_az.cell(row_idx, 4, a.severity or '')
                ws_az.cell(row_idx, 5, 'Yes' if a.enabled else 'No')

        summary = [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Data Source', 'Tracker monitoring_alert + Azure Monitor (synced)'),
            ('Configured Alerts (records)', len(alerts)),
            ('Azure Monitor Alert Rules', len(azure_alerts)),
        ]
        for sev, cnt in sorted(sev_counts.items()):
            summary.append((f'Alerts - {sev}', cnt))
        self._summary_sheet(wb, summary)
        file_path = self.get_file_path(evidence_name, 'RMM')
        wb.save(file_path)
        return file_path

    def generate_server_encryption_file(self, evidence_name):
        """Server encryption configuration, LIVE from azure_vm.disk_encryption
        (synced Azure VMs) and Intune server-class devices (is_encrypted). The
        actual current encryption state of server infrastructure.
        """
        vms = AzureVirtualMachine.query.filter_by(is_current=True).order_by(
            AzureVirtualMachine.name).all()
        # Intune server-class devices (Windows/Linux Server in the OS string)
        intune_servers = [
            d for d in IntuneDevice.query.filter_by(is_current=True).all()
            if 'server' in ((d.os_version or '') + ' ' + (d.model or '')).lower()
        ]

        wb, ws = self.create_styled_workbook('Azure VM Encryption')
        self.style_header_row(ws, ['VM Name', 'Resource Group', 'Location',
                                   'OS Type', 'VM Size', 'Disk Encryption'])
        enc_vms = 0
        for row_idx, vm in enumerate(vms, 2):
            if vm.disk_encryption:
                enc_vms += 1
            ws.cell(row_idx, 1, vm.name or '')
            ws.cell(row_idx, 2, vm.resource_group or '')
            ws.cell(row_idx, 3, vm.location or '')
            ws.cell(row_idx, 4, vm.os_type or '')
            ws.cell(row_idx, 5, vm.vm_size or '')
            ws.cell(row_idx, 6, 'Yes' if vm.disk_encryption else 'No')

        if intune_servers:
            ws_i = wb.create_sheet('Intune Servers')
            self.style_header_row(ws_i, ['Device Name', 'OS Version', 'Model',
                                         'Encrypted', 'Compliance State'])
            for row_idx, d in enumerate(intune_servers, 2):
                ws_i.cell(row_idx, 1, d.device_name or '')
                ws_i.cell(row_idx, 2, d.os_version or '')
                ws_i.cell(row_idx, 3, d.model or '')
                ws_i.cell(row_idx, 4, 'Yes' if d.is_encrypted else 'No')
                ws_i.cell(row_idx, 5, d.compliance_state or '')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Data Source', 'azure_vm.disk_encryption + Intune (is_encrypted)'),
            ('Azure VMs', len(vms)),
            ('Azure VMs Encrypted', enc_vms),
            ('Intune Server Devices', len(intune_servers)),
            ('Intune Servers Encrypted', sum(1 for d in intune_servers if d.is_encrypted)),
        ])
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path

    def generate_separation_of_environments_file(self, evidence_name):
        """Separation of prod/dev/test environments, LIVE from synced Azure VMs
        grouped by resource group with an environment classification derived
        from the resource-group / VM naming. Reports the actual deployed
        environments; classification heuristics are stated, not fabricated.
        """
        vms = AzureVirtualMachine.query.filter_by(is_current=True).order_by(
            AzureVirtualMachine.resource_group, AzureVirtualMachine.name).all()

        def classify(vm):
            blob = f"{vm.resource_group or ''} {vm.name or ''}".lower()
            if any(t in blob for t in ('prod', 'prd')):
                return 'Production'
            if any(t in blob for t in ('test', 'tst', 'qa', 'stage', 'stg', 'uat')):
                return 'Test/Staging'
            if any(t in blob for t in ('dev', 'leadgen', 'arm')):
                return 'Development'
            return 'Unclassified'

        wb, ws = self.create_styled_workbook('Environment Separation')
        self.style_header_row(ws, ['VM Name', 'Resource Group', 'Location',
                                   'OS Type', 'Derived Environment'])
        env_counts = {}
        rg_set = set()
        for row_idx, vm in enumerate(vms, 2):
            env = classify(vm)
            env_counts[env] = env_counts.get(env, 0) + 1
            rg_set.add(vm.resource_group or '')
            ws.cell(row_idx, 1, vm.name or '')
            ws.cell(row_idx, 2, vm.resource_group or '')
            ws.cell(row_idx, 3, vm.location or '')
            ws.cell(row_idx, 4, vm.os_type or '')
            ws.cell(row_idx, 5, env)

        summary = [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Data Source', 'azure_vm (synced), grouped by resource group'),
            ('Classification', 'Derived from resource-group / VM name tokens'),
            ('Total VMs', len(vms)),
            ('Distinct Resource Groups', len(rg_set)),
        ]
        for env, cnt in sorted(env_counts.items()):
            summary.append((f'Environment - {env}', cnt))
        self._summary_sheet(wb, summary)
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path

    def generate_ids_config_file(self, evidence_name):
        """Intrusion-detection posture, LIVE from synced Azure Defender for
        Cloud security alerts + assessments. Reports the actual detection
        signal currently held; if no synced detections exist the report still
        documents that state (it does not fabricate findings)."""
        alerts = AzureSecurityAlert.query.filter_by(is_current=True).all()
        assessments = AzureSecurityAssessment.query.filter_by(is_current=True).all()

        wb, ws = self.create_styled_workbook('Detection Alerts')
        self.style_header_row(ws, ['Alert Name', 'Severity', 'Status',
                                   'Affected Resource', 'Detection Time'])
        for row_idx, a in enumerate(alerts, 2):
            ws.cell(row_idx, 1, a.alert_name or '')
            ws.cell(row_idx, 2, a.severity or '')
            ws.cell(row_idx, 3, a.status or '')
            ws.cell(row_idx, 4, a.affected_resource or '')
            ws.cell(row_idx, 5, a.detection_time.strftime('%Y-%m-%d %H:%M') if a.detection_time else '')

        if assessments:
            ws_a = wb.create_sheet('Security Assessments')
            self.style_header_row(ws_a, ['Assessment', 'Status', 'Severity', 'Resource'])
            for row_idx, s in enumerate(assessments, 2):
                ws_a.cell(row_idx, 1, s.name or '')
                ws_a.cell(row_idx, 2, s.status or '')
                ws_a.cell(row_idx, 3, s.severity or '')
                ws_a.cell(row_idx, 4, s.resource_id or '')

        self._summary_sheet(wb, [
            ('Report Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')),
            ('Data Source', 'Azure Defender for Cloud (synced alerts + assessments)'),
            ('Security Alerts', len(alerts)),
            ('Security Assessments', len(assessments)),
            ('Note', 'Documented IDS standard is the Logging & Monitoring Procedure (ISMS).'),
        ])
        file_path = self.get_file_path(evidence_name, 'Azure')
        wb.save(file_path)
        return file_path

    # ------------------------------------------------------------------
    # ISMS-manual policy evidence generator
    # ------------------------------------------------------------------

    def _load_isms_manual_body(self):
        """Return (markdown_body, version_number) for the current published
        ISMS manual (isms_document slug=isms-manual). Returns (None, None) if
        the manual or its current version is unavailable."""
        from models import ISMSDocument, ISMSDocumentVersion

        doc = ISMSDocument.query.filter_by(slug='isms-manual').first()
        if not doc:
            return None, None
        version = None
        if doc.current_version_id:
            version = ISMSDocumentVersion.query.get(doc.current_version_id)
        if version is None:
            # fall back to the highest version_number for this document
            version = (ISMSDocumentVersion.query
                       .filter_by(document_id=doc.id)
                       .order_by(ISMSDocumentVersion.version_number.desc())
                       .first())
        if version is None or not version.markdown_body:
            return None, None
        return version.markdown_body, version.version_number

    def _extract_isms_section(self, body, is_id):
        """Extract one IS-section from the manual markdown body.

        A section runs from its ``## IS-<id>: Title`` heading up to (but not
        including) the next IS-section heading of the same or shallower level.
        Non-IS sub-headings (e.g. "## SOC 2 Trust Services Criteria Mapping")
        that appear *within* a policy are kept as part of the section.

        Returns (title, section_text) or None if the section isn't found.
        """
        if not body:
            return None
        lines = body.split('\n')
        start = level = None
        title = None
        for i, ln in enumerate(lines):
            m = _IS_HEADING_RE.match(ln.strip())
            if m and m.group(2) == is_id:
                start = i
                level = len(m.group(1))
                title = m.group(3).strip()
                break
        if start is None:
            return None
        end = len(lines)
        for j in range(start + 1, len(lines)):
            m = _IS_HEADING_RE.match(lines[j].strip())
            if m and len(m.group(1)) <= level:
                end = j
                break
        return title, '\n'.join(lines[start:end]).strip()

    def _find_isms_section_by_name(self, body, evidence_name):
        """Last-resort resolution: match a policy evidence name to a manual
        IS-section heading by title similarity. Returns the IS-ID or None."""
        if not body:
            return None
        target = evidence_name.lower()
        # strip generic suffixes so "Risk Management Policy and Procedures"
        # still matches a "Risk Management Policy" heading.
        for suffix in (' and procedures', ' policy and procedures', ' procedures',
                       ' policy', ' plan', ' schedule'):
            if target.endswith(suffix):
                target = target[: -len(suffix)]
                break
        target = target.strip()
        if not target:
            return None
        for ln in body.split('\n'):
            m = _IS_HEADING_RE.match(ln.strip())
            if m and target in m.group(3).strip().lower():
                return m.group(2)
        return None

    def _resolve_policy_sections(self, evidence_item, body):
        """Resolve a Policy evidence row to a list of ISMS IS-section IDs.

        Order: (a) the linked control's authoritative_docs IS-* IDs;
        (b) the verified POLICY_EVIDENCE_SECTION_MAP; (c) a manual title search.
        Returns (is_ids, method) where method is one of
        'authoritative_docs' | 'name-match' | 'title-search' | None.
        Only IS-IDs that actually exist as headings in ``body`` are returned.
        """
        name = evidence_item.evidence_name

        # Items explicitly reverted to manual/HR-sourced evidence are never
        # auto-resolved from the ISMS manual, even if the linked control still
        # carries authoritative_docs IS-* pointers (those remain as a doc ref).
        if name in POLICY_EVIDENCE_REVERTED_TO_MANUAL:
            return [], None

        def _existing(ids):
            out = []
            for sid in ids:
                if sid in out:
                    continue
                if self._extract_isms_section(body, sid):
                    out.append(sid)
            return out

        # (a) authoritative_docs on the linked control
        if evidence_item.control_id:
            control = SOC2Control.query.get(evidence_item.control_id)
            if control and control.authoritative_docs:
                # _IS_SECTION_RE only matches IS-* tokens, so external-file
                # references (e.g. ISMS-Manual.docx) are ignored; _existing()
                # further drops any IS-ID that isn't an extractable heading.
                ids = _IS_SECTION_RE.findall(control.authoritative_docs)
                resolved = _existing(ids)
                if resolved:
                    return resolved, 'authoritative_docs'

        # (b) explicit verified map
        if name in POLICY_EVIDENCE_SECTION_MAP:
            resolved = _existing(POLICY_EVIDENCE_SECTION_MAP[name])
            if resolved:
                return resolved, 'name-match'

        # (c) title search
        if name not in POLICY_EVIDENCE_NOT_IN_MANUAL:
            sid = self._find_isms_section_by_name(body, name)
            if sid:
                return [sid], 'title-search'

        return [], None

    def _markdown_section_to_flowables(self, section_text, styles):
        """Convert an extracted ISMS markdown section into reportlab flowables.

        Handles markdown headings (#..######), bullet lists (-/*), simple GFM
        tables (| a | b |), bold (**x**) and italics (*x*). This is a pragmatic
        renderer for the manual's structure, not a full markdown engine.
        """
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib import colors

        body_style = styles['isms_body']
        h2 = styles['isms_h2']
        h3 = styles['isms_h3']
        bullet_style = styles['isms_bullet']

        def _inline(text):
            text = text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            text = re.sub(r'\\([\\`*_{}\[\]()#+\-.!])', r'\1', text)
            text = re.sub(r'\*\*(.+?)\*\*', r'<b>\1</b>', text)
            text = re.sub(r'(?<!\*)\*(?!\*)(.+?)(?<!\*)\*(?!\*)', r'<i>\1</i>', text)
            return text

        flowables = []
        lines = section_text.split('\n')
        i = 0
        n = len(lines)
        while i < n:
            raw = lines[i]
            line = raw.strip()
            if not line:
                i += 1
                continue

            # GFM table block
            if line.startswith('|') and i + 1 < n and re.match(r'^\|[\s:|-]+\|?\s*$', lines[i + 1].strip()):
                table_rows = []
                while i < n and lines[i].strip().startswith('|'):
                    cells = [c.strip() for c in lines[i].strip().strip('|').split('|')]
                    table_rows.append(cells)
                    i += 1
                # row index 1 is the --- separator
                rendered = []
                for r_idx, cells in enumerate(table_rows):
                    if r_idx == 1:
                        continue
                    rendered.append([Paragraph(_inline(c), body_style) for c in cells])
                if rendered:
                    width = 6.9 * inch
                    ncols = max(len(r) for r in rendered)
                    col_w = width / ncols if ncols else width
                    # pad short rows
                    for r in rendered:
                        while len(r) < ncols:
                            r.append(Paragraph('', body_style))
                    tbl = Table(rendered, colWidths=[col_w] * ncols)
                    tbl.setStyle(TableStyle([
                        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cbd5e1')),
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
                        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 4),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 3),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ]))
                    flowables.append(tbl)
                    flowables.append(Spacer(1, 8))
                continue

            # headings
            hm = re.match(r'^(#{1,6})\s+(.*)$', line)
            if hm:
                depth = len(hm.group(1))
                text = _inline(hm.group(2).strip())
                flowables.append(Paragraph(text, h2 if depth <= 2 else h3))
                i += 1
                continue

            # bullets
            bm = re.match(r'^[-*]\s+(.*)$', line)
            if bm:
                flowables.append(Paragraph(_inline(bm.group(1)), bullet_style))
                i += 1
                continue

            flowables.append(Paragraph(_inline(line), body_style))
            i += 1
        return flowables

    def generate_isms_section_pdf(self, evidence_name):
        """Generate an auditor-ready PDF for a Policy evidence item by extracting
        the relevant section(s) directly from the published ISMS Manual.

        Resolves the evidence item to ISMS IS-section IDs (control
        authoritative_docs -> verified map -> title search), extracts each
        section from the current manual version's markdown_body, and renders a
        dated PDF stamped with the source. Returns the file path, or None if the
        item cannot be resolved to any manual section (left Manual + logged).
        """
        body, version_number = self._load_isms_manual_body()
        if not body:
            print(f"[ISMS] manual body unavailable; cannot generate '{evidence_name}'")
            return None

        evidence_item = StrikeGraphEvidence.query.filter_by(evidence_name=evidence_name).first()
        if not evidence_item:
            print(f"[ISMS] no catalog row for '{evidence_name}'")
            return None

        is_ids, method = self._resolve_policy_sections(evidence_item, body)
        if not is_ids:
            print(f"[ISMS] UNRESOLVED policy evidence '{evidence_name}' "
                  f"(no ISMS section); leaving Manual.")
            return None

        sections = []
        for sid in is_ids:
            extracted = self._extract_isms_section(body, sid)
            if extracted:
                sections.append((sid, extracted[0], extracted[1]))
        if not sections:
            print(f"[ISMS] resolved ids {is_ids} for '{evidence_name}' but no section text extracted")
            return None

        ver_label = f"v{version_number}" if version_number else "v4"
        generated = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')

        # styles
        base = getSampleStyleSheet()
        styles = {
            'title': ParagraphStyle('isms_title', parent=base['Heading1'], fontSize=18,
                                    textColor='#2D4639', spaceAfter=6, alignment=TA_CENTER),
            'subtitle': ParagraphStyle('isms_subtitle', parent=base['Normal'], fontSize=10,
                                       textColor='#5b6b62', alignment=TA_CENTER, spaceAfter=4),
            'source': ParagraphStyle('isms_source', parent=base['Normal'], fontSize=9,
                                     textColor='#5b6b62', alignment=TA_CENTER, spaceAfter=2),
            'isms_h2': ParagraphStyle('isms_h2', parent=base['Heading2'], fontSize=13,
                                      textColor='#2D4639', spaceBefore=12, spaceAfter=6),
            'isms_h3': ParagraphStyle('isms_h3', parent=base['Heading3'], fontSize=11,
                                      textColor='#2D4639', spaceBefore=8, spaceAfter=4),
            'isms_body': ParagraphStyle('isms_body', parent=base['BodyText'], fontSize=9.5,
                                        alignment=TA_JUSTIFY, spaceAfter=6, leading=13),
            'isms_bullet': ParagraphStyle('isms_bullet', parent=base['BodyText'], fontSize=9.5,
                                          leftIndent=16, bulletIndent=6, spaceAfter=3, leading=13),
        }

        story = []
        story.append(Paragraph("Cirque Corporation", styles['title']))
        story.append(Paragraph(evidence_name, styles['subtitle']))
        section_ids_label = ', '.join(s[0] for s in sections)
        story.append(Paragraph(
            f"Extracted from ISMS Manual {ver_label} (published), "
            f"section {section_ids_label} &mdash; generated {generated}",
            styles['source']))
        story.append(Paragraph(f"Resolution method: {method}", styles['source']))
        story.append(Spacer(1, 0.25 * inch))

        for idx, (sid, sec_title, sec_text) in enumerate(sections):
            if idx > 0:
                story.append(PageBreak())
            story.append(Paragraph(f"{sid}: {sec_title}", styles['isms_h2']))
            story.append(Spacer(1, 4))
            story.extend(self._markdown_section_to_flowables(sec_text, styles))

        file_path = self.get_file_path(evidence_name, 'ISMS', 'pdf')
        SimpleDocTemplate(file_path, pagesize=letter,
                          topMargin=0.7 * inch, bottomMargin=0.7 * inch).build(story)
        return file_path

    # ------------------------------------------------------------------
    # Catalog wiring: stamp StrikeGraphEvidence + write EvidenceSnapshot
    # ------------------------------------------------------------------

    AUTOMATION_SOURCE_BY_DIR = {
        'm365': 'M365/Intune',
        'M365/Defender': 'M365/Defender',
        'azure': 'Azure',
        'isms': 'ISMS',
        'teamviewer': 'TeamViewer',
        'rmm': 'RMM',
    }

    def _infer_automation_source(self, file_path):
        parent = os.path.basename(os.path.dirname(file_path))
        # the defender dir nests under M365; check the two-level tail
        tail2 = '/'.join(file_path.split('/')[-3:-1])
        if tail2.endswith('M365/Defender'):
            return 'M365/Defender'
        return self.AUTOMATION_SOURCE_BY_DIR.get(parent, 'M365/Intune')

    def _count_data_rows(self, file_path):
        """Best-effort count of data rows in the primary sheet of an .xlsx."""
        if not file_path.endswith('.xlsx'):
            return None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            # primary data sheet is the non-Summary sheet (or first sheet)
            data_sheet = None
            for name in wb.sheetnames:
                if name != 'Summary':
                    data_sheet = wb[name]
                    break
            data_sheet = data_sheet or wb[wb.sheetnames[0]]
            count = max(data_sheet.max_row - 1, 0)
            wb.close()
            return count
        except Exception:
            return None

    def _resolve_snapshot_control_id(self, evidence_item):
        """Resolve a control to attach the audit snapshot to. Prefer the
        evidence row's own control_id; otherwise match a control by name for a
        small set of known-unlinked rows. Returns None if none resolves (the
        snapshot is then skipped rather than fabricated)."""
        if evidence_item.control_id:
            return evidence_item.control_id
        name_to_control = {
            'Vulnerability Scan Results': 'Vulnerability Scan',
            'Vulnerability Remediation': 'Vulnerability Scan',
        }
        control_name = name_to_control.get(evidence_item.evidence_name)
        if control_name:
            control = SOC2Control.query.filter_by(control_name=control_name).first()
            if control:
                return control.id
        return None

    def stamp_evidence_collection(self, evidence_name, file_path, collected_by='automated'):
        """Record a generated artifact back to the catalog.

        - sets StrikeGraphEvidence.file_path / last_submitted_date / submission_status
        - sets automation_source when missing (idempotent backfill on first run)
        - updates the linked control's last_evidence_date
        - writes an EvidenceSnapshot audit row (when a control resolves)
        Returns the StrikeGraphEvidence row (or None if the name is unknown).
        """
        evidence_item = StrikeGraphEvidence.query.filter_by(evidence_name=evidence_name).first()
        if not evidence_item or not file_path or not os.path.exists(file_path):
            return None

        now = datetime.utcnow()
        record_count = self._count_data_rows(file_path)

        evidence_item.file_path = file_path
        evidence_item.last_submitted_date = now
        evidence_item.submission_status = 'Submitted'
        evidence_item.updated_at = now
        if not evidence_item.automation_source:
            evidence_item.automation_source = self._infer_automation_source(file_path)

        control_id = self._resolve_snapshot_control_id(evidence_item)
        if control_id:
            control = SOC2Control.query.get(control_id)
            if control:
                control.last_evidence_date = now
                control.updated_at = now
            db.session.add(EvidenceSnapshot(
                control_id=control_id,
                snapshot_date=now,
                evidence_type=evidence_item.automation_source or 'Automated',
                evidence_data=json.dumps({
                    'evidence_name': evidence_name,
                    'file': os.path.basename(file_path),
                    'record_count': record_count,
                }),
                record_count=record_count,
                status='collected',
                collected_by=collected_by or 'automated',
                notes=f'Auto-generated evidence artifact: {os.path.basename(file_path)}',
                file_path=file_path,
            ))

        db.session.commit()
        return evidence_item

    def generate_and_record(self, evidence_name, collected_by='automated'):
        """Generate an evidence artifact and stamp it back to the catalog.

        Returns dict: {evidence_name, file_path, record_count, stamped, control_id}.
        """
        file_path = self.generate_evidence_file_by_name(evidence_name)
        if not file_path or not os.path.exists(file_path):
            return {
                'evidence_name': evidence_name,
                'file_path': None,
                'stamped': False,
                'record_count': None,
                'control_id': None,
            }
        evidence_item = self.stamp_evidence_collection(evidence_name, file_path, collected_by=collected_by)
        return {
            'evidence_name': evidence_name,
            'file_path': file_path,
            'record_count': self._count_data_rows(file_path),
            'stamped': bool(evidence_item),
            'control_id': evidence_item.control_id if evidence_item else None,
        }

    def generate_evidence_file_by_name(self, evidence_name):
        """Generate evidence file based on evidence name"""
        # Map evidence names to generation functions
        evidence_map = {
            # --- Admin Access (catalog: per-layer) ----------------------
            # Backed by admin_role_snapshot (Entra directory roles) +
            # m365_user.is_admin. All four layers map to the same identity
            # population (a single Entra tenant is the IdP for every layer).
            'Administrator Access to Application': self.generate_admin_access_file,
            'Administrator Access to Database': self.generate_admin_access_file,
            'Administrator Access to Network/Cloud': self.generate_admin_access_file,
            'Administrator Access to Operating System': self.generate_admin_access_file,

            # --- User Lists (catalog names) -----------------------------
            'Application User List': self.generate_m365_users_file,
            'Database User List': self.generate_m365_users_file,
            'Network/Cloud User List': self.generate_m365_users_file,
            'Operating System User List': self.generate_m365_users_file,

            # --- Assets & Devices ---------------------------------------
            # Asset register enriched with Intune compliance/encryption.
            'Asset Inventory': self.generate_asset_inventory_file,

            # --- Provisioning (control 97) ------------------------------
            # New-hire access request: samples the latest approved onboard
            # (HR requested → IT approved + provisioned at /approvals).
            'Access Request - New Hire': self.generate_access_request_newhire_file,

            # --- Antivirus (catalog: per-layer) -------------------------
            # Live endpoint protection state from RMM telemetry.
            'Antivirus Configuration - Server': self.generate_rmm_antivirus_file,
            'Antivirus Configuration - Workstation': self.generate_rmm_antivirus_file,

            # --- Encryption ---------------------------------------------
            'Device Disk Encryption': self.generate_disk_encryption_file,

            # Azure Evidence - Network Security
            'Firewall Rules': self.generate_azure_nsg_file,
            'Current Network Diagram': self.generate_azure_network_topology_file,
            'Intrusion Detection System Configuration': self.generate_azure_security_alerts_file,
            'Monitoring Tools Enabled (legacy)': self.generate_azure_monitor_alerts_file,

            # Azure Evidence - Database Security
            'Database Encryption': self.generate_azure_databases_file,
            'SQL Server Database Encryption Configuration': self.generate_azure_databases_file,

            # Azure Evidence - Storage Security
            'Azure Storage Encryption Configuration': self.generate_azure_storage_file,

            # Azure Evidence - Server Security
            'Server Disk Encryption Configuration': self.generate_azure_vms_file,

            # --- Settings evidence (auto-collected) ---------------------
            # Live-config "screenshot" reports (actual current values) where a
            # real data source exists; ISMS documented-standard PDFs where the
            # control's authoritative_docs resolve to a manual section.
            #
            # LIVE config reports:
            'Server Encryption': self.generate_server_encryption_file,                       # azure_vm + Intune servers
            'Separation of Environments': self.generate_separation_of_environments_file,     # azure_vm by resource group
            'Monitoring Tools Enabled': self.generate_monitoring_tools_file,                 # Tracker monitoring subsystem
            'Performance Monitoring Alert Configuration': self.generate_monitoring_alert_config_file,
            'Password Settings - Network/Cloud': self.generate_m365_password_settings_file,  # live Entra/Graph
            #
            # ISMS documented-standard reports (resolved via the control's
            # authoritative_docs IS-IDs -> generate_isms_section_pdf):
            'Encryption in Transit': self.generate_isms_section_pdf,                         # Cryptography Policy
            'Intrusion Detection Configuration': self.generate_isms_section_pdf,             # Logging & Monitoring
            'Security Configuration Standards': self.generate_isms_section_pdf,              # Operations Security
            'Password Settings - Application': self.generate_isms_section_pdf,               # Access Control Policy
            'Password Settings - Database': self.generate_isms_section_pdf,                  # Access Control Policy
            'Password Settings - Operating System': self.generate_isms_section_pdf,          # Access Control Policy
            # 'Merge SOD Configuration Check' has no documented standard and no
            # live source (no GitLab integration) -> intentionally MANUAL.
            
            # --- Vulnerability & Patching (RMM-backed) ------------------
            # Live local data: vulnerability_cache + cve_patch_job + rmm_agent.
            'Vulnerability Scan Results': self.generate_rmm_vulnerability_scan_file,
            'Vulnerability Remediation': self.generate_rmm_vulnerability_remediation_file,
            'Patch Scan': self.generate_rmm_patch_scan_file,
            'Server Scan and Patch': self.generate_rmm_patch_scan_file,
            
            # Microsoft Defender Evidence - Security Events (NEW)
            'Security Incident Report': self.generate_security_incidents_file,
            'Security Incident History': self.generate_security_incidents_file,
            'Security Incident Resolution': self.generate_security_incidents_file,
            'Security Alert History': self.generate_security_alerts_file,
            'Security Alert Report': self.generate_security_alerts_file,
            'Security Event Log': self.generate_security_alerts_file,
            
            # M365 Evidence - MFA & Conditional Access (NEW)
            'MFA Status Report': self.generate_mfa_status_file,
            'Multi-Factor Authentication Report': self.generate_mfa_status_file,
            'User Authentication Report': self.generate_mfa_status_file,
            'Conditional Access Policy Report': self.generate_conditional_access_file,
            'Conditional Access Policies': self.generate_conditional_access_file,
            'Authentication Policy': self.generate_conditional_access_file,
            
            # Azure Evidence - RBAC (NEW)
            'Azure RBAC Report': self.generate_azure_rbac_file,
            'Azure Role Assignments': self.generate_azure_rbac_file,
            'Cloud Access Control': self.generate_azure_rbac_file,
            'Privileged Access Report': self.generate_azure_rbac_file,
            
            # Phase 2 Evidence - Software & Updates (NEW)
            'Software Inventory by Asset': self.generate_software_inventory_by_asset_file,
            'Application Inventory': self.generate_software_inventory_by_asset_file,
            'Installed Software Report': self.generate_software_inventory_by_asset_file,
            'System Updates Report': self.generate_system_updates_file,
            'Missing Hotfixes': self.generate_system_updates_file,
            'Windows Update Status': self.generate_system_updates_file,
            'Patch Status Report': self.generate_system_updates_file,
            
            # Phase 2 Evidence - Security Baseline (NEW)
            'Security Baseline Compliance': self.generate_security_baseline_file,
            'Secure Score Report': self.generate_security_baseline_file,
            'Security Configuration Assessment': self.generate_security_baseline_file,
            'Cloud Security Posture': self.generate_security_baseline_file,
            
            # Phase 2 Evidence - Key Vault (NEW)
            'Key Vault Access Policies': self.generate_key_vault_policies_file,
            'Secret Management Policies': self.generate_key_vault_policies_file,
            'Encryption Key Access': self.generate_key_vault_policies_file,
            
            # Phase 2 Evidence - Network Logs (NEW)
            'Network Traffic Logs': self.generate_network_traffic_logs_file,
            'NSG Flow Logs': self.generate_network_traffic_logs_file,
            'Network Monitoring Configuration': self.generate_network_traffic_logs_file,
            'Traffic Analysis Report': self.generate_network_traffic_logs_file,
            
            # ISMS Policy Evidence (PDF generated from the published ISMS Manual,
            # extracted by IS-section ID -> see generate_isms_section_pdf). These
            # are the Policy-type StrikeGraphEvidence items that resolve to a real
            # manual section. "Code of Conduct" is intentionally absent: it lives
            # in the Employee Handbook, not the ISMS manual.
            # 'Acceptable Use Policy', 'Business Continuity Plan', and 'Vendor
            # Management Policy and Procedures' are intentionally absent: they
            # were reverted to manual/HR-sourced evidence (see
            # POLICY_EVIDENCE_REVERTED_TO_MANUAL). Even via the Policy fallback
            # below, the resolver short-circuits them to "not collected".
            'Access Removal Procedures/Checklist': self.generate_isms_section_pdf,
            'Backup Policy': self.generate_isms_section_pdf,
            'Backup Restoration Procedures': self.generate_isms_section_pdf,
            'Change Management Policy': self.generate_isms_section_pdf,
            'Data Classification Policy': self.generate_isms_section_pdf,
            'Data Management Policy': self.generate_isms_section_pdf,
            'Incident Response Plan': self.generate_isms_section_pdf,
            'Information Security Policy': self.generate_isms_section_pdf,
            'Logical Access Policy and Procedures': self.generate_isms_section_pdf,
            'Password Policy': self.generate_isms_section_pdf,
            'Patch Management Policy': self.generate_isms_section_pdf,
            'Record Retention Schedule': self.generate_isms_section_pdf,
            'Risk Management Policy and Procedures': self.generate_isms_section_pdf,
            'System Description Document': self.generate_isms_section_pdf,
            'Vulnerability Management Policy': self.generate_isms_section_pdf,

            # Employee Handbook Evidence (PDF generation)
            '1-1. Welcome Statement': self.generate_employee_handbook_pdf,
            '1-6. Non-Disclosure Employee Assignment Agreements': self.generate_employee_handbook_pdf,
            '2-1. Employee Classifications': self.generate_employee_handbook_pdf,
            '2-3. Employment and Personnel Records': self.generate_employee_handbook_pdf,
            '2-10. Performance Reviews': self.generate_employee_handbook_pdf,
            '3-20. Tuition Reimbursement': self.generate_employee_handbook_pdf,
            '5-1. Workplace Conduct': self.generate_employee_handbook_pdf,
            '5-3. Use of Communication, Computer Systems and equipment': self.generate_employee_handbook_pdf,
            '5-9. Non-Disclosure of Confidential Information': self.generate_employee_handbook_pdf,
            '5-22. If You Must Leave Us': self.generate_employee_handbook_pdf,
            
            # M365 User Lists (new additions)
            'Organization Chart': self.generate_m365_users_file,
            # NOTE: the four 'Password Settings - *' items are wired in the
            # "Settings evidence" block above (live Entra for Network/Cloud;
            # ISMS Access Control Policy for Application/Database/Operating
            # System). The previous static generate_m365_password_policy_file
            # routing was removed because it emitted hard-coded placeholder
            # values rather than real configuration.
        }
        
        generator_func = evidence_map.get(evidence_name)
        if generator_func:
            return generator_func(evidence_name)

        evidence_item = StrikeGraphEvidence.query.filter_by(evidence_name=evidence_name).first()
        if evidence_item and (evidence_item.automation_source == 'ISMS'
                              or evidence_item.evidence_type == 'Policy'):
            return self.generate_isms_section_pdf(evidence_name)
        return None
    
    def generate_all_automated_evidence_files(self):
        """Generate files for all automated evidence items"""
        evidence_items = StrikeGraphEvidence.query.filter(
            StrikeGraphEvidence.automation_source.in_(['M365/Intune', 'M365/Defender', 'Azure', 'ISMS', 'TeamViewer'])
        ).all()
        
        results = []
        for item in evidence_items:
            try:
                file_path = self.generate_evidence_file_by_name(item.evidence_name)
                if file_path:
                    # Stamp the catalog + write the audit snapshot (commits).
                    self.stamp_evidence_collection(item.evidence_name, file_path)
                    results.append({
                        'evidence_name': item.evidence_name,
                        'status': 'success',
                        'file_path': file_path,
                        'source': item.automation_source
                    })
                else:
                    results.append({
                        'evidence_name': item.evidence_name,
                        'status': 'skipped',
                        'reason': 'No generator function',
                        'source': item.automation_source
                    })
            except Exception as e:
                results.append({
                    'evidence_name': item.evidence_name,
                    'status': 'error',
                    'error': str(e),
                    'source': item.automation_source
                })
        
        db.session.commit()
        return results
