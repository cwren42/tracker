"""Generate the ALAP ISMS Management Ledgers workbook from live Tracker data.

The ledger is an ALAP-supplied Excel template ([IS-APM02-F02..F09]) that Cirque
files a few times a year. Every sheet is a real Excel Table (``Info``,
``Supplement``, ``Worker_List``, ``SoftWare``, ...) with calculated columns,
cross-sheet XLOOKUPs and dropdowns bound to the ``Masters`` sheet, so this
module *fills the template in place* rather than building a workbook from
scratch: it writes values into the table body, extends each table ``ref``, and
copies the calculated-column formulas down.

Rows already present in the uploaded template carry ISMS judgement calls
Tracker does not hold (risk scoring, protect class, purpose text). Those are
preserved per row and only Tracker-owned fields are refreshed. Rows whose
underlying asset/worker is gone from Tracker are kept and stamped
retired/departed in Remarks rather than dropped.
"""

from __future__ import annotations

import io
import re
from copy import copy
from datetime import date, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import MultiCellRange
from fnmatch import fnmatch

from extensions import db


LEDGER_TEMPLATE_DIR = Path(__file__).resolve().parent / 'content' / 'isms' / 'ledgers'

# Header rows: r1 title, r2 metadata, r4/r5 headers, data starts at r6.
FIRST_DATA_ROW = 6

# Categories that belong on F04C. Deliberately excludes UniFi switches/APs,
# cameras, tablets and phone records -- matches the scope of the prior filing.
SUPPLEMENT_CATEGORIES = ('Laptop', 'Desktop', 'Workstation', 'Server', 'Computer', 'Mini PC')

RETIRED_ASSET_STATUSES = ('Retired', 'Disposed')

# An asset only reaches F04C if something has seen it within this window.
# "Seen" is the newest of the RMM heartbeat, the UniFi client record, the Intune
# sync and the AD last-logon -- an asset with no agent still counts as live if
# any other source has it. Assets with no signal from any source are excluded
# rather than carried, so the ledger reflects kit actually in service.
ONLINE_WITHIN_DAYS = 180

# Assets to keep off the ledger regardless of how live they are -- personal
# machines, bench/test boxes, anything not a company information asset. Ledger
# scope only: the asset stays untouched in Tracker. Case-insensitive, matched
# against asset name or tag, and fnmatch globs are allowed (e.g. "IT-NUC*").
# Editable at runtime via the `isms_ledger_excluded_assets` setting, one entry
# per line, which replaces this default when set.
LEDGER_EXCLUDED_ASSETS_DEFAULT = ('ChrisHome', 'ITWORKBENCH')
LEDGER_EXCLUSION_SETTING = 'isms_ledger_excluded_assets'

# Cloud services carried on F04C rather than in Tracker's asset register. ALAP's
# F02 supplement is explicit that a cloud service storing customer or derived
# data is registered as an F04 auxiliary asset, so these belong on the sheet.
CLOUD_SERVICE_NAMES = {'OFFICE365', 'ASANA', 'TEAMVIEWER', 'PAYLOCITY'}

# GREATEST ignores NULLs in Postgres, so this collapses to the newest signal
# present. unifi_last_seen is stored as text and is not always a timestamp.
LAST_SIGNAL_SQL = """
        GREATEST(
            a.last_seen,
            (CASE WHEN a.unifi_last_seen ~ '^[0-9]{4}-[0-9]{2}-[0-9]{2}'
                  THEN a.unifi_last_seen::timestamptz END),
            a.intune_last_sync,
            a.ad_last_logon::timestamptz
        )
"""

DEFAULT_SECURITY_MANAGER = 'Chris Wren'

# Masters!M External_Public. The template's own rows read "No", which is not a
# member of that list and fails the sheet's own validation.
EXTERNAL_PUBLIC_DEFAULT = 'Not Publicly Available'
DEFAULT_COMPANY = 'Cirque Corporation'

# Masters!N "Types of Assets" values.
ASSET_TYPE_BY_CATEGORY = {
    'Server': 'Server_Storage server',
    'Computer': 'Server_Storage server',
    'Laptop': 'Physical Object',
    'Desktop': 'Physical Object',
    'Workstation': 'Physical Object',
    'Mini PC': 'Physical Object',
}

PURPOSE_BY_CATEGORY = {
    'Laptop': 'Laptop /Workstation',
    'Desktop': 'Desktop /Workstation',
    'Workstation': 'Laptop /Workstation',
    'Mini PC': 'Desktop /Workstation',
    'Server': 'Server',
    'Computer': 'Server',
}

# Tracker's free-text asset.location values normalised onto the phrasing the
# ledger already uses for installation/storage location.
LOCATION_MAP = {
    'cirque-us': 'Cirque SLC office',
    'cirqueus': 'Cirque SLC office',
    'us': 'Cirque SLC office',
    'cirque us': 'Cirque SLC office',
    'cirque-taiwan': 'Cirque TW office',
    'cirque taiwan': 'Cirque TW office',
    'taiwan': 'Cirque TW office',
    'cirqueasia': 'Cirque TW office',
    'cirque-china': 'Cirque China office',
    'china': 'Cirque China office',
    'domain controllers': 'Cirque SLC server room',
    'server room': 'Cirque SLC server room',
    'cirque data center': 'Cirque SLC server room',
    'storage': 'Cirque SLC IT storage',
}

# Tracker license_type -> Masters!Q License_Type.
LICENSE_TYPE_MAP = {
    'subscription': 'Subscription',
    'retail': 'One-time Purchase',
    'per device': 'One-time Purchase',
    'perpetual': 'One-time Purchase',
    'open source': 'OSS(Open Source Software)',
}

# Tracker employee.work_type -> Masters!P Worker_Type.
WORKER_TYPE_MAP = {
    'local': 'full-time employee',
    'remote': 'full-time employee',
    'contractor': 'On-premises Outsourcing Contract',
    'temp': 'Temporary Staffing Agreement',
}

SHEET_SUPPLEMENT = 'F04C Asset- Supplement'
SHEET_WORKERS = 'F09A  Worker List'
SHEET_PARTNERS = 'F06B BusinessPartner'
SHEET_SOFTWARE = 'F08A Non-Std Software'
SHEET_WORKER_ACCESS = 'F09-A'
SHEET_ASSET_AREA = 'F02-A'
SHEET_PROTOTYPE = 'F03A Asset - Prototype'
SHEET_SOFTWARE_USERS = 'F09-B'
SHEET_INFO = 'F02B Asset - Info'

DATE_FORMAT = 'yyyy/mm/dd'

# Excel serial 55153 -- the placeholder "no expiry" contract date already used
# throughout the Worker List. Kept so preserved and generated rows agree.
CONTRACT_EXPIRE_PLACEHOLDER = date(2050, 12, 31)


# --------------------------------------------------------------------------
# template discovery
# --------------------------------------------------------------------------

def find_template():
    """Return the newest ledger template on disk, or None."""
    if not LEDGER_TEMPLATE_DIR.is_dir():
        return None
    candidates = [
        path for path in LEDGER_TEMPLATE_DIR.glob('*.xlsx')
        if not path.name.startswith('~$')
    ]
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def template_info():
    path = find_template()
    if path is None:
        return None
    stat = path.stat()
    return {
        'path': path,
        'name': path.name,
        'size_kb': round(stat.st_size / 1024),
        'modified': date.fromtimestamp(stat.st_mtime),
    }


# --------------------------------------------------------------------------
# small helpers
# --------------------------------------------------------------------------

def _clean(value):
    if value is None:
        return ''
    return str(value).strip()


def _key(value):
    return _clean(value).upper()


def _set(worksheet, row, column, value, *, number_format=None):
    cell = worksheet.cell(row=row, column=column)
    cell.value = value
    if number_format:
        cell.number_format = number_format
    return cell


def _set_date(worksheet, row, column, value):
    """Write a real date. The template stores raw Excel serials with a General
    format, which renders as ``45512``; normalising every date cell we own is
    strictly more readable and Excel reads it the same."""
    if value is None:
        return _set(worksheet, row, column, None)
    return _set(worksheet, row, column, value, number_format=DATE_FORMAT)


def _existing_date(cell_value):
    """Read a date back out of a template cell that may hold a real date, an
    Excel serial number, or a serial rendered as text."""
    if cell_value is None or cell_value == '':
        return None
    if isinstance(cell_value, date):
        return cell_value
    try:
        serial = int(float(str(cell_value).strip()))
    except (TypeError, ValueError):
        return None
    if serial <= 0:
        return None
    return date(1899, 12, 30) + timedelta(days=serial)


def _table_for(worksheet):
    """Return (name, table) for the sheet's single ListObject, or (None, None)."""
    for name in worksheet.tables:
        return name, worksheet.tables[name]
    return None, None


def _blank_row(worksheet, row, last_column):
    for column in range(1, last_column + 1):
        worksheet.cell(row=row, column=column).value = None


def _clone_row_style(worksheet, source_row, target_row, last_column):
    """Copy formatting + calculated-column formulas from a template row."""
    for column in range(1, last_column + 1):
        source = worksheet.cell(row=source_row, column=column)
        target = worksheet.cell(row=target_row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        if isinstance(source.value, str) and source.value.startswith('='):
            target.value = _shift_formula(source.value, source_row, target_row)


_ROW_REF = re.compile(r'(?<![A-Za-z0-9_$])(\$?)([A-Z]{1,3})(\$?)(\d+)')


def _shift_formula(formula, source_row, target_row):
    """Shift relative row references so a copied formula matches its new row.

    Structured references (``Table[[#This Row],[Col]]``) are row-relative
    already, so only plain A1 refs need adjusting.
    """
    delta = target_row - source_row
    if not delta:
        return formula

    def replace(match):
        col_abs, col, row_abs, row = match.groups()
        if row_abs:
            return match.group(0)
        return f'{col_abs}{col}{row_abs}{int(row) + delta}'

    # Protect structured references from the A1 regex.
    parts = re.split(r'(\[[^\]]*\])', formula)
    return ''.join(
        part if part.startswith('[') else _ROW_REF.sub(replace, part)
        for part in parts
    )


# --------------------------------------------------------------------------
# ALAP template revisions
#
# ALAP issued a revision list against these ledgers (2026-08). Each item is the
# same class of defect: row 6 is right and rows 7+ were dragged out of step.
# They are applied at generation time rather than baked into the stored
# template, so they hold for whichever template revision is uploaded and
# re-applying them to an already-corrected workbook is a no-op.
# --------------------------------------------------------------------------

# Revision 1 -- Importance must be the max of the row's own C/I/A triple.
# The template ships it as a drag artifact spanning the whole table body
# (``MAX(S6:U95)`` on F02B, ``MAX(U6:W100)`` on F04C), which scales with the
# table: one asset scored 4 drags every row above it to 4. ALAP's own wording:
# "Incorrect: =MAX(L6:N[highest row]) / Correct: =MAX(L6:N6)".
IMPORTANCE_FIXUPS = {
    SHEET_INFO: {'column': 22, 'formula': '=MAX(S{row}:U{row})'},          # F02, V
    SHEET_PROTOTYPE: {'column': 22, 'formula': '=MAX(S{row}:U{row})'},     # F03, V
    SHEET_SUPPLEMENT: {'column': 24, 'formula': '=MAX(U{row}:W{row})'},    # F04, X
    SHEET_SOFTWARE: {'column': 15, 'formula': '=MAX(L{row}:N{row})'},      # F08, O
}


def _fix_importance_column(worksheet, last_data_row):
    fixup = IMPORTANCE_FIXUPS.get(worksheet.title)
    if not fixup:
        return

    column = fixup['column']
    for row in range(FIRST_DATA_ROW, last_data_row + 1):
        worksheet.cell(row=row, column=column).value = fixup['formula'].format(row=row)

    # Excel re-broadcasts a table's stored calculated-column formula on open, so
    # the stored copy has to be corrected too or the cells revert.
    _name, table = _table_for(worksheet)
    if table is None:
        return
    for table_column in table.tableColumns:
        if table_column.name == 'Importance' and table_column.calculatedColumnFormula is not None:
            table_column.calculatedColumnFormula.attr_text = (
                fixup['formula'].format(row=FIRST_DATA_ROW).lstrip('=')
            )


def _replace_validation(worksheet, match_formula, new_sqref, new_formula=None, *, drop=False):
    """Retarget (or remove) the data validation whose formula1 matches.

    openpyxl keeps validations as a flat list keyed by their ``sqref`` range,
    so a revision is expressed by rewriting that range rather than by editing
    individual cells.
    """
    validations = worksheet.data_validations.dataValidation
    for validation in list(validations):
        if _clean(validation.formula1) != match_formula:
            continue
        if drop:
            validations.remove(validation)
            return True
        validation.sqref = MultiCellRange(new_sqref)
        if new_formula is not None:
            validation.formula1 = new_formula
        return True
    return False


def _apply_validation_revisions(workbook):
    """Revisions 2 and 3 -- data-validation ranges that skip a column or bind
    the wrong list from row 7 down."""
    applied = []

    # Revision 2 [F02-A]: the "Area select" dropdown sits in column E on row 6
    # but column F from row 7 down, so the dependent INDIRECT lookup is one
    # column out too. Pull rows 7+ back into line with row 6 and drop the
    # stray dependent validation left in column G. ALAP's instruction is the
    # manual equivalent: delete E7 downward and shift the cells left.
    area = workbook[SHEET_ASSET_AREA]
    if _replace_validation(area, 'Store_Area', 'E6:E1048576'):
        applied.append('F02-A: Area select realigned to column E for all rows')
    if _replace_validation(area, 'INDIRECT($E6)', 'F6:F1048576'):
        applied.append('F02-A: dependent storage-location lookup realigned to column F')
    if _replace_validation(area, 'INDIRECT($F7)', None, drop=True):
        applied.append('F02-A: removed the stray column-G lookup left by the shift')

    # Revision 3 [F09-A][F09-B]: column A row 6 validates against Name_Mail
    # (``Name(email)``), but rows 7+ validate against Worker_Name (bare name).
    # The XLOOKUPs in both sheets key on Name (Mail), so the row-7+ rule
    # rejects the only value that actually resolves -- and it is the value this
    # generator writes.
    for sheet_name in (SHEET_WORKER_ACCESS, SHEET_SOFTWARE_USERS):
        worksheet = workbook[sheet_name]
        # Drop row 6's standalone Name_Mail rule first, so widening the
        # Worker_Name rule to cover column A does not leave two overlapping
        # validations on A6.
        _replace_validation(worksheet, 'Name_Mail', None, drop=True)
        if _replace_validation(worksheet, 'Worker_Name', 'A6:A1048576', 'Name_Mail'):
            applied.append(f'{sheet_name}: column A now validates against Name_Mail for all rows')

    return applied


def _resize_table(worksheet, last_data_row):
    """Extend/shrink the sheet's ListObject to cover the written rows."""
    name, table = _table_for(worksheet)
    if table is None:
        return
    start, end = table.ref.split(':')
    start_col = re.match(r'([A-Z]+)', start).group(1)
    end_col = re.match(r'([A-Z]+)', end).group(1)
    header_row = int(re.search(r'(\d+)', start).group(1))
    # A table must always contain at least one data row.
    final_row = max(last_data_row, header_row + 1)
    table.ref = f'{start_col}{header_row}:{end_col}{final_row}'
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def _write_rows(worksheet, rows, last_column, *, template_row=FIRST_DATA_ROW):
    """Write ``rows`` (list of {column_index: value_or_(value, fmt)}) starting at
    FIRST_DATA_ROW, cloning style/formulas from ``template_row``, then resize
    the table and clear any leftover rows below."""
    previous_last = worksheet.max_row
    style_source = template_row

    for offset, row_values in enumerate(rows):
        row = FIRST_DATA_ROW + offset
        if row != style_source:
            _clone_row_style(worksheet, style_source, row, last_column)
        for column, value in row_values.items():
            if isinstance(value, tuple):
                _set(worksheet, row, column, value[0], number_format=value[1])
            elif isinstance(value, date):
                _set_date(worksheet, row, column, value)
            else:
                _set(worksheet, row, column, value)

    last_data_row = FIRST_DATA_ROW + len(rows) - 1 if rows else FIRST_DATA_ROW - 1
    for row in range(max(last_data_row + 1, FIRST_DATA_ROW), previous_last + 1):
        _blank_row(worksheet, row, last_column)

    _resize_table(worksheet, last_data_row)
    _fix_importance_column(worksheet, last_data_row)
    return len(rows)


def _snapshot_rows(worksheet, key_column, last_column):
    """Index the template's existing data rows by a key column so curated
    values can be carried forward."""
    snapshot = {}
    for row in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        key = _key(worksheet.cell(row=row, column=key_column).value)
        if not key:
            continue
        snapshot[key] = {
            column: worksheet.cell(row=row, column=column).value
            for column in range(1, last_column + 1)
        }
    return snapshot


def _carry(existing, column, fallback=None):
    """Prefer the curated template value; fall back to a Tracker-derived one."""
    if existing:
        value = existing.get(column)
        if value not in (None, '') and not (isinstance(value, str) and value.startswith('=')):
            return value
    return fallback


# --------------------------------------------------------------------------
# Tracker queries
# --------------------------------------------------------------------------

def _fetch(sql, params=None):
    return db.session.execute(db.text(sql), params or {}).mappings().all()


def ledger_exclusions():
    """Patterns for assets held off the ledger. Setting overrides the default."""
    from models import Setting
    row = Setting.query.filter_by(key=LEDGER_EXCLUSION_SETTING).first()
    if row is not None and _clean(row.value):
        patterns = [line.strip() for line in str(row.value).replace(',', '\n').splitlines()]
        return tuple(p for p in patterns if p)
    return LEDGER_EXCLUDED_ASSETS_DEFAULT


def _is_excluded(asset, patterns):
    name, tag = _key(asset['name']), _key(asset['asset_tag'])
    for pattern in patterns:
        candidate = _key(pattern)
        if name == candidate or (tag and tag == candidate):
            return True
        if fnmatch(name, candidate) or (tag and fnmatch(tag, candidate)):
            return True
    return False


def _load_assets(*, stale=False):
    """In-scope assets, split on whether anything has seen them recently.

    ``stale=False`` returns the assets that belong on F04C; ``stale=True``
    returns the ones excluded for having no signal inside ONLINE_WITHIN_DAYS
    (including those that have never reported from any source).
    """
    comparison = '<' if stale else '>='
    null_clause = 'OR {sig} IS NULL'.format(sig=LAST_SIGNAL_SQL) if stale else ''
    rows = _fetch(
        """
        SELECT a.id, a.asset_tag, a.name, a.category, a.status, a.location,
               a.serial_number, a.model, a.manufacturer, a.os_version,
               e.name AS employee_name, e.department AS employee_department,
               {signal} AS last_signal
        FROM asset a
        LEFT JOIN employee e ON e.id = a.employee_id
        WHERE a.category IN :categories
          AND ({signal} {comparison} now() - make_interval(days => :window) {null_clause})
        ORDER BY a.name
        """.format(
            signal=LAST_SIGNAL_SQL,
            comparison=comparison,
            null_clause=null_clause,
        ),
        {'categories': tuple(SUPPLEMENT_CATEGORIES), 'window': ONLINE_WITHIN_DAYS},
    )
    patterns = ledger_exclusions()
    if not patterns:
        return rows
    return [row for row in rows if not _is_excluded(row, patterns)]


def _load_employees():
    return _fetch(
        """
        SELECT e.id, e.name, e.email, e.sam_account_name, e.department,
               e.job_title, e.work_type, e.start_date, e.location,
               e.ad_enabled, e.offboarded_at
        FROM employee e
        WHERE e.is_visible = TRUE AND e.offboarded_at IS NULL
        ORDER BY e.name
        """
    )


def _load_training():
    rows = _fetch(
        """
        SELECT trainee_name, trainee_email, employee_id,
               MAX(training_date) AS last_training,
               MAX(completion_status) AS completion_status
        FROM soc2_security_training_record
        WHERE completion_status ILIKE 'complete%'
        GROUP BY trainee_name, trainee_email, employee_id
        """
    )
    by_employee, by_email, by_name = {}, {}, {}
    for row in rows:
        if row['employee_id']:
            by_employee[row['employee_id']] = row
        if row['trainee_email']:
            by_email[row['trainee_email'].strip().lower()] = row
        if row['trainee_name']:
            by_name[_key(row['trainee_name'])] = row
    return by_employee, by_email, by_name


def _load_assigned_devices():
    """employee_id -> {'pc': asset_tag_or_name, 'other': [...]} for F09A Q/R/S."""
    rows = _fetch(
        """
        SELECT a.employee_id, a.asset_tag, a.name, a.category, a.status
        FROM asset a
        WHERE a.employee_id IS NOT NULL
          AND (a.status IS NULL OR a.status NOT IN :retired)
        ORDER BY a.employee_id, {signal} DESC NULLS LAST, a.name
        """.format(signal=LAST_SIGNAL_SQL),
        {'retired': RETIRED_ASSET_STATUSES},
    )
    devices = {}
    for row in rows:
        bucket = devices.setdefault(row['employee_id'], {'pc': [], 'other': []})
        label = _clean(row['asset_tag']) or _clean(row['name'])
        if row['category'] in SUPPLEMENT_CATEGORIES:
            bucket['pc'].append((label, _clean(row['name'])))
        else:
            bucket['other'].append(label)
    return devices


def _load_asset_operations():
    """Per-asset operational evidence for F04C's System Operation Record and
    inspection columns: last patch installed, last Defender sync, and open
    Critical/High findings."""
    rows = _fetch(
        """
        SELECT a.id,
               (SELECT MAX(p.installed_on)::date
                  FROM rmm_patch p JOIN rmm_agent g ON g.agent_id = p.agent_id
                 WHERE g.asset_id = a.id) AS last_patch,
               (SELECT MAX(v.synced_at)::date
                  FROM device_vulnerability v WHERE v.asset_id = a.id) AS last_scan,
               (SELECT COUNT(*) FROM device_vulnerability v
                 WHERE v.asset_id = a.id AND v.status = 'Open'
                   AND v.severity = 'Critical') AS open_critical,
               (SELECT COUNT(*) FROM device_vulnerability v
                 WHERE v.asset_id = a.id AND v.status = 'Open'
                   AND v.severity = 'High') AS open_high
        FROM asset a
        WHERE a.category IN :categories
        """,
        {'categories': tuple(SUPPLEMENT_CATEGORIES)},
    )
    return {row['id']: row for row in rows}


def _next_quarter_start(today=None):
    today = today or date.today()
    quarter_month = ((today.month - 1) // 3 + 1) * 3 + 1
    if quarter_month > 12:
        return date(today.year + 1, 1, 1)
    return date(today.year, quarter_month, 1)


# F04C column M. ALAP: "Restrictions on members and locations authorized to use
# the asset, etc."
def _permitted_scope(asset, is_cloud=False):
    if is_cloud:
        return ('Licensed Cirque users only; access via corporate identity '
                '(Entra ID) with multi-factor authentication.')
    if asset['category'] in ('Server', 'Computer'):
        return ('IT administrators only. Cirque server room; remote administration '
                'over the company VPN.')
    if not asset['employee_name']:
        return 'IT department only; held as unissued stock at a Cirque site.'
    return ('Assigned user only. Cirque offices and approved remote locations '
            'over the company VPN.')


# F04C column AD. Factual record of the operational activity actually run
# against the asset -- ALAP asks for "records of activities such as
# vulnerability response".
def _system_operation_record(operations):
    parts = []
    last_patch = operations.get('last_patch') if operations else None
    last_scan = operations.get('last_scan') if operations else None
    if last_patch:
        parts.append(f'Patch management via CirqueRMM agent; last update installed {last_patch}.')
    else:
        parts.append('Patch management via CirqueRMM agent.')
    if last_scan:
        parts.append(
            f'Microsoft Defender vulnerability assessment, last synchronised {last_scan}: '
            f"{operations['open_critical']} open Critical, {operations['open_high']} open High, "
            'tracked to remediation.')
    else:
        parts.append('Not yet enrolled in Defender vulnerability assessment.')
    return ' '.join(parts)


def _load_vendors():
    return _fetch(
        """
        SELECT vendor_name, service_description, vendor_type, criticality,
               owner, contract_status, last_review_date, next_review_date,
               data_access_scope, assurance_status, notes,
               nda_executed_date, isms_notified_date, contact_department,
               onsite_access_scope, required_availability, training_required,
               data_return_on_termination
        FROM soc2_vendor
        WHERE is_active = TRUE
        ORDER BY vendor_name
        """
    )


def _load_licenses():
    return _fetch(
        """
        SELECT l.id, l.software_name, l.vendor, l.license_type, l.total_licenses,
               l.status, l.notes, l.version,
               (SELECT s.version FROM rmm_software s
                 WHERE s.name = l.software_name AND s.version IS NOT NULL
                 GROUP BY s.version ORDER BY COUNT(*) DESC LIMIT 1) AS common_version
        FROM license l
        WHERE l.status IS NULL OR l.status <> 'Retired'
        ORDER BY l.software_name
        """
    )


def _load_partner_system_accounts():
    """Accounts Cirque staff hold on a business-partner-provided system.

    ALAP asks for these on F09-A: the worker, the partner system they were
    instructed to use, and their account on it. Tracker only holds this for
    Microsoft 365 (`m365_user`, linked to an employee and to the Microsoft row
    on F06B). Accounts on the other partner platforms -- GitLab, Cadence,
    Altium, Cliosoft, Arena/Omnify -- and on any customer-provided system are
    not recorded anywhere in Tracker, so they cannot be emitted here.
    """
    return _fetch(
        """
        SELECT v.vendor_name,
               e.name AS employee_name, e.email, e.sam_account_name,
               e.department, e.start_date,
               m.user_principal_name, m.is_admin
        FROM m365_user m
        JOIN employee e ON e.id = m.employee_id
        JOIN soc2_vendor v ON v.vendor_name = 'Microsoft' AND v.is_active = TRUE
        WHERE m.is_current = TRUE
          AND m.account_enabled = TRUE
          AND e.is_visible = TRUE
          AND e.offboarded_at IS NULL
        ORDER BY e.name
        """
    )


def _load_license_assignments():
    return _fetch(
        """
        SELECT l.software_name, e.name AS employee_name, e.email, e.sam_account_name,
               la.assigned_date, la.status
        FROM license_assignment la
        JOIN license l ON l.id = la.license_id
        JOIN employee e ON e.id = la.employee_id
        WHERE la.status IS NULL OR la.status = 'Active'
        ORDER BY l.software_name, e.name
        """
    )


# --------------------------------------------------------------------------
# derived values
# --------------------------------------------------------------------------

def _installation_location(asset):
    location = _clean(asset['location'])
    mapped = LOCATION_MAP.get(location.lower())
    if mapped:
        if asset['category'] in ('Laptop', 'Desktop', 'Workstation') and asset['employee_name']:
            return f"{asset['employee_name']}'s Desk"
        return mapped
    if location:
        return location
    if asset['employee_name']:
        return f"{asset['employee_name']}'s Desk"
    return 'Cirque SLC office'


def _name_mail(name, email):
    """Match the Worker_List ``Name (Mail)`` calculated column exactly -- the
    F09-A/F09-B XLOOKUPs key on it."""
    return f'{_clean(name)}({_clean(email)})'


def _planned_training_date(last_training, prior_planned):
    """F09A column N. ALAP: "Required for those who have NOT attended training
    in the past year due to extended leave, etc." -- so anyone current is left
    blank, and only a stale or missing record carries a planned date."""
    if last_training and last_training >= date.today() - timedelta(days=365):
        return None
    if last_training:
        return last_training + timedelta(days=365)
    return prior_planned


def _worker_type(employee):
    return WORKER_TYPE_MAP.get(_clean(employee['work_type']).lower(), 'full-time employee')


def _license_type(value):
    return LICENSE_TYPE_MAP.get(_clean(value).lower(), 'Others')


# --------------------------------------------------------------------------
# sheet builders
# --------------------------------------------------------------------------

def _fill_supplement(worksheet):
    """F04C -- supporting assets (endpoints, servers, lab/conference PCs)."""
    last_column = 37  # AK
    existing = _snapshot_rows(worksheet, 3, last_column)  # keyed on Asset Name
    # Boxes get renamed (offboard -> repurpose), so the same physical asset can
    # sit under an old hostname in the template. The Cirque asset tag is the
    # stable identity; index on it as well so a rename updates the row instead
    # of leaving a duplicate ghost behind.
    existing_by_tag = {}
    for key, snapshot in existing.items():
        tag = _key(snapshot.get(2))
        if tag and tag not in ('0', 'NA'):
            existing_by_tag.setdefault(tag, key)

    assets = _load_assets()

    # Assets excluded for having no signal inside the window. A template row
    # matching one of these must be dropped outright -- falling through to the
    # carry-forward branch would put it straight back on the sheet, mislabelled
    # as "no matching Tracker asset".
    excluded = _load_assets(stale=True)
    exclusion_patterns = ledger_exclusions()
    operations = _load_asset_operations()
    inspection_due = _next_quarter_start()
    # ALAP: "Continuous Defender vulnerability assessment" is the inspection of
    # record. No grade is asserted -- ALAP defines no scoring yardstick, and a
    # self-assigned score would be an invention rather than evidence.
    inspection_result = ('Continuous Microsoft Defender vulnerability assessment; '
                         'findings tracked to remediation.')
    excluded_keys = {_key(asset['name']) for asset in excluded}
    excluded_keys |= {
        _key(asset['asset_tag']) for asset in excluded
        if _key(asset['asset_tag']) not in ('', '0', 'NA')
    }

    rows = []
    matched_keys = set()
    stats = {
        'total': 0, 'new': 0, 'carried': 0, 'retired': 0, 'renamed': 0,
        'excluded_offline': len(excluded),
        'excluded_by_list': 0,
    }

    for asset in assets:
        key = _key(asset['name'])
        prior = existing.get(key)
        if prior is None:
            prior_key = existing_by_tag.get(_key(asset['asset_tag']))
            if prior_key:
                prior = existing[prior_key]
                matched_keys.add(prior_key)
                stats['renamed'] += 1
        if prior is not None:
            matched_keys.add(key)
        else:
            stats['new'] += 1

        retired = _clean(asset['status']) in RETIRED_ASSET_STATUSES
        remarks = _carry(prior, 35)
        if retired:
            note = f"Retired in Tracker (status: {_clean(asset['status'])})"
            remarks = note if not remarks else f'{remarks} | {note}'
            stats['retired'] += 1


        asset_ops = operations.get(asset['id'])
        last_scan = asset_ops.get('last_scan') if asset_ops else None

        rows.append({
            2: _carry(prior, 2),                                   # FY25 ledger number
            3: _clean(asset['name']),
            4: _carry(prior, 4),   # FY25
            5: _carry(prior, 5),   # FY25
            6: _carry(prior, 6),   # FY25 asset manager
            7: _carry(prior, 7, 'Cirque'),
            8: _carry(prior, 8, 'Cirque IT'),   # IT Operations Organization (required)
            9: _carry(prior, 9, PURPOSE_BY_CATEGORY.get(asset['category'], 'IT equipment')),
            10: EXTERNAL_PUBLIC_DEFAULT,   # Masters External_Public; 'No' is not a valid member
            11: _carry(prior, 11, ASSET_TYPE_BY_CATEGORY.get(asset['category'], 'Physical Object')),
            12: _carry(prior, 12),  # FY25
            13: _carry(prior, 13, _permitted_scope(asset)),
            14: _carry(prior, 14),  # FY25
            15: _carry(prior, 15, _installation_location(asset)),
            16: _carry(prior, 16, 'None'),
            17: _carry(prior, 17),
            18: _carry(prior, 18),
            19: _carry(prior, 19),
            20: _carry(prior, 20, 'No'),   # Training Requirement; F09-A auto-displays it
            21: _carry(prior, 21, 2),
            22: _carry(prior, 22, 1),
            23: _carry(prior, 23, 1),
            25: _carry(prior, 25, 2),
            26: _carry(prior, 26, 3),
            29: _carry(prior, 29),   # FY25
            30: _carry(prior, 30, _system_operation_record(asset_ops)),
            31: _existing_date(_carry(prior, 31)) or last_scan,
            32: _carry(prior, 32, inspection_result if last_scan else None),
            33: (_existing_date(_carry(prior, 33))
                 or (last_scan + timedelta(days=90) if last_scan else inspection_due)),
            34: _carry(prior, 34),
            35: remarks,
        })

    # Template rows Tracker no longer knows about -- keep, stamp as retired.
    for key, prior in existing.items():
        if key in matched_keys:
            continue
        if (key in excluded_keys or _key(prior.get(2)) in excluded_keys
                or _is_excluded({'name': key, 'asset_tag': prior.get(2)}, exclusion_patterns)):
            stats['excluded_from_template'] = stats.get('excluded_from_template', 0) + 1
            continue
        remarks = _carry(prior, 35)
        note = 'No matching Tracker asset at generation time - confirm still in service (renamed, retired, or a cloud service held outside the asset register)'
        row = {column: prior.get(column) for column in range(2, last_column + 1)
               if not (isinstance(prior.get(column), str) and str(prior.get(column)).startswith('='))}
        row[35] = note if not remarks else f'{remarks} | {note}'
        row[10] = EXTERNAL_PUBLIC_DEFAULT
        is_cloud = key in CLOUD_SERVICE_NAMES
        row[8] = row.get(8) or 'Cirque IT'
        row[11] = row.get(11) or ('SaaS' if is_cloud else 'Physical Object')
        row[15] = row.get(15) or ('Cloud service environment' if is_cloud else 'Cirque SLC office')
        row[20] = row.get(20) or 'No'
        row[13] = row.get(13) or _permitted_scope(
            {'category': 'Other', 'employee_name': None}, is_cloud=is_cloud)
        row[16] = row.get(16) or ('SLA' if is_cloud else 'None')
        row[30] = row.get(30) or (
            'Vendor-managed service; patching and monitoring performed by the provider under contract.'
            if is_cloud else 'No Tracker operational record at generation time.')
        # No inspection has been performed on these, so no date or result is
        # asserted -- only the scheduled next one. ALAP requires cloud services
        # to be audited, so the gap is left visible rather than filled in.
        row[33] = _existing_date(row.get(33)) or inspection_due
        for column in (31, 33):
            row[column] = _existing_date(row.get(column))
        rows.append(row)
        stats['carried'] += 1

    stats['total'] = _write_rows(worksheet, rows, last_column)
    return stats


def _fill_workers(worksheet):
    """F09A -- worker list, with ISMS training dates and issued equipment."""
    last_column = 26  # Z
    existing = _snapshot_rows(worksheet, 5, last_column)  # keyed on worker name
    employees = _load_employees()
    training_by_employee, training_by_email, training_by_name = _load_training()
    devices = _load_assigned_devices()

    rows = []
    matched_keys = set()
    stats = {'total': 0, 'new': 0, 'departed': 0, 'with_training': 0, 'without_training': 0}

    for employee in employees:
        key = _key(employee['name'])
        prior = existing.get(key)
        if prior:
            matched_keys.add(key)
        else:
            stats['new'] += 1

        training = (
            training_by_employee.get(employee['id'])
            or training_by_email.get(_clean(employee['email']).lower())
            or training_by_name.get(key)
        )
        last_training = training['last_training'] if training else None
        if last_training:
            stats['with_training'] += 1
        else:
            stats['without_training'] += 1

        # ALAP marks Q/R/S as FY25-inventory carry-over, but the annual refresh
        # means the carried laptop is often one the worker no longer has. So:
        # keep the FY25 machine while they still actively hold it, and fall to
        # their newest-seen machine once it is gone (replaced or retired).
        #
        # Picking "newest seen" unconditionally is not an option -- for someone
        # holding several machines that all heartbeat constantly the winner
        # churns between generations, and a ledger column must not move on its
        # own between filings.
        assigned = devices.get(employee['id'], {'pc': [], 'other': []})
        held_tags = {tag for tag, _name in assigned['pc'] if tag}
        fy25_pc = _clean(_carry(prior, 17))
        if fy25_pc and fy25_pc in held_tags:
            pc_label = fy25_pc
        elif assigned['pc']:
            pc_label = assigned['pc'][0][0]
        else:
            pc_label = fy25_pc or 'NA'
        other_label = ', '.join(assigned['other']) if assigned['other'] else _carry(prior, 18, 'NA')

        rows.append({
            3: DEFAULT_COMPANY,
            4: _carry(prior, 4),   # FY25 ALAP employee number
            5: _clean(employee['name']),
            6: _clean(employee['email']),
            8: _carry(prior, 8, _worker_type(employee)),
            9: _clean(employee['sam_account_name']) or _carry(prior, 9),
            10: employee['start_date'] or _existing_date(_carry(prior, 10)),
            11: _existing_date(_carry(prior, 11)) or CONTRACT_EXPIRE_PLACEHOLDER,
            12: last_training or _existing_date(_carry(prior, 12)),
            13: 'Pass' if last_training else _carry(prior, 13),
            14: _planned_training_date(last_training, _existing_date(_carry(prior, 14))),
            15: _existing_date(_carry(prior, 15)),
            16: _carry(prior, 16),  # FY25
            17: pc_label,
            18: other_label,
            19: _carry(prior, 19),  # FY25
            20: _carry(prior, 20, 'NA'),
            21: _carry(prior, 21, 'NA'),
            22: _carry(prior, 22, _clean(employee['name'])),
            23: _carry(prior, 23, 'NA'),
            24: _carry(prior, 24),
        })

    for key, prior in existing.items():
        if key in matched_keys:
            continue
        remarks = _carry(prior, 24)
        note = 'Departed - no active Tracker employee record at generation time'
        row = {column: prior.get(column) for column in range(3, last_column + 1)
               if not (isinstance(prior.get(column), str) and str(prior.get(column)).startswith('='))}
        for column in (10, 11, 12, 14, 15):
            row[column] = _existing_date(row.get(column))
        row[24] = note if not remarks else f'{remarks} | {note}'
        rows.append(row)
        stats['departed'] += 1

    stats['total'] = _write_rows(worksheet, rows, last_column)
    return stats


def _fill_partners(worksheet):
    """F06B -- business partners, from the SOC2 vendor register."""
    last_column = 28  # AB
    # Prior rows put the tool name in D ("Contact Department") and our own
    # company in C. Index on both so either convention matches.
    existing = {}
    for row in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        snapshot = {column: worksheet.cell(row=row, column=column).value
                    for column in range(1, last_column + 1)}
        for column in (3, 4):
            key = _key(snapshot.get(column))
            if key and key != _key(DEFAULT_COMPANY):
                existing.setdefault(key, snapshot)

    vendors = _load_vendors()
    inspection_due = _next_quarter_start()
    rows = []
    matched_keys = set()
    stats = {'total': 0, 'new': 0, 'carried': 0, 'renamed': 0}

    def _match_partner(vendor_key):
        """Prior rows name the tool, not the company ("Cadence Tools" for the
        Cadence vendor), so fall back to word-level containment."""
        if vendor_key in existing:
            return vendor_key, existing[vendor_key]
        for candidate in existing:
            if vendor_key in candidate.split() or candidate.startswith(vendor_key + ' '):
                return candidate, existing[candidate]
        return None, None

    for vendor in vendors:
        key = _key(vendor['vendor_name'])
        matched_key, prior = _match_partner(key)
        if prior:
            matched_keys.add(matched_key)
        else:
            stats['new'] += 1

        rows.append({
            3: _clean(vendor['vendor_name']),
            4: _clean(vendor['contact_department']) or None,
            5: _carry(prior, 5, 'IT'),
            6: _clean(vendor['service_description']) or _carry(prior, 6),
            7: _carry(prior, 7, _clean(vendor['onsite_access_scope'])),   # FY25 first
            8: _carry(prior, 8, DEFAULT_SECURITY_MANAGER),
            9: vendor['nda_executed_date'] or _existing_date(_carry(prior, 9)),
            10: vendor['isms_notified_date'] or _existing_date(_carry(prior, 10)),
            11: vendor['last_review_date'] or _existing_date(_carry(prior, 11)),
            12: _carry(prior, 12, 'Pass' if vendor['assurance_status'] else None),
            13: vendor['next_review_date'] or _existing_date(_carry(prior, 13)),
            14: _carry(prior, 14),  # FY25
            15: _carry(prior, 15, 3 if _clean(vendor['criticality']) in ('Critical', 'High') else 2),
            16: _carry(prior, 16, 1),
            18: _carry(prior, 18),
            19: _existing_date(_carry(prior, 19)),
            20: _existing_date(_carry(prior, 20)),
            21: _carry(prior, 21),
            22: _carry(prior, 22),
            23: _carry(prior, 23, vendor['required_availability'] or 2),
            24: _carry(prior, 24, _clean(vendor['training_required']) or 'Yes'),
            25: _carry(prior, 25, _clean(vendor['data_return_on_termination']) or 'Yes'),
            26: _carry(prior, 26, _clean(vendor['data_access_scope'])),
        })

    for key, prior in existing.items():
        if key in matched_keys:
            continue
        remarks = _carry(prior, 26)
        note = 'Not in Tracker vendor register at generation time'
        row = {column: prior.get(column) for column in range(3, last_column + 1)
               if not (isinstance(prior.get(column), str) and str(prior.get(column)).startswith('='))}
        for column in (9, 10, 11, 13, 19, 20):
            row[column] = _existing_date(row.get(column))
        # FY25 filed these with our own company in "Business Partner Name" and
        # the partner in "Contact Department at Business Partner". ALAP's guide
        # is explicit that column C is the contracting party's official company
        # name, so lift the partner name into C. Neither column is FY25
        # carry-over, so correcting them is in scope.
        partner_name = _clean(row.get(4))
        if partner_name and _key(row.get(3)) == _key(DEFAULT_COMPANY):
            row[3] = partner_name
            row[4] = None
            stats['renamed'] += 1
        # No review is scheduled for these -- they are not in Tracker's vendor
        # register -- so carry a due date rather than leaving the column empty.
        row[13] = _existing_date(row.get(13)) or inspection_due
        row[26] = note if not remarks else f'{remarks} | {note}'
        row[5] = row.get(5) or 'IT'
        row[15] = row.get(15) or 2
        row[16] = row.get(16) or 1
        row[23] = row.get(23) or 2
        row[24] = row.get(24) or 'Yes'
        row[25] = row.get(25) or 'Yes'
        rows.append(row)
        stats['carried'] += 1

    stats['total'] = _write_rows(worksheet, rows, last_column)
    return stats


def _fill_software(worksheet):
    """F08A -- licensed / non-standard software, from the license register."""
    last_column = 22  # V
    existing = _snapshot_rows(worksheet, 3, last_column)
    licenses = _load_licenses()

    rows = []
    stats = {'total': 0, 'new': 0}

    for entry in licenses:
        key = _key(entry['software_name'])
        prior = existing.get(key)
        if not prior:
            stats['new'] += 1

        rows.append({
            3: _clean(entry['software_name']),
            # The recorded version wins; the RMM-derived one is only a hint,
            # and loose name matching there returns the wrong product's version.
            4: _clean(entry['version']) or _clean(entry['common_version']) or _carry(prior, 4),
            5: _carry(prior, 5, _clean(entry['notes']) or 'Business / engineering tooling'),
            6: _carry(prior, 6),   # FY25
            7: _carry(prior, 7),
            8: _carry(prior, 8),
            9: _license_type(entry['license_type']),
            10: entry['total_licenses'],
            11: _carry(prior, 11, 'No'),
            12: _carry(prior, 12, 2),
            13: _carry(prior, 13, 1),
            14: _carry(prior, 14, 1),
            16: _carry(prior, 16, 2),
            17: _carry(prior, 17, 3),
            20: _carry(prior, 20, f"Vendor: {_clean(entry['vendor'])}" if entry['vendor'] else None),
        })

    stats['total'] = _write_rows(worksheet, rows, last_column)
    return stats


def _load_information_assets():
    return _fetch(
        """
        SELECT asset_name, required_protect_class, critical_classification,
               customer_name, information_category, information_category_fy25,
               asset_manager, owning_department, business_area, purpose,
               media_form, media_form_fy25, stored_on, viewing_authority,
               permitted_scope_of_use, other_requirements,
               confidentiality, integrity, availability,
               threat_class, vulnerability_class, remarks
        FROM isms_information_asset
        WHERE is_active = TRUE
        ORDER BY id
        """
    )


def _fill_information_assets(worksheet):
    """F02B -- generated from Tracker's information-asset register.

    Previously this sheet was carried through untouched because Tracker had no
    information-asset concept. The register now holds all 90, so the sheet is
    generated and the FY26 columns ALAP added fill in as they are completed.
    """
    last_column = 29  # AC
    records = _load_information_assets()
    if not records:
        # Nothing in the register -- leave the sheet exactly as filed.
        rows = sum(1 for row in range(FIRST_DATA_ROW, worksheet.max_row + 1)
                   if _clean(worksheet.cell(row=row, column=3).value))
        _fix_importance_column(worksheet, max(FIRST_DATA_ROW, FIRST_DATA_ROW + rows - 1))
        return {'total': rows, 'from_register': 0, 'incomplete': 0}

    rows, incomplete = [], 0
    for record in records:
        missing = [f for f in (record['critical_classification'],
                               record['information_category'],
                               record['business_area'],
                               record['media_form'],
                               record['permitted_scope_of_use'])
                   if not _clean(f)]
        if missing:
            incomplete += 1
        rows.append({
            3: _clean(record['asset_name']),
            4: _clean(record['required_protect_class']),
            5: _clean(record['critical_classification']),
            6: _clean(record['customer_name']) or None,
            7: _clean(record['information_category']) or None,
            8: _clean(record['information_category_fy25']),      # FY25
            9: _clean(record['asset_manager']),                  # FY25
            10: _clean(record['owning_department']),
            11: _clean(record['business_area']) or None,
            12: _clean(record['purpose']),
            13: _clean(record['media_form']) or None,
            14: _clean(record['media_form_fy25']),               # FY25
            15: _clean(record['stored_on']),                     # FY25
            16: _clean(record['viewing_authority']),             # FY25
            17: _clean(record['permitted_scope_of_use']) or None,
            18: _clean(record['other_requirements']),
            19: record['confidentiality'],
            20: record['integrity'],
            21: record['availability'],
            23: record['threat_class'],
            24: record['vulnerability_class'],
            27: _clean(record['remarks']),
        })

    total = _write_rows(worksheet, rows, last_column)
    return {'total': total, 'from_register': total, 'incomplete': incomplete}


def _fill_worker_access(worksheet):
    """F09-A -- which worker can reach which supporting asset (device issue)."""
    last_column = 14  # N
    employees = {employee['id']: employee for employee in _load_employees()}
    devices = _load_assigned_devices()
    # Column D is resolved against Supplement[Asset Name] by the sheet's
    # XLOOKUP, so an endpoint excluded from F04C would render "Not Found".
    on_supplement = {_key(asset['name']) for asset in _load_assets()}

    rows = []
    for employee_id, bucket in sorted(devices.items()):
        employee = employees.get(employee_id)
        if employee is None:
            continue
        for _tag, asset_name in bucket['pc']:
            if not asset_name or _key(asset_name) not in on_supplement:
                continue
            rows.append({
                1: _name_mail(employee['name'], employee['email']),
                3: 'SupplimentName',
                4: asset_name,
                6: _clean(employee['sam_account_name']) or _clean(employee['email']),
                7: employee['start_date'],
                8: 'Adminstrator' if _clean(employee['department']).lower() == 'it' else 'User',
                11: 'Assigned endpoint (Tracker asset register)',
            })

    # Partner-provided systems staff were instructed to use (ALAP's follow-up
    # request). Column D must match Business_Partner_List[Business Partner Name]
    # exactly -- the sheet's XLOOKUP resolves the partner's Auto_No from it.
    partner_accounts = _load_partner_system_accounts()
    for account in partner_accounts:
        rows.append({
            1: _name_mail(account['employee_name'], account['email']),
            3: 'BP_Name',
            4: _clean(account['vendor_name']),
            6: _clean(account['user_principal_name']),
            7: account['start_date'],
            8: 'Adminstrator' if account['is_admin'] else 'User',
            11: 'Microsoft 365 account (Entra ID, synced to Tracker)',
        })

    stats = {
        'total': _write_rows(worksheet, rows, last_column),
        'endpoints': len(rows) - len(partner_accounts),
        'partner_systems': len(partner_accounts),
    }
    return stats


def _fill_software_users(worksheet):
    """F09-B -- software user list, from license assignments."""
    last_column = 14  # N
    assignments = _load_license_assignments()

    rows = []
    for assignment in assignments:
        rows.append({
            1: _name_mail(assignment['employee_name'], assignment['email']),
            3: 'Software_Name',
            4: _clean(assignment['software_name']),
            6: _clean(assignment['sam_account_name']) or _clean(assignment['email']),
            7: assignment['assigned_date'],
            8: 'User',
            11: 'Named license assignment (Tracker license register)',
        })

    stats = {'total': _write_rows(worksheet, rows, last_column)}
    return stats


# --------------------------------------------------------------------------
# entry points
# --------------------------------------------------------------------------

def build_ledger_workbook(template_path=None):
    """Fill the ledger template from live Tracker data.

    Returns ``(BytesIO, summary_dict)``.
    """
    path = Path(template_path) if template_path else find_template()
    if path is None or not path.exists():
        raise FileNotFoundError(
            f'No ISMS ledger template found. Upload one into {LEDGER_TEMPLATE_DIR}.'
        )

    workbook = load_workbook(path)

    summary = {
        'template': path.name,
        'generated_on': date.today(),
        'sheets': {},
    }
    summary['sheets']['F04C Supporting Assets'] = _fill_supplement(workbook[SHEET_SUPPLEMENT])
    summary['sheets']['F09A Worker List'] = _fill_workers(workbook[SHEET_WORKERS])
    summary['sheets']['F06B Business Partners'] = _fill_partners(workbook[SHEET_PARTNERS])
    summary['sheets']['F08A Software'] = _fill_software(workbook[SHEET_SOFTWARE])
    summary['sheets']['F09-A Worker Access'] = _fill_worker_access(workbook[SHEET_WORKER_ACCESS])
    summary['sheets']['F09-B Software Users'] = _fill_software_users(workbook[SHEET_SOFTWARE_USERS])

    summary['sheets']['F02B Information Assets'] = _fill_information_assets(workbook[SHEET_INFO])

    # F03A (prototypes) is likewise not Tracker-held -- Cirque files no
    # prototype assets -- but carries the same Importance column.
    prototype_sheet = workbook[SHEET_PROTOTYPE]
    prototype_rows = sum(
        1 for row in range(FIRST_DATA_ROW, prototype_sheet.max_row + 1)
        if _clean(prototype_sheet.cell(row=row, column=3).value)
    )
    _fix_importance_column(prototype_sheet, max(FIRST_DATA_ROW, FIRST_DATA_ROW + prototype_rows - 1))
    summary['sheets']['F03A Prototypes (manual)'] = {'total': prototype_rows}

    summary['revisions'] = _apply_validation_revisions(workbook)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output, summary


def ledger_coverage():
    """Row counts Tracker would contribute, for the ledger page (no file work)."""
    training_by_employee, training_by_email, training_by_name = _load_training()
    employees = _load_employees()
    trained = sum(
        1 for employee in employees
        if training_by_employee.get(employee['id'])
        or training_by_email.get(_clean(employee['email']).lower())
        or training_by_name.get(_key(employee['name']))
    )
    assets = _load_assets()
    return {
        'assets': len(assets),
        'assets_excluded_offline': len(_load_assets(stale=True)),
        'online_window_days': ONLINE_WITHIN_DAYS,
        'assets_retired': sum(1 for a in assets if _clean(a['status']) in RETIRED_ASSET_STATUSES),
        'assets_unassigned': sum(1 for a in assets if not a['employee_name']),
        'employees': len(employees),
        'employees_trained': trained,
        'employees_untrained': len(employees) - trained,
        'vendors': len(_load_vendors()),
        'licenses': len(_load_licenses()),
        'license_assignments': len(_load_license_assignments()),
    }


# What ALAP requires as input on each generated sheet, and where it comes from.
# Anything marked "no input required" (FY25 carry-over) or auto-calculated is
# deliberately absent -- a blank there is correct, not a gap.
READINESS_SPEC = {
    SHEET_SUPPLEMENT: (3, [
        (8, 'IT Operations Organization', None),
        (9, 'Purpose of Using the Asset', None),
        (11, 'Types of Assets', None),
        (13, 'Permitted Scope of Use', None),
        (15, 'Installation / Storage Location', None),
        (20, 'Training Requirement', None),
        (31, 'Latest Inspection Date', 'Defender vulnerability scan coverage'),
        (32, 'Inspection Result', 'Defender vulnerability scan coverage'),
    ]),
    SHEET_PARTNERS: (3, [
        (4, 'Contact Department at Partner', 'Vendors'),
        (6, 'Overview of Transaction', 'Vendors'),
        (9, 'xNDA-007 Execution Date', 'Vendors'),
        (10, 'ISMS Announcement Date', 'Vendors'),
        (11, 'Latest Audit Date', 'Vendors'),
        (13, 'Next Audit Date', 'Vendors'),
        (23, 'Required Availability', 'Vendors'),
    ]),
    SHEET_SOFTWARE: (3, [
        (4, 'Software Version', 'Licenses'),
        (5, 'Purpose of Using the Asset', 'Licenses'),
        (9, 'License Type', 'Licenses'),
        (10, 'Number of Licenses', 'Licenses'),
    ]),
    SHEET_WORKERS: (5, [
        (6, 'Email address', 'Employees'),
        (8, 'Relationship with our company', 'Employees'),
        (9, 'AD Account ID', 'Employees'),
        (10, 'Start Date of Employment', 'Employees'),
        (12, 'Last ISMS Training Date', 'Security Training'),
    ]),
    SHEET_INFO: (3, [
        (5, 'Critical Information Classification', 'Information Assets'),
        (7, 'Information Category', 'Information Assets'),
        (11, 'Primary Business Area', 'Information Assets'),
        (13, 'Media form', 'Information Assets'),
        (17, 'Permitted Scope of Use', 'Information Assets'),
    ]),
}


def ledger_readiness(workbook=None):
    """Per-column fill rate on a freshly generated workbook.

    Answers "is this ready to file?" against ALAP's required-input columns,
    rather than the generic gap count the page used to show.
    """
    if workbook is None:
        output, _summary = build_ledger_workbook()
        workbook = load_workbook(output)

    report = []
    for sheet_name, (key_column, columns) in READINESS_SPEC.items():
        worksheet = workbook[sheet_name]
        rows = [r for r in range(FIRST_DATA_ROW, worksheet.max_row + 1)
                if _clean(worksheet.cell(row=r, column=key_column).value)]
        if not rows:
            continue
        entries = []
        for column, label, source in columns:
            filled = sum(1 for r in rows
                         if _clean(worksheet.cell(row=r, column=column).value))
            entries.append({
                'column': get_column_letter(column),
                'label': label,
                'source': source,
                'filled': filled,
                'total': len(rows),
                'complete': filled == len(rows),
            })
        report.append({
            'sheet': sheet_name,
            'rows': len(rows),
            'columns': entries,
            'gaps': sum(1 for e in entries if not e['complete']),
        })
    return report


def export_filename(today=None):
    stamp = (today or date.today()).strftime('%Y%m%d')
    return f'ISMS-Management-Ledgers_CIRQUE_{stamp}.xlsx'
