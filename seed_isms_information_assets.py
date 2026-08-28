"""Seed the F02B information-asset register from the ALAP ledger template.

The 90 information assets already exist in the workbook; this lifts them into
Tracker so they stop living only in a spreadsheet. Whatever FY25 recorded is
carried across, and the FY26 columns ALAP added (critical classification,
information category, business area, media form, permitted scope of use) come
across empty where they are empty — the point is to make the gaps visible and
fillable, not to invent values.

Idempotent: matches on asset name and updates rather than duplicating.
Run: venv/bin/python seed_isms_information_assets.py
"""
import sys

from app import app
from extensions import db
from models import ISMSInformationAsset
from isms_ledger_service import find_template, FIRST_DATA_ROW

from openpyxl import load_workbook

SHEET = 'F02B Asset - Info'

# F02B column index -> model attribute
COLUMNS = {
    3:  'asset_name',
    4:  'required_protect_class',
    5:  'critical_classification',
    6:  'customer_name',
    7:  'information_category',
    8:  'information_category_fy25',
    9:  'asset_manager',
    10: 'owning_department',
    11: 'business_area',
    12: 'purpose',
    13: 'media_form',
    14: 'media_form_fy25',
    15: 'stored_on',
    16: 'viewing_authority',
    17: 'permitted_scope_of_use',
    18: 'other_requirements',
    19: 'confidentiality',
    20: 'integrity',
    21: 'availability',
    23: 'threat_class',
    24: 'vulnerability_class',
    27: 'remarks',
}
INT_FIELDS = {'confidentiality', 'integrity', 'availability',
              'threat_class', 'vulnerability_class'}


def _clean(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def seed():
    path = find_template()
    if path is None:
        print('No ledger template found; nothing to seed.')
        return 1

    worksheet = load_workbook(path, data_only=True)[SHEET]
    created = updated = skipped = 0

    for row in range(FIRST_DATA_ROW, worksheet.max_row + 1):
        name = _clean(worksheet.cell(row=row, column=3).value)
        if not name:
            continue

        values = {}
        for column, attribute in COLUMNS.items():
            value = _clean(worksheet.cell(row=row, column=column).value)
            if value is not None and attribute in INT_FIELDS:
                try:
                    value = int(float(value))
                except (TypeError, ValueError):
                    value = None
            values[attribute] = value

        record = ISMSInformationAsset.query.filter(
            db.func.lower(ISMSInformationAsset.asset_name) == name.lower()).first()
        if record is None:
            record = ISMSInformationAsset(source='f02b-template', created_by='seed')
            db.session.add(record)
            created += 1
        else:
            # Only backfill blanks — never clobber a value someone has entered.
            values = {k: v for k, v in values.items()
                      if v is not None and getattr(record, k, None) in (None, '')}
            if not values:
                skipped += 1
                continue
            updated += 1

        for attribute, value in values.items():
            setattr(record, attribute, value)

    db.session.commit()
    total = ISMSInformationAsset.query.count()
    incomplete = sum(1 for r in ISMSInformationAsset.query.all() if r.missing_ledger_fields)
    print(f'created={created} updated={updated} unchanged={skipped}')
    print(f'register now holds {total} information assets; {incomplete} have missing ALAP fields')
    return 0


if __name__ == '__main__':
    with app.app_context():
        sys.exit(seed())
